# -*- coding: utf-8 -*-
"""Generate mode1 input files for missing monthly balance fill testing."""

from __future__ import annotations

import csv
import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.bank_integration.balances import get_monthly_balances
from src.bank_integration.config import BANK_READ_CONFIG
from src.bank_integration.readers import read_bank_file
from src.bank_integration.scanner import scan_source_files


RAW_DIR = ROOT / "data" / "input" / "raw" / "1"
OUTPUT_DIR = ROOT / "data" / "input" / "fill_missing" / "1"

START_DATE = "20260101"
END_DATE = "20260331"

FILES = {
    "招商银行": {
        "raw": RAW_DIR / "招商银行.xlsx",
        "output": OUTPUT_DIR / "补测A-招商银行.xlsx",
        "expected": [("2026-01-31", 1000.0), ("2026-02-28", 1000.0), ("2026-03-03", 900.0)],
    },
    "中信银行": {
        "raw": RAW_DIR / "中信银行.xlsx",
        "output": OUTPUT_DIR / "补测B-中信银行.xlsx",
        "expected": [("2026-01-31", 1000.0), ("2026-02-28", 1000.0), ("2026-03-10", 1050.0)],
    },
    "中国银行": {
        "raw": RAW_DIR / "中国银行.csv",
        "output": OUTPUT_DIR / "补测C-中国银行.csv",
        "expected": [("2026-01-31", 6474.27), ("2026-02-28", 6474.27), ("2026-03-26", 6075.08)],
    },
}


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    _generate_cmb()
    _generate_citic()
    _generate_boc()
    _validate_outputs()

    print("Generated fill-missing input files:")
    for bank_name, spec in FILES.items():
        print(f"  {bank_name}: {spec['output'].resolve()}")
        print(f"    expected: {spec['expected']}")


def _generate_cmb() -> None:
    bank_name = "招商银行"
    spec = FILES[bank_name]
    shutil.copy2(spec["raw"], spec["output"])

    columns = _read_xlsx_columns(bank_name, spec["raw"])
    row = _blank_row(columns)
    row.update(
        {
            "账号": "755966716810301",
            "账号名称": "补测招商账户",
            "币种": "人民币",
            "交易日": "2026-03-03",
            "交易时间": "09:00:00",
            "起息日": "2026-03-03",
            "交易类型": "补测借方交易",
            "借方金额": "100.00",
            "贷方金额": "",
            "余额": "900.00",
            "摘要": "补测缺失月份倒推",
            "流水号": "FILL-CMB-20260303",
        }
    )

    _write_xlsx_rows(
        spec["output"],
        bank_name,
        [row],
        {
            "查询开始日期": START_DATE,
            "查询结束日期": END_DATE,
            "记录条数": "1",
            "累计借金额": "100.00",
            "累计借总笔数": "1",
            "累计贷金额": "0.00",
            "累计贷总笔数": "0",
            "对账单期初余额": "1000.00",
            "对账单余额": "900.00",
        },
    )


def _generate_citic() -> None:
    bank_name = "中信银行"
    spec = FILES[bank_name]
    shutil.copy2(spec["raw"], spec["output"])

    columns = _read_xlsx_columns(bank_name, spec["raw"])
    row = _blank_row(columns)
    row.update(
        {
            "交易日期": "2026-03-10",
            "交易时间": "10:00:00",
            "对方账号": "6222002600000001",
            "对方账户名称": "补测客户",
            "借方发生额": "",
            "贷方发生额": "50.00",
            "账户余额": "1,050.00",
            "摘要": "补测贷方倒推",
            "退汇标识": "否",
            "币种": "CNY",
            "交易账号": "8110301012300546583",
            "发起方流水号": "FILL-CITIC-20260310",
        }
    )

    _write_xlsx_rows(
        spec["output"],
        bank_name,
        [row],
        {
            "起始日期": "2026-01-01",
            "截止日期": "2026-03-31",
        },
    )


def _generate_boc() -> None:
    bank_name = "中国银行"
    spec = FILES[bank_name]
    cfg = BANK_READ_CONFIG[bank_name]

    with spec["raw"].open("r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))

    header_idx = cfg["header"]
    meta_rows = rows[:header_idx]
    header_row = rows[header_idx]
    for row in meta_rows:
        if row and row[0].startswith("查询时间范围"):
            _ensure_len(row, 2)
            row[1] = f"{START_DATE}-{END_DATE}"
        elif row and row[0].startswith("总笔数"):
            _ensure_len(row, 2)
            row[1] = "2"
        elif row and row[0].startswith("借方发生总笔数"):
            _ensure_len(row, 2)
            row[1] = "1"
        elif row and row[0].startswith("借方发生总额"):
            _ensure_len(row, 2)
            row[1] = "400.00"
        elif row and row[0].startswith("贷方发生总笔数"):
            _ensure_len(row, 2)
            row[1] = "1"
        elif row and row[0].startswith("贷方发生总额"):
            _ensure_len(row, 2)
            row[1] = "0.81"

    columns = [str(c).split("[")[0].strip() for c in header_row]
    data_rows = [
        _csv_values(
            columns,
            {
                "交易类型": "来账",
                "业务类型": "补测",
                "收款人开户行行号": "06261",
                "收款人开户行名": "中国银行南京后宰门支行",
                "收款人账号": "496261064671",
                "收款人名称": "补测中国银行账户",
                "交易日期": "20260320",
                "交易时间": "22:43:58",
                "交易货币": "CNY",
                "交易金额": "+0.81",
                "交易后余额": "6,475.08",
                "起息日期": "20260321",
                "汇率": "1.000000",
                "交易流水号": "FILL-BOC-20260320",
                "记录标识号": "FILL-BOC-20260320-ID",
                "摘要": "补测贷方交易",
            },
        ),
        _csv_values(
            columns,
            {
                "交易类型": "往账",
                "业务类型": "补测",
                "付款人开户行号": "06261",
                "付款人开户行名": "中国银行南京后宰门支行",
                "付款人账号": "496261064671",
                "付款人名称": "补测中国银行账户",
                "收款人开户行行号": "00011",
                "收款人开户行名": "中国银行江苏省分行",
                "收款人账号": "5338",
                "收款人名称": "中国银行江苏省分行收费业务专用账号",
                "交易日期": "20260326",
                "交易时间": "18:29:09",
                "交易货币": "CNY",
                "交易金额": "-400.00",
                "交易后余额": "6,075.08",
                "起息日期": "20260326",
                "汇率": "1.000000",
                "交易流水号": "FILL-BOC-20260326",
                "记录标识号": "FILL-BOC-20260326-ID",
                "摘要": "补测借方交易",
            },
        ),
    ]

    with spec["output"].open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerows(meta_rows)
        writer.writerow(header_row)
        writer.writerows(data_rows)


def _read_xlsx_columns(bank_name: str, path: Path) -> list[str]:
    cfg = BANK_READ_CONFIG[bank_name]
    df = pd.read_excel(path, header=cfg["header"], dtype=str, engine=cfg["engine"])
    return [str(c).strip() for c in df.columns]


def _blank_row(columns: list[str]) -> dict[str, str]:
    return {column: "" for column in columns}


def _write_xlsx_rows(path: Path, bank_name: str, row_dicts: list[dict[str, str]], header_values: dict[str, str]) -> None:
    cfg = BANK_READ_CONFIG[bank_name]
    columns = _read_xlsx_columns(bank_name, path)
    wb = load_workbook(path)
    ws = wb.active

    _set_header_values(ws, header_values)

    data_start_row = cfg["header"] + 2
    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)

    for row_idx, row_dict in enumerate(row_dicts, start=data_start_row):
        for col_idx, column in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx).value = row_dict.get(column, "")

    wb.save(path)


def _set_header_values(ws, values: dict[str, str]) -> None:
    remaining = dict(values)
    for row in ws.iter_rows(min_row=1, max_row=20):
        for idx, cell in enumerate(row):
            label = str(cell.value or "").strip()
            if label in remaining:
                target = row[idx + 1] if idx + 1 < len(row) else None
                if target is not None:
                    target.value = remaining.pop(label)
        if not remaining:
            return


def _csv_values(columns: list[str], values: dict[str, str]) -> list[str]:
    return [values.get(column, "") for column in columns]


def _ensure_len(row: list[str], length: int) -> None:
    while len(row) < length:
        row.append("")


def _validate_outputs() -> None:
    sources = scan_source_files(OUTPUT_DIR)
    if len(sources) != 3:
        raise AssertionError(f"Expected 3 generated sources, found {len(sources)}")

    for item in sources:
        bank_name = item["bank_name"]
        expected = FILES[bank_name]["expected"]
        df = read_bank_file(item["filepath"], bank_name)
        actual = get_monthly_balances(df, bank_name)
        if len(actual) != len(expected):
            raise AssertionError(f"{bank_name}: expected {expected}, got {actual}")
        for (actual_date, actual_balance), (expected_date, expected_balance) in zip(actual, expected):
            if actual_date != expected_date or round(actual_balance, 2) != round(expected_balance, 2):
                raise AssertionError(f"{bank_name}: expected {expected}, got {actual}")


if __name__ == "__main__":
    main()
