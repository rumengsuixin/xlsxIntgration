# -*- coding: utf-8 -*-
"""Generate full-coverage fake bank input files by editing raw-file copies."""

from __future__ import annotations

import csv
import json
import shutil
import subprocess
import sys
import tempfile
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.bank_integration.config import BANK_READ_CONFIG, INPUT_DIR, RAW_INPUT_DIR


TARGETS = [
    ("A", "中信银行", "xlsx"),
    ("B", "招商银行", "xlsx"),
    ("C", "建设银行", "xls"),
    ("D", "浦发银行", "xls"),
    ("E", "工商银行", "xlsx"),
    ("F", "中国银行", "csv"),
    ("G", "农业银行", "xls"),
]

DATE_COL = {
    "工商银行": "交易时间",
    "中信银行": "交易日期",
    "招商银行": "交易日",
    "农业银行": "交易时间",
    "建设银行": "交易时间",
    "浦发银行": "交易日期",
    "中国银行": "交易日期",
}

BALANCE_COL = {
    "工商银行": "余额",
    "中信银行": "账户余额",
    "招商银行": "余额",
    "农业银行": "账户余额",
    "建设银行": "余额",
    "浦发银行": "余额",
    "中国银行": "交易后余额",
}

RAW_FILE = {
    "工商银行": "工商银行.xlsx",
    "中信银行": "中信银行.xlsx",
    "招商银行": "招商银行.xlsx",
    "农业银行": "农业银行.xls",
    "建设银行": "建设银行.xls",
    "浦发银行": "浦发银行.xls",
    "中国银行": "中国银行.csv",
}


def main() -> None:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    _clear_input_dir()

    for company, bank_name, ext in TARGETS:
        rows = _build_rows(bank_name)
        output_path = INPUT_DIR / f"{company}-{bank_name}.{ext}"
        shutil.copy2(RAW_INPUT_DIR / RAW_FILE[bank_name], output_path)

        if ext == "csv":
            _write_csv(output_path, bank_name, rows)
        elif ext == "xlsx":
            _write_xlsx_copy(output_path, bank_name, rows)
        else:
            _write_xls_copy_with_excel(output_path, bank_name, rows)

    print(f"Generated {len(TARGETS)} bank test files in {INPUT_DIR}")


def _clear_input_dir() -> None:
    for path in INPUT_DIR.iterdir():
        if path.name in {".gitkeep", "raw"}:
            continue
        if path.is_file():
            path.unlink()
        elif path.is_dir():
            shutil.rmtree(path)


def _read_sample_columns(bank_name: str) -> list[str]:
    cfg = BANK_READ_CONFIG[bank_name]
    path = RAW_INPUT_DIR / RAW_FILE[bank_name]

    if cfg["is_csv"]:
        df = pd.read_csv(path, header=cfg["header"], dtype=str, encoding="utf-8-sig")
        return [str(c).split("[")[0].strip() for c in df.columns]

    df = pd.read_excel(path, header=cfg["header"], dtype=str, engine=cfg["engine"])
    return [str(c).strip() for c in df.columns]


def _build_rows(bank_name: str) -> list[list[str]]:
    columns = _read_sample_columns(bank_name)
    rows = []

    for month in range(1, 13):
        day = min(20 + month, 28)
        trade_date = date(2026, month, day)
        balance = 10000 + month * 123.45
        amount = 80 + month * 7.5
        row = {column: "" for column in columns}

        _fill_common_fields(row, month, amount, balance)
        row[DATE_COL[bank_name]] = _format_date(bank_name, trade_date, month)
        row[BALANCE_COL[bank_name]] = _format_money(balance)

        if bank_name == "中信银行":
            row["交易时间"] = f"{8 + month % 10:02d}:15:{month:02d}"
            row["贷方发生额"] = _format_money(amount)
            row["摘要"] = f"测试结息{month:02d}"
        elif bank_name == "招商银行":
            row["交易时间"] = f"{9 + month % 8:02d}:20:{month:02d}"
            row["贷方金额"] = _format_money(amount)
            row["交易类型"] = "测试转入"
            row["摘要"] = f"测试交易{month:02d}"
        elif bank_name == "建设银行":
            row["贷方发生额/元(收入)"] = _format_money(amount)
            row["借方发生额/元(支取)"] = "0"
            row["记账日期"] = trade_date.strftime("%Y%m%d")
            row["摘要"] = f"测试收入{month:02d}"
        elif bank_name == "浦发银行":
            row["交易时间"] = f"{10 + month % 8:02d}{month:02d}30"
            row["申请日期"] = trade_date.strftime("%Y%m%d")
            row["贷方金额"] = _format_money(amount)
            row["摘要"] = f"测试入账{month:02d}"
        elif bank_name == "工商银行":
            row["借贷标志"] = "贷"
            row["转入金额"] = _format_money(amount)
            row["摘要"] = f"测试入账{month:02d}"
            row["用途"] = "测试"
        elif bank_name == "中国银行":
            row["交易类型"] = "来账"
            row["业务类型"] = "测试"
            row["交易时间"] = f"{8 + month % 10:02d}:35:{month:02d}"
            row["交易货币"] = "CNY"
            row["交易金额"] = f"+{_format_money(amount)}"
            row["起息日期"] = trade_date.strftime("%Y%m%d")
            row["摘要"] = f"测试来账{month:02d}"
        elif bank_name == "农业银行":
            row["收入金额"] = _format_money(amount)
            row["摘要"] = f"测试收入{month:02d}"

        rows.append([row[column] for column in columns])

    if bank_name == "农业银行":
        rows.append(["总收入笔数", "总收入金额", "总支出笔数", "总支出金额", "", "", "", ""])
        rows.append(["", "", "", "", "", "", "", ""])

    return rows


def _fill_common_fields(row: dict[str, str], month: int, amount: float, balance: float) -> None:
    for key in row:
        if "账号" in key:
            row[key] = f"622200260000{month:04d}"
        elif "户名" in key or "名称" in key or "单位" in key:
            row[key] = f"测试客户{month:02d}"
        elif "流水" in key:
            row[key] = f"TEST2026{month:02d}0001"
        elif key == "币种":
            row[key] = "人民币"
        elif key == "交易货币":
            row[key] = "CNY"
        elif "借方" in key or "支出" in key or "转出" in key:
            row[key] = ""
        elif "贷方" in key or "收入" in key or "转入" in key:
            row[key] = _format_money(amount)
        elif key == "余额" or key.endswith("余额"):
            row[key] = _format_money(balance)


def _format_date(bank_name: str, trade_date: date, month: int) -> str:
    if bank_name in {"建设银行", "浦发银行", "中国银行"}:
        base = trade_date.strftime("%Y%m%d")
    else:
        base = trade_date.isoformat()

    if bank_name in {"工商银行", "农业银行"}:
        return f"{trade_date.isoformat()} {8 + month % 10:02d}:00:{month:02d}"
    if bank_name == "建设银行":
        return f"{base} {8 + month % 10:02d}:10:{month:02d}"
    return base


def _format_money(value: float) -> str:
    return f"{value:,.2f}"


def _write_xlsx_copy(path: Path, bank_name: str, rows: list[list[str]]) -> None:
    cfg = BANK_READ_CONFIG[bank_name]
    wb = load_workbook(path)
    ws = wb.active
    data_start_row = cfg["header"] + 2

    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)
    for row_idx, row in enumerate(rows, start=data_start_row):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    wb.save(path)


def _write_xls_copy_with_excel(path: Path, bank_name: str, rows: list[list[str]]) -> None:
    cfg = BANK_READ_CONFIG[bank_name]
    payload = {
        "path": str(path.resolve()),
        "dataStartRow": cfg["header"] + 2,
        "rows": rows,
    }

    ps_script = r"""
$payloadPath = $args[0]
$payload = Get-Content -LiteralPath $payloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {
    $workbook = $excel.Workbooks.Open($payload.path)
    $worksheet = $workbook.Worksheets.Item(1)
    $dataStartRow = [int]$payload.dataStartRow
    $usedRows = [int]$worksheet.UsedRange.Rows.Count
    if ($usedRows -ge $dataStartRow) {
        $worksheet.Range($worksheet.Rows.Item($dataStartRow), $worksheet.Rows.Item([Math]::Max($usedRows, $dataStartRow + 20))).ClearContents() | Out-Null
    }

    for ($r = 0; $r -lt $payload.rows.Count; $r++) {
        $rowValues = $payload.rows[$r]
        for ($c = 0; $c -lt $rowValues.Count; $c++) {
            $cell = $worksheet.Cells.Item($dataStartRow + $r, $c + 1)
            $cell.NumberFormat = "@"
            $cell.Value2 = [string]$rowValues[$c]
        }
    }

    $workbook.Save()
    $workbook.Close($true)
}
finally {
    $excel.Quit()
}
"""

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp = Path(tmp_dir)
        payload_path = tmp / "payload.json"
        script_path = tmp / "write_xls.ps1"
        payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        script_path.write_text(ps_script, encoding="utf-8")
        subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                str(payload_path),
            ],
            check=True,
        )


def _write_csv(path: Path, bank_name: str, rows: list[list[str]]) -> None:
    cfg = BANK_READ_CONFIG[bank_name]
    columns = _read_sample_columns(bank_name)
    raw_path = RAW_INPUT_DIR / RAW_FILE[bank_name]

    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        with raw_path.open("r", encoding="utf-8-sig", newline="") as raw_fh:
            reader = csv.reader(raw_fh)
            for idx, row in enumerate(reader):
                if idx >= cfg["header"]:
                    break
                csv.writer(fh).writerow(row)

        writer = csv.writer(fh)
        writer.writerow([f"{column}[test]" for column in columns])
        writer.writerows(rows)


if __name__ == "__main__":
    main()
