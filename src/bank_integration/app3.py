"""代号3（游戏订单匹配）应用逻辑。

数据流：
    data/input/3/  →  scan_source_files_3()
                           │
                      read_admin / read_adyen / read_huawei / read_google
                           │
                  build_adyen_lookup / build_huawei_lookup / build_google_lookup
                           │
                      enrich_admin()   ← left-join 三个查找表
                           │
                  data/output/订单匹配结果_{YYYYMMDD}.xlsx

输出新增列（追加在 admin 原始列末尾，共 10 列）：
    原Charge金额  - Google 原始 Charge / Charge refund 的 Amount (Buyer Currency) 绝对值
    平台订单金额  - 各平台匹配到的订单交易金额；Google 为平台币种含税总额
    平台币种      - 对应货币
    是否匹配      - 是 / 否 / 平台多余
    状态          - 成功 / 失败 / 退款
    结算金额      - 扣除平台手续费后实际到账金额
                    Adyen: Payable(SC) [USD/HKD]
                    Google: Merchant Charge - |Merchant fee|；退款: |Merchant refund| - |Merchant fee_refund|
                    华为: 空（平台报表无手续费数据）
    手续费        - 平台收取的手续费
                    Adyen: Markup+Scheme Fees+Interchange [USD/HKD]
                    Google: |Google fee Amount(Merchant Currency)|
                    华为: 空
    国家税费      - Google 专属，= abs(Charge Amount (Buyer Currency)) * 0.2 * Currency Conversion Rate
"""

import logging
import csv
from datetime import date
from pathlib import Path
from typing import List, Optional, Set

import pandas as pd

from .config import OUTPUT_DIR
from .config3 import (
    ADMIN_AMOUNT_COL,
    ADMIN_DATE_COL,
    ADMIN_JOIN_COL,
    ADMIN_PAYMENT_COL,
    ADMIN_REFUND_COL,
    ADMIN_SHEET,
    ADYEN_AMOUNT_COL,
    ADYEN_CURRENCY_COL,
    ADYEN_DATE_COL,
    ADYEN_INTERCHANGE_COL,
    ADYEN_JOIN_COL,
    ADYEN_MARKUP_COL,
    ADYEN_PAYABLE_COL,
    ADYEN_RECORD_TYPE_COL,
    ADYEN_RECORD_TYPE_PRIORITY,
    ADYEN_SCHEME_FEES_COL,
    ADYEN_SETTLEMENT_CURRENCY_COL,
    ADYEN_SHEET,
    ADYEN_SETTLE_SHEET,
    ADYEN_SETTLE_HEADER,
    ADYEN_SETTLE_JOURNAL_COL,
    ADYEN_SETTLE_PAYOUT_TYPE,
    ADYEN_SETTLE_TAX_TYPE,
    ADYEN_SETTLE_DATE_COL,
    ADYEN_SETTLE_CURRENCY_COL,
    ADYEN_SETTLE_AMOUNT_COL,
    COUNTRY_TAX_COL,
    FEE_COL,
    TAX_COL,
    GOOGLE_BUYER_AMOUNT_COL,
    GOOGLE_BUYER_CURRENCY_COL,
    GOOGLE_CHARGE_TYPE,
    GOOGLE_CONVERSION_RATE_COL,
    GOOGLE_DATE_COL,
    GOOGLE_FEE_REFUND_TYPE,
    GOOGLE_FEE_TYPE,
    GOOGLE_JOIN_COL,
    GOOGLE_MERCHANT_AMOUNT_COL,
    GOOGLE_MERCHANT_CURRENCY_COL,
    GOOGLE_REFUND_TYPE,
    GOOGLE_TRANSACTION_TYPE_COL,
    HUAWEI_AMOUNT_COL,
    HUAWEI_CURRENCY_COL,
    HUAWEI_DATE_COL,
    HUAWEI_JOIN_COL,
    HUAWEI_SHEET,
    HUAWEI_SETTLE_AMOUNT_COL,
    HUAWEI_SETTLE_CURRENCY_COL,
    HUAWEI_SETTLE_DATE_COL,
    HUAWEI_SETTLE_RATE_COL,
    HUAWEI_SETTLE_TOTAL_TRX_COL,
    HUAWEI_SETTLE_VAT_COL,
    INPUT_DIR_3,
    MATCH_STATUS_COL,
    ORIGINAL_CHARGE_AMOUNT_COL,
    OUTPUT_APPLE_SHEET_3,
    OUTPUT_DIFF_SHEET_3,
    OUTPUT_FAILED_SHEET_3,
    OUTPUT_FILE_TEMPLATE,
    OUTPUT_SHEET_3,
    OUTPUT_SUMMARY_SHEET_3,
    PLATFORM_AMOUNT_COL,
    PLATFORM_CURRENCY_COL,
    SETTLEMENT_CURRENCY_COL,
    SETTLEMENT_AMOUNT_COL,
    STATUS_COL,
    TRANSACTION_DATE_COL,
)


GOOGLE_CHARGE_NET_INTERNAL_COL = "_google_charge_net_amount"
CSV_ENCODINGS = ("utf-8-sig", "utf-16", "gbk")


def scan_source_files_3(input_dir: Path) -> dict:
    """扫描输入目录，按平台识别文件。

    识别规则（文件名 stem 小写前缀匹配）：
        admin   → admin 文件
        adyen-  → Adyen 文件（含连字符以避免误匹配）
        华为    → 华为文件
        googol- 或 google-  → Google Play 文件
        苹果    → 苹果文件

    返回 {"admin": [Path, ...], ...}，同平台多文件时全部收录，由调用方合并。
    """
    result: dict = {"admin": [], "adyen": [], "adyen_settlement": [], "huawei": [], "huawei_settlement": [], "google": [], "apple": []}

    for f in sorted(input_dir.iterdir()):
        if not f.is_file():
            continue
        if f.name.startswith("~$"):
            continue
        suffix = f.suffix.lower()
        stem = f.stem.lower()
        if suffix not in {".xlsx", ".xls", ".csv"}:
            continue
        if stem.startswith("admin"):
            result["admin"].append(f)
        elif stem.startswith("adyen-") and "settlement" in stem:
            result["adyen_settlement"].append(f)
        elif stem.startswith("adyen-"):
            result["adyen"].append(f)
        elif stem.startswith("华为平台结算"):
            result["huawei_settlement"].append(f)
        elif stem.startswith("华为"):
            result["huawei"].append(f)
        elif stem.startswith("googol-") or stem.startswith("google-"):
            result["google"].append(f)
        elif stem.startswith("苹果"):
            result["apple"].append(f)

    for key, file_list in result.items():
        if len(file_list) > 1:
            logging.info(
                "发现 %d 个 %s 文件，将合并读取：%s",
                len(file_list), key, [f.name for f in file_list],
            )

    return result


def _normalize_columns(columns) -> Set[str]:
    return {str(col).strip() for col in columns}


def _columns_match(
    columns,
    required_columns: List[str],
    any_column_groups: Optional[List[List[str]]] = None,
) -> bool:
    normalized = _normalize_columns(columns)
    required = [str(col).strip() for col in required_columns]
    any_groups = [[str(col).strip() for col in group] for group in (any_column_groups or [])]
    if not all(col in normalized for col in required):
        return False
    return not any_groups or any(all(col in normalized for col in group) for group in any_groups)


def _read_csv_file(filepath: Path, *, header: int = 0, label: str = "CSV") -> pd.DataFrame:
    """按常见编码读取 CSV，全列字符串，不限制数据大小。"""
    last_error = None
    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(filepath, dtype=str, encoding=encoding, header=header)
        except UnicodeError as exc:
            last_error = exc
    raise UnicodeError(f"读取 {label} CSV 文件失败: {filepath.name}; 已尝试 utf-8-sig / utf-16 / gbk") from last_error


def _find_csv_header_row(
    filepath: Path,
    *,
    required_columns: List[str],
    any_column_groups: Optional[List[List[str]]] = None,
    max_scan_rows: int = 50,
) -> Optional[int]:
    last_error = None
    for encoding in CSV_ENCODINGS:
        try:
            with filepath.open("r", encoding=encoding, newline="") as fh:
                for idx, row in enumerate(csv.reader(fh)):
                    if idx >= max_scan_rows:
                        break
                    if _columns_match(row, required_columns, any_column_groups):
                        return idx
            return None
        except UnicodeError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise UnicodeError(f"读取 CSV 文件失败: {filepath.name}; 已尝试 utf-8-sig / utf-16 / gbk") from last_error
    return None


def _read_csv_by_columns(
    filepath: Path,
    *,
    header: int,
    required_columns: List[str],
    label: str,
    any_column_groups: Optional[List[List[str]]] = None,
) -> pd.DataFrame:
    df = _read_csv_file(filepath, header=header, label=label).dropna(how="all").fillna("")
    if _columns_match(df.columns, required_columns, any_column_groups):
        return df

    detected_header = _find_csv_header_row(
        filepath,
        required_columns=required_columns,
        any_column_groups=any_column_groups,
    )
    if detected_header is not None and detected_header != header:
        logging.warning(
            "%s CSV %s configured header=%d not usable; using detected header=%d",
            label,
            filepath.name,
            header,
            detected_header,
        )
        df = _read_csv_file(filepath, header=detected_header, label=label).dropna(how="all").fillna("")
        if _columns_match(df.columns, required_columns, any_column_groups):
            return df

    if not _columns_match(df.columns, required_columns, any_column_groups):
        required_desc = list(required_columns)
        if any_column_groups:
            required_desc = required_desc + [f"one of {any_column_groups}"]
        raise ValueError(
            f"{label} CSV {filepath.name} does not contain required columns {required_desc!r}. "
            f"Actual columns: {list(df.columns)!r}"
        )
    return df


def _select_sheet_by_columns(
    filepath: Path,
    *,
    default_sheet,
    header: int,
    required_columns: List[str],
    label: str,
    any_column_groups: Optional[List[List[str]]] = None,
):
    """Return the first sheet whose header contains the required platform columns."""
    with pd.ExcelFile(filepath) as xls:
        sheet_names = list(xls.sheet_names)
        if isinstance(default_sheet, int):
            default_name = sheet_names[default_sheet] if -len(sheet_names) <= default_sheet < len(sheet_names) else None
        else:
            default_name = default_sheet if default_sheet in sheet_names else None

        ordered_sheets = []
        if default_name is not None:
            ordered_sheets.append(default_name)
        ordered_sheets.extend(sheet for sheet in sheet_names if sheet != default_name)

        for sheet_name in ordered_sheets:
            try:
                preview = pd.read_excel(xls, sheet_name=sheet_name, header=header, nrows=0, dtype=str)
            except Exception:
                logging.warning("Failed to inspect %s sheet header: %s [%s]", label, filepath.name, sheet_name)
                continue

            if _columns_match(preview.columns, required_columns, any_column_groups):
                if sheet_name != default_name:
                    expected = default_sheet if default_name is None else default_name
                    logging.warning(
                        "%s sheet %r not usable in %s; using sheet %r because it contains required columns %s. "
                        "Available sheets: %s",
                        label,
                        expected,
                        filepath.name,
                        sheet_name,
                        required,
                        sheet_names,
                    )
                return sheet_name

    required_desc = list(required_columns)
    if any_column_groups:
        required_desc = required_desc + [f"one of {any_column_groups}"]
    raise ValueError(
        f"{label} workbook {filepath.name} does not contain usable sheet {default_sheet!r} "
        f"or any sheet with required columns {required_desc!r}. Available sheets: {sheet_names}"
    )


def _select_admin_sheet(filepath: Path) -> str:
    """Return the admin worksheet name, falling back to the sheet with the join column."""
    with pd.ExcelFile(filepath) as xls:
        sheet_names = list(xls.sheet_names)
        if ADMIN_SHEET in sheet_names:
            try:
                header = pd.read_excel(xls, sheet_name=ADMIN_SHEET, nrows=0, dtype=str)
                if ADMIN_JOIN_COL in _normalize_columns(header.columns):
                    return ADMIN_SHEET
            except Exception:
                logging.warning("Failed to inspect admin sheet header: %s [%s]", filepath.name, ADMIN_SHEET)

        for sheet_name in sheet_names:
            if sheet_name == ADMIN_SHEET:
                continue
            try:
                header = pd.read_excel(xls, sheet_name=sheet_name, nrows=0, dtype=str)
            except Exception:
                logging.warning("Failed to inspect admin sheet header: %s [%s]", filepath.name, sheet_name)
                continue

            columns = _normalize_columns(header.columns)
            if ADMIN_JOIN_COL in columns:
                logging.warning(
                    "Admin sheet %r not usable in %s; using sheet %r because it contains %r. "
                    "Available sheets: %s",
                    ADMIN_SHEET,
                    filepath.name,
                    sheet_name,
                    ADMIN_JOIN_COL,
                    sheet_names,
                )
                return sheet_name

    raise ValueError(
        f"Admin workbook {filepath.name} does not contain sheet {ADMIN_SHEET!r} "
        f"or any sheet with column {ADMIN_JOIN_COL!r}. Available sheets: {sheet_names}"
    )


def read_admin(filepath: Path) -> pd.DataFrame:
    """读取 admin 订单表（工作表：汇总），全列字符串。"""
    if filepath.suffix.lower() == ".csv":
        return _read_csv_by_columns(
            filepath,
            header=0,
            required_columns=[ADMIN_JOIN_COL],
            label="Admin",
        )
    sheet_name = _select_admin_sheet(filepath)
    df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
    return df.dropna(how="all").fillna("")


def read_adyen(filepath: Path) -> pd.DataFrame:
    """读取 Adyen 报告（工作表：Data），全列字符串。"""
    if filepath.suffix.lower() == ".csv":
        return _read_csv_by_columns(
            filepath,
            header=0,
            required_columns=[ADYEN_JOIN_COL, ADYEN_RECORD_TYPE_COL],
            label="Adyen",
        )
    sheet_name = _select_sheet_by_columns(
        filepath,
        default_sheet=ADYEN_SHEET,
        header=0,
        required_columns=[ADYEN_JOIN_COL, ADYEN_RECORD_TYPE_COL],
        label="Adyen",
    )
    return pd.read_excel(filepath, sheet_name=sheet_name, dtype=str).fillna("")


def read_adyen_settlement(filepath: Path) -> pd.DataFrame:
    """读取 ADYEN 结算报告（工作表 Data，header=6），全列字符串。"""
    required_columns = [
        ADYEN_SETTLE_JOURNAL_COL,
        ADYEN_SETTLE_DATE_COL,
        ADYEN_SETTLE_CURRENCY_COL,
        ADYEN_SETTLE_AMOUNT_COL,
    ]
    if filepath.suffix.lower() == ".csv":
        return _read_csv_by_columns(
            filepath,
            header=ADYEN_SETTLE_HEADER,
            required_columns=required_columns,
            label="ADYEN settlement",
        )
    sheet_name = _select_sheet_by_columns(
        filepath,
        default_sheet=ADYEN_SETTLE_SHEET,
        header=ADYEN_SETTLE_HEADER,
        required_columns=required_columns,
        label="ADYEN settlement",
    )
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=ADYEN_SETTLE_HEADER, dtype=str)
    return df.dropna(how="all").fillna("")


def read_huawei(filepath: Path) -> pd.DataFrame:
    """读取华为平台文件（工作表：Sheet0），全列字符串。"""
    if filepath.suffix.lower() == ".csv":
        return _read_csv_by_columns(
            filepath,
            header=0,
            required_columns=[HUAWEI_JOIN_COL, HUAWEI_AMOUNT_COL, HUAWEI_CURRENCY_COL],
            label="Huawei",
        )
    sheet_name = _select_sheet_by_columns(
        filepath,
        default_sheet=HUAWEI_SHEET,
        header=0,
        required_columns=[HUAWEI_JOIN_COL, HUAWEI_AMOUNT_COL, HUAWEI_CURRENCY_COL],
        label="Huawei",
    )
    df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
    return df.dropna(how="all").fillna("")


def read_huawei_settlement(filepath: Path) -> pd.DataFrame:
    """读取华为月度结算文件。

    文件格式：第1行=中文列名，第2行=英文列名（header=1），第3行起=数据。
    """
    required_columns = [
        HUAWEI_SETTLE_DATE_COL,
        HUAWEI_SETTLE_AMOUNT_COL,
        HUAWEI_SETTLE_CURRENCY_COL,
    ]
    if filepath.suffix.lower() == ".csv":
        return _read_csv_by_columns(
            filepath,
            header=1,
            required_columns=required_columns,
            label="Huawei settlement",
        )
    sheet_name = _select_sheet_by_columns(
        filepath,
        default_sheet=0,
        header=1,
        required_columns=required_columns,
        label="Huawei settlement",
    )
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=1, dtype=str)
    return df.dropna(how="all").fillna("")


def read_google(filepath: Path) -> pd.DataFrame:
    """读取 Google Play 报告（第一个工作表），返回全部行。"""
    if filepath.suffix.lower() == ".csv":
        df = _read_csv_by_columns(
            filepath,
            header=0,
            required_columns=[GOOGLE_JOIN_COL, GOOGLE_TRANSACTION_TYPE_COL, GOOGLE_BUYER_AMOUNT_COL],
            label="Google Play",
        )
    else:
        sheet_name = _select_sheet_by_columns(
            filepath,
            default_sheet=0,
            header=0,
            required_columns=[GOOGLE_JOIN_COL, GOOGLE_TRANSACTION_TYPE_COL, GOOGLE_BUYER_AMOUNT_COL],
            label="Google Play",
        )
        df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
    return df.fillna("")


def read_apple(filepath: Path) -> pd.DataFrame:
    """读取 App Store Connect 报表（header 在第 4 行，即 header=3），全列字符串。"""
    required_columns = ["Extended Partner Share"]
    any_column_groups = [
        ["Settlement Date"],
        ["Transaction Date"],
        ["Begin Date"],
        ["End Date"],
        ["Report Date"],
        ["Date"],
        ["日期"],
        ["Quantity", "Customer Price"],
    ]
    if filepath.suffix.lower() == ".csv":
        return _read_csv_by_columns(
            filepath,
            header=3,
            required_columns=required_columns,
            any_column_groups=any_column_groups,
            label="Apple",
        )
    sheet_name = _select_sheet_by_columns(
        filepath,
        default_sheet=0,
        header=3,
        required_columns=required_columns,
        any_column_groups=any_column_groups,
        label="Apple",
    )
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=3, dtype=str)
    return df.dropna(how="all").fillna("")


def _dedup_lookup(df: pd.DataFrame, key_col: str, label: str) -> pd.DataFrame:
    """对 lookup 表按 key_col 去重，空 key 行过滤掉，发现重复时打 warning。"""
    df = df[df[key_col].str.strip() != ""].copy()
    before = len(df)
    df = df.drop_duplicates(subset=[key_col], keep="first")
    dupes = before - len(df)
    if dupes > 0:
        logging.warning("【%s】去重时发现 %d 条重复流水号，已保留首行", label, dupes)
    return df


def build_adyen_lookup(df: pd.DataFrame) -> pd.DataFrame:
    """构建以 Psp Reference 为索引的 Adyen 查找表，只保留 SentForSettle 行。"""
    df2 = df[df[ADYEN_JOIN_COL].str.strip() != ""].copy()
    record_types = df2[ADYEN_RECORD_TYPE_COL].str.strip()

    settle_mask = record_types.isin(ADYEN_RECORD_TYPE_PRIORITY)
    filtered_count = len(df2) - int(settle_mask.sum())
    if filtered_count > 0:
        logging.info("【Adyen】过滤 %d 条非 SentForSettle 记录，不计入匹配", filtered_count)
    df2 = df2[settle_mask].copy()

    result = df2.drop_duplicates(subset=[ADYEN_JOIN_COL], keep="first")
    dupes = len(df2) - len(result)
    if dupes > 0:
        logging.warning("【Adyen】同一 PSP Reference 存在多行 SentForSettle，去重 %d 行", dupes)

    keep_cols = [c for c in [
        ADYEN_JOIN_COL, ADYEN_AMOUNT_COL, ADYEN_CURRENCY_COL,
        ADYEN_SETTLEMENT_CURRENCY_COL,
        ADYEN_PAYABLE_COL, ADYEN_MARKUP_COL, ADYEN_SCHEME_FEES_COL, ADYEN_INTERCHANGE_COL,
        ADYEN_DATE_COL,
    ] if c in result.columns]
    return result[keep_cols].fillna("").set_index(ADYEN_JOIN_COL)


def build_huawei_lookup(df: pd.DataFrame) -> pd.DataFrame:
    """构建以 华为订单号 为索引的华为查找表。

    退款行在华为数据中为独立新行（华为订单号不同），正常不存在重复键；
    若出现重复则保留首行（原始成功记录在前）。
    """
    keep_cols = [c for c in [HUAWEI_JOIN_COL, HUAWEI_AMOUNT_COL, HUAWEI_CURRENCY_COL, HUAWEI_DATE_COL]
                 if c in df.columns]
    result = _dedup_lookup(df[keep_cols], HUAWEI_JOIN_COL, "华为")
    return result.fillna("").set_index(HUAWEI_JOIN_COL)


def build_adyen_settlement_monthly(df: pd.DataFrame) -> pd.DataFrame:
    """从 ADYEN 结算文件提取月度实际到账和税金。

    实际到账(_adyen_payout) = MerchantPayout 行 Net Debit (NC) 合计
    税金(_adyen_tax)        = InvoiceDeduction 行 Net Debit (NC) 合计
    """
    df = df.copy()
    df["_month"] = pd.to_datetime(
        df[ADYEN_SETTLE_DATE_COL].astype(str).str.strip(), errors="coerce"
    ).dt.strftime("%Y-%m")
    df["_amount"] = pd.to_numeric(
        df[ADYEN_SETTLE_AMOUNT_COL].astype(str).str.replace(",", "", regex=False),
        errors="coerce",
    ).fillna(0.0)
    df["_journal"]  = df[ADYEN_SETTLE_JOURNAL_COL].astype(str).str.strip().str.lower()
    df["_currency"] = df[ADYEN_SETTLE_CURRENCY_COL].astype(str).str.strip()

    gkeys = ["_month", "_currency"]
    payout = (
        df[df["_journal"] == ADYEN_SETTLE_PAYOUT_TYPE.lower()]
        .groupby(gkeys)["_amount"].sum().rename("_adyen_payout").reset_index()
    )
    tax = (
        df[df["_journal"] == ADYEN_SETTLE_TAX_TYPE.lower()]
        .groupby(gkeys)["_amount"].sum().rename("_adyen_tax").reset_index()
    )
    result = payout.merge(tax, on=gkeys, how="outer").fillna(0.0)
    return result.rename(columns={"_month": TRANSACTION_DATE_COL,
                                   "_currency": SETTLEMENT_CURRENCY_COL})


def build_huawei_fee_summary(df: pd.DataFrame) -> pd.DataFrame:
    """从华为结算文件按(月份, 结算货币)计算手续费与结算总额。

    手续费 = 结算金额(结算货币) - ((总交易额(交易货币) - 销项税额(交易货币)) * 汇率)
    End Date 列为 YYYYMM 格式，转换为 YYYY-MM。
    """
    summary_cols = [
        TRANSACTION_DATE_COL, ADMIN_PAYMENT_COL, SETTLEMENT_CURRENCY_COL,
        "成功笔数", "成功金额", "退款笔数", "退款金额", "净交易金额", "手续费", TAX_COL,
    ]
    if df.empty:
        return pd.DataFrame(columns=summary_cols)

    def _col_to_float(col_name: str) -> "pd.Series":
        return pd.to_numeric(
            df[col_name].astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        ).fillna(0.0)

    missing = [c for c in [
        HUAWEI_SETTLE_DATE_COL, HUAWEI_SETTLE_AMOUNT_COL, HUAWEI_SETTLE_CURRENCY_COL,
        HUAWEI_SETTLE_TOTAL_TRX_COL, HUAWEI_SETTLE_VAT_COL, HUAWEI_SETTLE_RATE_COL,
    ] if c not in df.columns]
    if missing:
        logging.warning("华为结算文件缺少列: %s，跳过手续费计算", missing)
        return pd.DataFrame(columns=summary_cols)

    work = df.copy()
    settle_amt = _col_to_float(HUAWEI_SETTLE_AMOUNT_COL)
    total_trx  = _col_to_float(HUAWEI_SETTLE_TOTAL_TRX_COL)
    vat        = _col_to_float(HUAWEI_SETTLE_VAT_COL)
    rate       = _col_to_float(HUAWEI_SETTLE_RATE_COL)

    work["_fee"]        = settle_amt - (total_trx - vat) * rate
    work["_settle_amt"] = settle_amt
    work["_settle_ccy"] = work[HUAWEI_SETTLE_CURRENCY_COL].astype(str).str.strip()

    def _to_month(val: str) -> str:
        v = val.strip()
        if len(v) == 6 and v.isdigit():
            return f"{v[:4]}-{v[4:6]}"
        return v[:7]

    work["_month"] = work[HUAWEI_SETTLE_DATE_COL].astype(str).apply(_to_month)

    grouped = (
        work.groupby(["_month", "_settle_ccy"], as_index=False)
        .agg(成功金额=("_settle_amt", "sum"), 手续费=("_fee", "sum"))
    )
    grouped["成功金额"] = grouped["成功金额"].round(2)
    grouped["手续费"]   = grouped["手续费"].round(2)

    result = pd.DataFrame({
        TRANSACTION_DATE_COL:    grouped["_month"],
        ADMIN_PAYMENT_COL:       "华为支付",
        SETTLEMENT_CURRENCY_COL: grouped["_settle_ccy"],
        "成功笔数":  0,
        "成功金额":  grouped["成功金额"],
        "退款笔数":  0,
        "退款金额":  0.0,
        "净交易金额": grouped["成功金额"],
        "手续费":    grouped["手续费"],
        TAX_COL:     0.0,
    })
    return result[summary_cols]


def _format_date(val) -> str:
    try:
        if pd.isna(val):
            return ""
        s = str(val).strip()
        if not s:
            return ""
        parsed = pd.to_datetime(s, errors="coerce")
        if not pd.isna(parsed):
            return parsed.strftime("%Y-%m-%d")
        if len(s) >= 10 and s[4:5] == "-" and s[7:8] == "-":
            return s[:10]
        return ""
    except Exception:
        return ""


def _to_float(val) -> Optional[float]:
    try:
        return float(str(val).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None


def _google_charge_net_amount(merchant_charge, merchant_fee) -> Optional[float]:
    """Google 原始 Charge 扣除 Google fee 后的商户币种净额。"""
    mc = _to_float(merchant_charge)
    if mc is None:
        return None
    mf = _to_float(merchant_fee)
    return round((mc or 0.0) - abs(mf or 0.0), 2)


def build_google_lookup(df: pd.DataFrame) -> pd.DataFrame:
    """构建以 Description 为索引的 Google Play 查找表，合并同一流水号的 4 种行类型。

    每笔订单在 Google 数据中最多出现 4 行：
      Charge          → 原始收款（正数）
      Google fee      → 平台手续费（负数）
      Charge refund   → 退款（负数）
      Google fee refund → 手续费退回（正数）

    对应 admin.是否退款：
      正常   → 平台金额 = |Charge| * 1.2；国家税费 = |Charge| * 0.2 * 汇率（Merchant Currency）
      已退款 → 平台金额 = |Charge refund| * 1.2；国家税费 = |Charge refund| * 0.2 * 汇率（Merchant Currency）
    """
    type_col = GOOGLE_TRANSACTION_TYPE_COL
    join = GOOGLE_JOIN_COL
    amt = GOOGLE_BUYER_AMOUNT_COL
    merchant_amt = GOOGLE_MERCHANT_AMOUNT_COL
    ccy = GOOGLE_BUYER_CURRENCY_COL
    merchant_ccy = GOOGLE_MERCHANT_CURRENCY_COL
    conversion_rate = GOOGLE_CONVERSION_RATE_COL

    def _pick(type_val, out_amt_col, include_merchant_amt=False, merchant_amt_out_col="merchant_amt",
              include_ccy=False, ccy_out_col="ccy",
              include_merchant_ccy=False, merchant_ccy_out_col="merchant_ccy",
              include_date=False, date_out_col="transaction_date",
              include_conversion_rate=False, conversion_rate_out_col="conversion_rate"):
        sub = df[df[type_col].str.strip() == type_val].copy()
        sub = sub[sub[join].str.strip() != ""]
        cols = (
            [join, amt]
            + ([merchant_amt] if include_merchant_amt and merchant_amt in sub.columns else [])
            + ([ccy] if include_ccy else [])
            + ([merchant_ccy] if include_merchant_ccy and merchant_ccy in sub.columns else [])
            + ([GOOGLE_DATE_COL] if include_date and GOOGLE_DATE_COL in sub.columns else [])
            + ([conversion_rate] if include_conversion_rate and conversion_rate in sub.columns else [])
        )
        sub = sub[[c for c in cols if c in sub.columns]]
        before = len(sub)
        sub = sub.drop_duplicates(subset=[join], keep="first")
        if len(sub) < before:
            logging.warning("【Google-%s】发现 %d 条重复流水号，已保留首行", type_val, before - len(sub))
        rename = {amt: out_amt_col}
        if include_merchant_amt and merchant_amt in sub.columns:
            rename[merchant_amt] = merchant_amt_out_col
        if include_ccy:
            rename[ccy] = ccy_out_col
        if include_merchant_ccy and merchant_ccy in sub.columns:
            rename[merchant_ccy] = merchant_ccy_out_col
        if include_date and GOOGLE_DATE_COL in sub.columns:
            rename[GOOGLE_DATE_COL] = date_out_col
        if include_conversion_rate and conversion_rate in sub.columns:
            rename[conversion_rate] = conversion_rate_out_col
        return sub.rename(columns=rename)

    charge = _pick(
        GOOGLE_CHARGE_TYPE,
        "charge_amt",
        include_merchant_amt=True,
        merchant_amt_out_col="merchant_charge_amt",
        include_ccy=True,
        include_merchant_ccy=True,
        include_date=True,
        include_conversion_rate=True,
    )
    fee = _pick(
        GOOGLE_FEE_TYPE,
        "fee_amt",
        include_merchant_amt=True,
        merchant_amt_out_col="merchant_fee_amt",
    )
    refund = _pick(GOOGLE_REFUND_TYPE, "refund_amt",
                   include_merchant_amt=True, merchant_amt_out_col="merchant_refund_amt",
                   include_ccy=True, ccy_out_col="refund_ccy",
                   include_merchant_ccy=True, merchant_ccy_out_col="refund_merchant_ccy",
                   include_date=True, date_out_col="refund_date",
                   include_conversion_rate=True, conversion_rate_out_col="refund_conversion_rate")
    fee_refund = _pick(
        GOOGLE_FEE_REFUND_TYPE,
        "fee_refund_amt",
        include_merchant_amt=True,
        merchant_amt_out_col="merchant_fee_refund_amt",
    )

    result = charge
    for part in (fee, refund, fee_refund):
        if not part.empty:
            result = result.merge(part, on=join, how="outer")

    # 退款订单（无原始 Charge 行）用 refund_date / refund_ccy 补充缺失字段
    if "refund_date" in result.columns:
        if "transaction_date" not in result.columns:
            result["transaction_date"] = None
        mask = result["transaction_date"].isna() | (result["transaction_date"] == "")
        result.loc[mask, "transaction_date"] = result.loc[mask, "refund_date"]
        result = result.drop(columns=["refund_date"])

    if "refund_ccy" in result.columns:
        if "ccy" not in result.columns:
            result["ccy"] = None
        mask = result["ccy"].isna() | (result["ccy"] == "")
        result.loc[mask, "ccy"] = result.loc[mask, "refund_ccy"]
        result = result.drop(columns=["refund_ccy"])

    if "refund_merchant_ccy" in result.columns:
        if "merchant_ccy" not in result.columns:
            result["merchant_ccy"] = None
        mask = result["merchant_ccy"].isna() | (result["merchant_ccy"] == "")
        result.loc[mask, "merchant_ccy"] = result.loc[mask, "refund_merchant_ccy"]
        result = result.drop(columns=["refund_merchant_ccy"])

    if "refund_conversion_rate" in result.columns:
        if "conversion_rate" not in result.columns:
            result["conversion_rate"] = None
        mask = result["conversion_rate"].isna() | (result["conversion_rate"] == "")
        result.loc[mask, "conversion_rate"] = result.loc[mask, "refund_conversion_rate"]
        result = result.drop(columns=["refund_conversion_rate"])

    return result.fillna("").drop_duplicates(subset=[join]).set_index(join)


def _build_platform_only_rows(
    result_cols: list,
    admin_keys: set,
    adyen_lk: Optional[pd.DataFrame],
    huawei_lk: Optional[pd.DataFrame],
    google_lk: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """构建平台有、admin 无对应记录的多余行（回调丢失/网络延迟场景）。

    所有 admin 原始列留空，填充可从平台数据计算的新增列，标记为平台多余。
    """
    extra = []

    # ── Adyen 多余行 ────────────────────────────────────────
    if adyen_lk is not None:
        for key in adyen_lk.index:
            if key in admin_keys:
                continue
            row: dict = {c: "" for c in result_cols}
            row[ADMIN_JOIN_COL]     = key
            row[ADMIN_PAYMENT_COL]  = "Adyen"
            row[PLATFORM_AMOUNT_COL]   = str(adyen_lk.at[key, ADYEN_AMOUNT_COL]).strip()
            row[PLATFORM_CURRENCY_COL] = str(adyen_lk.at[key, ADYEN_CURRENCY_COL]).strip()
            row[SETTLEMENT_CURRENCY_COL] = str(adyen_lk.at[key, ADYEN_SETTLEMENT_CURRENCY_COL]).strip()
            row[MATCH_STATUS_COL]      = "平台多余"
            row[STATUS_COL]            = "成功"
            row[SETTLEMENT_AMOUNT_COL] = str(adyen_lk.at[key, ADYEN_PAYABLE_COL]).strip()
            row[FEE_COL] = "0"
            if ADYEN_DATE_COL in adyen_lk.columns:
                row[TRANSACTION_DATE_COL] = _format_date(adyen_lk.at[key, ADYEN_DATE_COL])
            extra.append(row)

    # ── 华为多余行 ────────────────────────────────────────
    if huawei_lk is not None:
        for key in huawei_lk.index:
            if key in admin_keys:
                continue
            row = {c: "" for c in result_cols}
            row[ADMIN_JOIN_COL]     = key
            row[ADMIN_PAYMENT_COL]  = "华为支付"
            amt = str(huawei_lk.at[key, HUAWEI_AMOUNT_COL]).strip()
            row[PLATFORM_AMOUNT_COL]   = amt
            row[PLATFORM_CURRENCY_COL] = str(huawei_lk.at[key, HUAWEI_CURRENCY_COL]).strip()
            row[SETTLEMENT_CURRENCY_COL] = row[PLATFORM_CURRENCY_COL]
            row[MATCH_STATUS_COL] = "平台多余"
            row[SETTLEMENT_AMOUNT_COL] = amt  # 暂无手续费数据
            row[STATUS_COL]            = "成功"
            if HUAWEI_DATE_COL in huawei_lk.columns:
                row[TRANSACTION_DATE_COL] = _format_date(huawei_lk.at[key, HUAWEI_DATE_COL])
            extra.append(row)

    # ── Google 多余行 ─────────────────────────────────────
    if google_lk is not None:
        for key in google_lk.index:
            if key in admin_keys:
                continue
            row = {c: "" for c in result_cols}
            row[ADMIN_JOIN_COL]    = key
            row[ADMIN_PAYMENT_COL] = "Google支付"
            r  = _to_float(google_lk.at[key, "refund_amt"])
            c_ = _to_float(google_lk.at[key, "charge_amt"])
            mr  = _to_float(google_lk.at[key, "merchant_refund_amt"])
            mc  = _to_float(google_lk.at[key, "merchant_charge_amt"])
            mf  = _to_float(google_lk.at[key, "merchant_fee_amt"])
            mfr = _to_float(google_lk.at[key, "merchant_fee_refund_amt"])
            rate = _to_float(google_lk.at[key, "conversion_rate"]) if "conversion_rate" in google_lk.columns else None
            row[PLATFORM_CURRENCY_COL] = str(google_lk.at[key, "ccy"]).strip()
            row[SETTLEMENT_CURRENCY_COL] = str(google_lk.at[key, "merchant_ccy"]).strip()
            row[MATCH_STATUS_COL] = "平台多余"
            charge_net = _google_charge_net_amount(mc, mf)
            if GOOGLE_CHARGE_NET_INTERNAL_COL in row and charge_net is not None:
                row[GOOGLE_CHARGE_NET_INTERNAL_COL] = str(charge_net)
            if r is not None:  # 有退款记录 → 退款
                refund_amount = abs(r or 0.0)
                row[ORIGINAL_CHARGE_AMOUNT_COL] = str(round(refund_amount, 2))
                row[PLATFORM_AMOUNT_COL] = str(round(refund_amount * 1.2, 2))
                if rate is not None:
                    row[COUNTRY_TAX_COL] = str(round(refund_amount * 0.2 * rate, 2))
                if mr is not None:
                    row[SETTLEMENT_AMOUNT_COL] = str(round(abs(mr or 0.0) - abs(mfr or 0.0), 2))
                if mfr is not None:
                    row[FEE_COL] = str(round(abs(mfr), 2))
                row[STATUS_COL] = "退款"
            else:              # 无退款记录 → 成功
                if c_ is not None:
                    charge_amount = abs(c_ or 0.0)
                    row[ORIGINAL_CHARGE_AMOUNT_COL] = str(round(charge_amount, 2))
                    row[PLATFORM_AMOUNT_COL] = str(round(charge_amount * 1.2, 2))
                    if rate is not None:
                        row[COUNTRY_TAX_COL] = str(round(charge_amount * 0.2 * rate, 2))
                if mc is not None:
                    row[SETTLEMENT_AMOUNT_COL] = str(round((mc or 0.0) - abs(mf or 0.0), 2))
                if mf is not None:
                    row[FEE_COL] = str(round(abs(mf), 2))
                row[STATUS_COL] = "成功"
            if "transaction_date" in google_lk.columns:
                row[TRANSACTION_DATE_COL] = _format_date(google_lk.at[key, "transaction_date"])
            extra.append(row)

    if not extra:
        return pd.DataFrame(columns=result_cols)
    return pd.DataFrame(extra, columns=result_cols)


def enrich_admin(
    admin_df: pd.DataFrame,
    adyen_lk: Optional[pd.DataFrame],
    huawei_lk: Optional[pd.DataFrame],
    google_lk: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """以 admin 为主表，通过流水号与三个平台查找表 left-join，追加平台匹配列。"""
    result = admin_df.copy()
    admin_col_count = len(admin_df.columns)

    expected_rows = len(admin_df)

    def _safe_merge(base, lookup, prefix, platform_name):
        merged = base.merge(
            lookup.add_prefix(prefix),
            left_on=ADMIN_JOIN_COL, right_index=True, how="left",
        )
        if len(merged) != expected_rows:
            raise ValueError(
                f"【{platform_name}】merge 后行数从 {expected_rows} 变为 {len(merged)}，"
                f"平台数据仍存在重复流水号，请检查 build_{platform_name.lower()}_lookup 去重逻辑"
            )
        for c in lookup.columns:
            merged[f"{prefix}{c}"] = merged[f"{prefix}{c}"].fillna("")
        return merged

    # ── Adyen join ───────────────────────────────────────────
    adyen_avail = adyen_lk is not None and not adyen_lk.empty
    if adyen_avail:
        result = _safe_merge(result, adyen_lk, "_a_", "Adyen")

    # ── 华为 join ─────────────────────────────────────────────
    huawei_avail = huawei_lk is not None and not huawei_lk.empty
    if huawei_avail:
        result = _safe_merge(result, huawei_lk, "_h_", "华为")

    # ── Google join ───────────────────────────────────────────
    google_avail = google_lk is not None and not google_lk.empty
    if google_avail:
        result = _safe_merge(result, google_lk, "_g_", "Google")

    # ── 构建新增列 ────────────────────────────────────────────
    original_charge_amt_list = []
    platform_amt_list     = []
    platform_ccy_list     = []
    settlement_ccy_list   = []
    match_status_list     = []
    status_list           = []
    settlement_list       = []
    fee_list              = []
    country_tax_list      = []
    transaction_date_list = []
    google_charge_net_list = []

    for _, row in result.iterrows():
        src = str(row.get(ADMIN_PAYMENT_COL, "")).strip()
        refund_flag = str(row.get(ADMIN_REFUND_COL, "")).strip()

        settle_amt       = ""
        settle_ccy       = ""
        fee_amt          = ""
        country_tax      = ""
        transaction_date = ""
        original_charge_amt = ""
        google_charge_net = ""
        match_tag        = None  # None 表示由 matched 决定；非 None 时直接用此值

        if src == "Adyen" and adyen_avail:
            amt = str(row.get(f"_a_{ADYEN_AMOUNT_COL}", "")).strip()
            ccy = str(row.get(f"_a_{ADYEN_CURRENCY_COL}", "")).strip()
            settle_ccy = str(row.get(f"_a_{ADYEN_SETTLEMENT_CURRENCY_COL}", "")).strip()
            matched = amt != ""
            # 结算金额：Payable (SC)
            settle_amt = str(row.get(f"_a_{ADYEN_PAYABLE_COL}", "")).strip()
            fee_amt = "0"
            transaction_date = _format_date(row.get(ADMIN_DATE_COL, ""))

        elif src == "华为支付" and huawei_avail:
            amt = str(row.get(f"_h_{HUAWEI_AMOUNT_COL}", "")).strip()
            ccy = str(row.get(f"_h_{HUAWEI_CURRENCY_COL}", "")).strip()
            settle_ccy = ccy
            matched = amt != ""
            settle_amt = amt  # 暂无手续费数据，结算金额 = 平台订单金额
            transaction_date = _format_date(row.get(ADMIN_DATE_COL, ""))

        elif "Google" in src and google_avail:
            ccy = str(row.get("_g_ccy", "")).strip()
            settle_ccy = str(row.get("_g_merchant_ccy", "")).strip()
            rate = _to_float(row.get("_g_conversion_rate", ""))
            transaction_date = _format_date(row.get("_g_transaction_date", "")) or _format_date(row.get(ADMIN_DATE_COL, ""))
            if refund_flag == "已退款":
                # 平台订单金额：|Charge refund| * 1.2（含 Google 代扣税费）
                r  = _to_float(row.get("_g_refund_amt", ""))
                mr  = _to_float(row.get("_g_merchant_refund_amt", ""))
                mfr = _to_float(row.get("_g_merchant_fee_refund_amt", ""))
                mc_charge = row.get("_g_merchant_charge_amt", "")
                mf_charge = row.get("_g_merchant_fee_amt", "")
                charge_net = _google_charge_net_amount(mc_charge, mf_charge)
                if charge_net is not None:
                    google_charge_net = str(charge_net)
                if r is not None:
                    refund_amount = abs(r or 0.0)
                    original_charge_amt = str(round(refund_amount, 2))
                    amt = str(round(refund_amount * 1.2, 2))
                    if rate is not None:
                        country_tax = str(round(refund_amount * 0.2 * rate, 2))
                    # 结算金额：|merchant refund| - |merchant fee_refund|
                    if mr is not None:
                        settle_amt = str(round(abs(mr or 0.0) - abs(mfr or 0.0), 2))
                    # 手续费：|merchant fee_refund|
                    if mfr is not None:
                        fee_amt = str(round(abs(mfr), 2))
                else:
                    # Google报表暂无Charge refund行（退款发生在报表周期外），以Charge行兜底
                    # 标记为"退款待确认"，归入匹配订单差异sheet
                    c_fb = _to_float(row.get("_g_charge_amt", ""))
                    mc_fb = _to_float(row.get("_g_merchant_charge_amt", ""))
                    mf_fb = _to_float(row.get("_g_merchant_fee_amt", ""))
                    if c_fb is not None:
                        charge_amount = abs(c_fb or 0.0)
                        original_charge_amt = str(round(charge_amount, 2))
                        amt = str(round(charge_amount * 1.2, 2))
                        if rate is not None:
                            country_tax = str(round(charge_amount * 0.2 * rate, 2))
                        match_tag = "退款待确认"
                    else:
                        amt = ""
                    if mc_fb is not None:
                        settle_amt = str(round((mc_fb or 0.0) - abs(mf_fb or 0.0), 2))
                    if mf_fb is not None:
                        fee_amt = str(round(abs(mf_fb), 2))
            else:
                # 平台订单金额：|Charge| * 1.2（含 Google 代扣税费）
                c = _to_float(row.get("_g_charge_amt", ""))
                mc = _to_float(row.get("_g_merchant_charge_amt", ""))
                mf = _to_float(row.get("_g_merchant_fee_amt", ""))
                charge_net = _google_charge_net_amount(row.get("_g_merchant_charge_amt", ""), row.get("_g_merchant_fee_amt", ""))
                if charge_net is not None:
                    google_charge_net = str(charge_net)
                if c is not None:
                    charge_amount = abs(c or 0.0)
                    original_charge_amt = str(round(charge_amount, 2))
                    amt = str(round(charge_amount * 1.2, 2))
                    if rate is not None:
                        country_tax = str(round(charge_amount * 0.2 * rate, 2))
                else:
                    amt = ""
                # 结算金额：merchant charge - |merchant fee|
                if mc is not None:
                    settle_amt = str(round((mc or 0.0) - abs(mf or 0.0), 2))
                # 手续费：|merchant Google fee|
                if mf is not None:
                    fee_amt = str(round(abs(mf), 2))
            matched = amt != ""

        elif "苹果支付" in src:
            amt        = str(row.get(ADMIN_AMOUNT_COL, "")).strip()
            ccy        = ""   # Admin 无币种列，留空
            settle_ccy = ""
            matched    = True  # Admin 中已确认付款，无需平台文件验证
            settle_amt = amt   # 暂无苹果手续费数据，结算金额取 Admin 金额
            transaction_date = _format_date(row.get(ADMIN_DATE_COL, ""))

        else:
            amt = ccy = ""
            settle_ccy = ""
            matched = False

        original_charge_amt_list.append(original_charge_amt)
        platform_amt_list.append(amt)
        platform_ccy_list.append(ccy)
        settlement_ccy_list.append(settle_ccy)
        match_status_list.append(match_tag if match_tag is not None else ("是" if matched else "否"))
        settlement_list.append(settle_amt)
        fee_list.append(fee_amt)
        country_tax_list.append(country_tax)
        transaction_date_list.append(transaction_date)
        google_charge_net_list.append(google_charge_net)

        psp_ref = str(row.get(ADMIN_JOIN_COL, "")).strip()
        if refund_flag == "已退款":
            status_list.append("退款")
        elif src == "Adyen":
            if matched:
                status_list.append("成功")
            else:
                status_list.append("")
        else:
            if matched:
                status_list.append("成功")
            else:
                status_list.append("")

    # ── 清理临时列，插入新增列 ────────────────────────────────
    tmp_cols = [c for c in result.columns if c.startswith(("_a_", "_h_", "_g_"))]
    result = result.drop(columns=tmp_cols)

    result.insert(admin_col_count + 0, ORIGINAL_CHARGE_AMOUNT_COL, original_charge_amt_list)
    result.insert(admin_col_count + 1, PLATFORM_AMOUNT_COL,   platform_amt_list)
    result.insert(admin_col_count + 2, PLATFORM_CURRENCY_COL, platform_ccy_list)
    result.insert(admin_col_count + 3, SETTLEMENT_CURRENCY_COL, settlement_ccy_list)
    result.insert(admin_col_count + 4, MATCH_STATUS_COL,      match_status_list)
    result.insert(admin_col_count + 5, STATUS_COL,            status_list)
    result.insert(admin_col_count + 6, SETTLEMENT_AMOUNT_COL, settlement_list)
    result.insert(admin_col_count + 7, FEE_COL,               fee_list)
    result.insert(admin_col_count + 8, COUNTRY_TAX_COL,       country_tax_list)
    result.insert(admin_col_count + 9, TRANSACTION_DATE_COL,  transaction_date_list)
    result[GOOGLE_CHARGE_NET_INTERNAL_COL] = google_charge_net_list

    # ── 追加平台多余行（平台有、admin 无对应记录）────────────
    admin_keys = set(admin_df[ADMIN_JOIN_COL].str.strip())
    extra_df = _build_platform_only_rows(
        list(result.columns),
        admin_keys,
        adyen_lk if adyen_avail else None,
        huawei_lk if huawei_avail else None,
        google_lk if google_avail else None,
    )
    if not extra_df.empty:
        logging.info(
            "追加平台多余行：共 %d 行（在平台报表中存在但 admin 无对应记录，回调丢失或网络延迟）",
            len(extra_df),
        )
        result = pd.concat([result, extra_df], ignore_index=True)

    return result.fillna("")


def log_match_stats(result_df: pd.DataFrame) -> None:
    """按平台打印匹配统计（成功 / 失败 / 退款）。"""
    if ADMIN_PAYMENT_COL not in result_df.columns:
        return

    platform_map = [("Adyen", "Adyen"), ("华为支付", "华为支付"), ("Google支付", "Google支付")]
    for label, pay_val in platform_map:
        rows = result_df[result_df[ADMIN_PAYMENT_COL].str.strip() == pay_val]
        if rows.empty:
            continue
        success = (rows[STATUS_COL] == "成功").sum()
        failed = (rows[STATUS_COL] == "失败").sum()
        refund = (rows[STATUS_COL] == "退款").sum()
        logging.info("【%s】共 %d 条 — 成功 %d，失败 %d，退款 %d", label, len(rows), success, failed, refund)
        unmatched_count = (
            rows[MATCH_STATUS_COL].astype(str).str.strip().eq("否").sum()
            if MATCH_STATUS_COL in rows.columns
            else failed
        )
        if unmatched_count > 0:
            logging.warning("【%s】存在 %d 条未匹配记录，请查看【匹配失败订单】sheet", label, unmatched_count)


def build_google_cashflow_summary(google_raw_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """按 Google 源文件现金流日期构建月度汇总行。"""
    summary_cols = [
        TRANSACTION_DATE_COL,
        ADMIN_PAYMENT_COL,
        SETTLEMENT_CURRENCY_COL,
        "成功笔数",
        "成功金额",
        "退款笔数",
        "退款金额",
        "退款待确认笔数",
        "退款待确认金额",
        "净交易金额",
        "手续费",
        TAX_COL,
    ]
    if google_raw_df is None or google_raw_df.empty:
        return pd.DataFrame(columns=summary_cols)

    required = {
        GOOGLE_JOIN_COL,
        GOOGLE_TRANSACTION_TYPE_COL,
        GOOGLE_DATE_COL,
        GOOGLE_MERCHANT_AMOUNT_COL,
        GOOGLE_MERCHANT_CURRENCY_COL,
    }
    if not required.issubset(set(google_raw_df.columns)):
        missing = sorted(required - set(google_raw_df.columns))
        logging.warning("Google 源文件缺少现金流汇总必需列，跳过源流水汇总: %s", ", ".join(missing))
        return pd.DataFrame(columns=summary_cols)

    df = google_raw_df.copy()
    df[GOOGLE_JOIN_COL] = df[GOOGLE_JOIN_COL].fillna("").astype(str).str.strip()
    df[GOOGLE_TRANSACTION_TYPE_COL] = df[GOOGLE_TRANSACTION_TYPE_COL].fillna("").astype(str).str.strip()
    df[GOOGLE_DATE_COL] = df[GOOGLE_DATE_COL].fillna("").astype(str).str.strip()
    df[GOOGLE_MERCHANT_CURRENCY_COL] = (
        df[GOOGLE_MERCHANT_CURRENCY_COL].fillna("").astype(str).str.strip()
    )
    amount_raw = (
        df[GOOGLE_MERCHANT_AMOUNT_COL]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.replace(",", "", regex=False)
    )
    df["_merchant_amount"] = pd.to_numeric(amount_raw, errors="coerce").fillna(0.0)
    df["_date"] = df[GOOGLE_DATE_COL].map(_format_date)
    df["_month"] = df["_date"].astype(str).str[:7]
    df = df[(df[GOOGLE_JOIN_COL] != "") & (df["_month"].str.len() == 7)].copy()
    if df.empty:
        return pd.DataFrame(columns=summary_cols)

    key_cols = [GOOGLE_JOIN_COL, GOOGLE_DATE_COL, GOOGLE_MERCHANT_CURRENCY_COL]

    def _typed(type_name: str, amount_col: str, count_col: str) -> pd.DataFrame:
        sub = df[df[GOOGLE_TRANSACTION_TYPE_COL] == type_name].copy()
        if sub.empty:
            return pd.DataFrame(columns=key_cols + ["_month", amount_col, count_col])
        return (
            sub.groupby(key_cols, as_index=False)
            .agg(
                _month=("_month", "first"),
                **{
                    amount_col: ("_merchant_amount", "sum"),
                    count_col: (GOOGLE_JOIN_COL, "size"),
                },
            )
        )

    charges = _typed(GOOGLE_CHARGE_TYPE, "_charge_amount", "_charge_count")
    fees = _typed(GOOGLE_FEE_TYPE, "_fee_amount", "_fee_count").drop(columns=["_month"], errors="ignore")
    refunds = _typed(GOOGLE_REFUND_TYPE, "_refund_amount", "_refund_count")
    fee_refunds = _typed(GOOGLE_FEE_REFUND_TYPE, "_fee_refund_amount", "_fee_refund_count").drop(
        columns=["_month"], errors="ignore"
    )

    success_rows = pd.DataFrame()
    if not charges.empty:
        success_rows = charges.merge(fees, on=key_cols, how="left")
        success_rows["_fee_amount"] = success_rows["_fee_amount"].fillna(0.0)
        success_rows["_success_count"] = success_rows["_charge_count"].fillna(0).astype(int)
        success_rows["_refund_count_out"] = 0
        success_rows["_success_amount"] = success_rows["_charge_amount"] - success_rows["_fee_amount"].abs()
        success_rows["_refund_amount_out"] = 0.0
        success_rows["_fee_success"] = success_rows["_fee_amount"].abs()
        success_rows["_fee_refund"] = 0.0

    refund_rows = pd.DataFrame()
    if not refunds.empty:
        refund_rows = refunds.merge(fee_refunds, on=key_cols, how="left")
        refund_rows["_fee_refund_amount"] = refund_rows["_fee_refund_amount"].fillna(0.0)
        refund_rows["_success_count"] = 0
        refund_rows["_refund_count_out"] = refund_rows["_refund_count"].fillna(0).astype(int)
        refund_rows["_success_amount"] = 0.0
        refund_rows["_refund_amount_out"] = (
            refund_rows["_refund_amount"].abs() - refund_rows["_fee_refund_amount"].abs()
        )
        refund_rows["_fee_success"] = 0.0
        refund_rows["_fee_refund"] = refund_rows["_fee_refund_amount"].abs()

    cashflow = pd.concat([success_rows, refund_rows], ignore_index=True, sort=False)
    if cashflow.empty:
        return pd.DataFrame(columns=summary_cols)

    cashflow[ADMIN_PAYMENT_COL] = "Google支付"
    grouped = (
        cashflow.groupby(["_month", ADMIN_PAYMENT_COL, GOOGLE_MERCHANT_CURRENCY_COL], as_index=False)
        .agg(
            成功笔数=("_success_count", "sum"),
            成功金额=("_success_amount", "sum"),
            退款笔数=("_refund_count_out", "sum"),
            退款金额=("_refund_amount_out", "sum"),
            _fee_success=("_fee_success", "sum"),
            _fee_refund=("_fee_refund", "sum"),
        )
        .rename(
            columns={
                "_month": TRANSACTION_DATE_COL,
                GOOGLE_MERCHANT_CURRENCY_COL: SETTLEMENT_CURRENCY_COL,
            }
        )
    )
    grouped["退款待确认笔数"] = 0
    grouped["退款待确认金额"] = 0.0
    grouped["净交易金额"] = grouped["成功金额"] - grouped["退款金额"]
    grouped["手续费"] = (grouped["_fee_success"] - grouped["_fee_refund"]).abs()
    grouped[TAX_COL] = 0.0

    amount_cols = ["成功金额", "退款金额", "退款待确认金额", "净交易金额", "手续费", TAX_COL]
    grouped[amount_cols] = grouped[amount_cols].round(4)
    grouped = grouped.drop(columns=["_fee_success", "_fee_refund"])
    return grouped[summary_cols].sort_values(
        [TRANSACTION_DATE_COL, ADMIN_PAYMENT_COL, SETTLEMENT_CURRENCY_COL],
        kind="stable",
    ).reset_index(drop=True)


def build_summary_sheet(
    result_df: pd.DataFrame,
    huawei_settle_df: Optional[pd.DataFrame] = None,
    adyen_settle_df: Optional[pd.DataFrame] = None,
    google_raw_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """按交易日期、平台、结算币种汇总结算金额，附手续费和税金列。

    手续费来源：
        Adyen  — 来自 adyen_settle_df（月度结算文件）：手续费 = 成功金额 - 净交易金额 - 税金
        Google — 优先聚合 Google 源文件现金流行；未传源文件时聚合 result_df 中的 FEE_COL
        华为   — 来自 huawei_settle_df（月度结算文件），追加为单独的 HKD 结算行
    Adyen 净交易金额 = 结算文件 MerchantPayout 月度合计；税金 = InvoiceDeduction 月度合计。
    """
    summary_cols = [
        TRANSACTION_DATE_COL,
        ADMIN_PAYMENT_COL,
        SETTLEMENT_CURRENCY_COL,
        "成功笔数",
        "成功金额",
        "退款笔数",
        "退款金额",
        "退款待确认笔数",
        "退款待确认金额",
        "净交易金额",
        "手续费",
        TAX_COL,
    ]

    if result_df.empty:
        return pd.DataFrame(columns=summary_cols)

    df = result_df.copy()
    df[STATUS_COL] = df[STATUS_COL].astype(str).str.strip()
    df = df[df[STATUS_COL].isin(["成功", "退款"])].copy()
    if df.empty:
        return pd.DataFrame(columns=summary_cols)

    group_cols = [TRANSACTION_DATE_COL, ADMIN_PAYMENT_COL, SETTLEMENT_CURRENCY_COL]
    for col in group_cols:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df[TRANSACTION_DATE_COL] = df[TRANSACTION_DATE_COL].str[:7]

    amount = df[SETTLEMENT_AMOUNT_COL].fillna("").astype(str).str.strip().str.replace(",", "", regex=False)
    df["_amount"] = pd.to_numeric(amount, errors="coerce").fillna(0.0)
    if GOOGLE_CHARGE_NET_INTERNAL_COL in df.columns:
        charge_net_raw = (
            df[GOOGLE_CHARGE_NET_INTERNAL_COL]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.replace(",", "", regex=False)
        )
        df["_google_charge_net"] = pd.to_numeric(charge_net_raw, errors="coerce").fillna(0.0)
    else:
        df["_google_charge_net"] = 0.0

    if MATCH_STATUS_COL in df.columns:
        pending_mask = df[MATCH_STATUS_COL].astype(str).str.strip().eq("退款待确认")
    else:
        pending_mask = pd.Series(False, index=df.index)

    google_mask = df[ADMIN_PAYMENT_COL].astype(str).str.contains("Google", na=False)
    google_refund_charge_mask = google_mask & (df[STATUS_COL] == "退款")

    df["_success_count"]  = ((df[STATUS_COL] == "成功") & ~pending_mask).astype(int)
    df["_refund_count"]   = ((df[STATUS_COL] == "退款") & ~pending_mask).astype(int)
    df["_pending_count"]  = pending_mask.astype(int)
    df["_success_amount"] = (
        df["_amount"].where((df[STATUS_COL] == "成功") & ~pending_mask, 0.0)
        + df["_google_charge_net"].where(google_refund_charge_mask, 0.0)
    )
    df["_refund_amount"]  = df["_amount"].where((df[STATUS_COL] == "退款") & ~pending_mask, 0.0)
    df["_pending_amount"] = df["_amount"].where(pending_mask, 0.0)

    if FEE_COL in df.columns:
        fee_raw = df[FEE_COL].fillna("").astype(str).str.strip().str.replace(",", "", regex=False)
        df["_fee"] = pd.to_numeric(fee_raw, errors="coerce").fillna(0.0)
    else:
        df["_fee"] = 0.0
    df["_fee_success"] = df["_fee"].where((df[STATUS_COL] == "成功") & ~pending_mask, 0.0)
    df["_fee_refund"]  = df["_fee"].where((df[STATUS_COL] == "退款") & ~pending_mask, 0.0)

    summary = (
        df.groupby(group_cols, as_index=False)
        .agg(
            成功笔数=("_success_count", "sum"),
            成功金额=("_success_amount", "sum"),
            退款笔数=("_refund_count", "sum"),
            退款金额=("_refund_amount", "sum"),
            退款待确认笔数=("_pending_count", "sum"),
            退款待确认金额=("_pending_amount", "sum"),
            _fee_success=("_fee_success", "sum"),
            _fee_refund=("_fee_refund", "sum"),
        )
        .sort_values(group_cols, kind="stable")
        .reset_index(drop=True)
    )
    summary["净交易金额"] = summary["成功金额"] - summary["退款金额"]
    summary["手续费"] = (summary["_fee_success"] - summary["_fee_refund"]).round(4)
    summary = summary.drop(columns=["_fee_success", "_fee_refund"])
    summary[TAX_COL] = 0.0

    google_cashflow_summary = build_google_cashflow_summary(google_raw_df)
    if not google_cashflow_summary.empty:
        google_summary_mask = summary[ADMIN_PAYMENT_COL].astype(str).str.contains("Google", na=False)
        summary = (
            pd.concat([summary[~google_summary_mask], google_cashflow_summary], ignore_index=True)
            .sort_values(group_cols, kind="stable")
            .reset_index(drop=True)
        )

    if huawei_settle_df is not None and not huawei_settle_df.empty:
        huawei_settle_rows = build_huawei_fee_summary(huawei_settle_df)
        if not huawei_settle_rows.empty:
            # 用结算文件 HKD 行替换原有华为 TRY 订单行，避免同日期双行
            hw_mask = summary[ADMIN_PAYMENT_COL].str.strip() == "华为支付"
            non_huawei_summary = summary[~hw_mask]
            huawei_order_rows  = summary[hw_mask]

            # 从订单匹配行取笔数（按月合并所有交易货币），填入 HKD 结算行
            if not huawei_order_rows.empty:
                count_by_month = (
                    huawei_order_rows
                    .groupby(TRANSACTION_DATE_COL, as_index=False)
                    .agg(_cnt_s=("成功笔数", "sum"), _cnt_r=("退款笔数", "sum"))
                )
                huawei_settle_rows = huawei_settle_rows.merge(
                    count_by_month, on=TRANSACTION_DATE_COL, how="left"
                )
                huawei_settle_rows["成功笔数"] = huawei_settle_rows["_cnt_s"].fillna(0).astype(int)
                huawei_settle_rows["退款笔数"] = huawei_settle_rows["_cnt_r"].fillna(0).astype(int)
                huawei_settle_rows = huawei_settle_rows.drop(columns=["_cnt_s", "_cnt_r"])

            summary = (
                pd.concat([non_huawei_summary, huawei_settle_rows], ignore_index=True)
                .sort_values(group_cols, kind="stable")
                .reset_index(drop=True)
            )

    # ── ADYEN 结算文件：替换净交易金额、手续费、税金 ────────────
    if adyen_settle_df is not None and not adyen_settle_df.empty:
        monthly = build_adyen_settlement_monthly(adyen_settle_df)
        if not monthly.empty:
            adyen_mask = summary[ADMIN_PAYMENT_COL].str.strip() == "Adyen"
            if adyen_mask.any():
                adyen_rows = summary[adyen_mask].merge(
                    monthly, on=[TRANSACTION_DATE_COL, SETTLEMENT_CURRENCY_COL], how="left"
                )
                adyen_rows["_adyen_payout"] = adyen_rows["_adyen_payout"].fillna(0.0)
                adyen_rows["_adyen_tax"]    = adyen_rows["_adyen_tax"].fillna(0.0)
                # 净交易金额 = MerchantPayout（实际到账）
                adyen_rows["净交易金额"] = adyen_rows["_adyen_payout"].round(4)
                adyen_rows[TAX_COL]     = adyen_rows["_adyen_tax"].round(4)
                # 手续费 = 成功金额（订单匹配结果）- 净交易金额 - 税金
                adyen_rows["手续费"] = (
                    adyen_rows["成功金额"] - adyen_rows["净交易金额"] - adyen_rows[TAX_COL]
                ).round(4)
                adyen_rows = adyen_rows.drop(columns=["_adyen_payout", "_adyen_tax"])
                summary = (
                    pd.concat([summary[~adyen_mask], adyen_rows], ignore_index=True)
                    .sort_values(group_cols, kind="stable")
                    .reset_index(drop=True)
                )
            else:
                logging.info("ADYEN 结算文件已读取，但汇总表中未找到 Adyen 平台行，结算数据未应用")

    summary["手续费"] = pd.to_numeric(summary["手续费"], errors="coerce").abs().round(4)
    for _col in ("退款待确认笔数", "退款待确认金额"):
        if _col in summary.columns:
            summary[_col] = summary[_col].fillna(0)
    return summary[summary_cols]


def build_apple_platform_summary(apple_raw_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """从苹果平台报表数据构建交易金额汇总行。

    以 Settlement Date × 结算货币 分组，按 Quantity 正负区分成功/退款。
    """
    summary_cols = [
        TRANSACTION_DATE_COL, ADMIN_PAYMENT_COL, SETTLEMENT_CURRENCY_COL,
        "成功笔数", "成功金额", "退款笔数", "退款金额", "净交易金额", "手续费", TAX_COL,
    ]
    if apple_raw_df is None or apple_raw_df.empty:
        return pd.DataFrame(columns=summary_cols)

    df = apple_raw_df.copy()

    date_col = next(
        (c for c in ["Settlement Date", "Transaction Date"] if c in df.columns), None
    )
    if date_col is None:
        logging.warning("苹果平台报表未找到日期列，跳过苹果汇总")
        return pd.DataFrame(columns=summary_cols)
    df["_date"] = (
        pd.to_datetime(df[date_col].astype(str).str.strip(), errors="coerce")
        .dt.strftime("%Y-%m").fillna("")
    )
    df = df[df["_date"] != ""]

    currency_col = next(
        (c for c in ["Currency of Proceeds", "Partner Share Currency"] if c in df.columns), None
    )
    df["_currency"] = df[currency_col].astype(str).str.strip() if currency_col else ""

    if "Extended Partner Share" not in df.columns:
        logging.warning("苹果平台报表未找到 Extended Partner Share 列，跳过苹果汇总")
        return pd.DataFrame(columns=summary_cols)
    df["_eps"] = pd.to_numeric(
        df["Extended Partner Share"].astype(str).str.replace(",", "", regex=False),
        errors="coerce",
    ).fillna(0.0)

    if "Quantity" in df.columns:
        df["_qty"] = pd.to_numeric(
            df["Quantity"].astype(str).str.replace(",", "", regex=False), errors="coerce"
        ).fillna(0.0)
    else:
        df["_qty"] = df["_eps"].apply(lambda x: 1.0 if x >= 0 else -1.0)

    fee_cols = ["Quantity", "Customer Price", "Extended Partner Share"]
    has_fee_cols = all(c in df.columns for c in fee_cols)
    if has_fee_cols:
        customer_price = pd.to_numeric(
            df["Customer Price"].astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        ).fillna(0.0)
        df["_fee"] = (df["_qty"] * customer_price - df["_eps"]).fillna(0.0)
    else:
        logging.warning("苹果平台报表缺少手续费计算列: %s，苹果汇总手续费留空", [c for c in fee_cols if c not in df.columns])
        df["_fee"] = 0.0

    df["_is_success"] = df["_qty"] > 0
    df["_is_refund"]  = df["_qty"] < 0

    rows = []
    for (dt, cur), grp in df.groupby(["_date", "_currency"], sort=True):
        rows.append({
            TRANSACTION_DATE_COL:    dt,
            ADMIN_PAYMENT_COL:       "苹果支付Lua",
            SETTLEMENT_CURRENCY_COL: cur,
            "成功笔数": int(grp.loc[grp["_is_success"], "_qty"].sum()),
            "成功金额": round(grp.loc[grp["_is_success"], "_eps"].sum(), 2),
            "退款笔数": int(grp.loc[grp["_is_refund"], "_qty"].abs().sum()),
            "退款金额": round(grp.loc[grp["_is_refund"], "_eps"].abs().sum(), 2),
            "手续费": round(grp["_fee"].sum(), 4) if has_fee_cols else "",
            TAX_COL: 0.0,
        })

    if not rows:
        return pd.DataFrame(columns=summary_cols)

    result = pd.DataFrame(rows)
    result["净交易金额"] = result["成功金额"] - result["退款金额"]
    if has_fee_cols:
        result["手续费"] = pd.to_numeric(result["手续费"], errors="coerce").abs().round(4)
    return result[summary_cols]


def build_monthly_comparison(
    summary_df: pd.DataFrame,
    apple_admin_df: pd.DataFrame,
    result_df: pd.DataFrame,
) -> pd.DataFrame:
    """按月汇总对账差异：Admin结算金额（匹配成功的 admin 订单毛额）vs 平台净到账（成功-退款）。

    返回列：月份, 支付方式, Admin结算币种, Admin结算金额, 结算币种, 平台净到账
    苹果行：Admin侧取 apple_admin_df 正常订单金额，同月多币种时仅填第一个币种行。
    """
    comp_cols = [
        "月份", ADMIN_PAYMENT_COL, "Admin结算币种", "Admin结算金额", SETTLEMENT_CURRENCY_COL, "平台净到账",
    ]
    if summary_df.empty:
        return pd.DataFrame(columns=comp_cols)

    segments = []

    # ── 非苹果部分 ──────────────────────────────────────────
    non_apple = summary_df[
        summary_df[ADMIN_PAYMENT_COL].astype(str).str.strip() != "苹果支付Lua"
    ].copy()
    if not non_apple.empty:
        non_apple["月份"] = non_apple[TRANSACTION_DATE_COL].astype(str).str[:7]
        grp_cols = ["月份", ADMIN_PAYMENT_COL, SETTLEMENT_CURRENCY_COL]

        # 平台侧：平台净到账 来自 summary_df
        plat_monthly = (
            non_apple.groupby(grp_cols, as_index=False, sort=True)
            .agg(平台净到账=("净交易金额", "sum"))
        )

        # Admin侧：仅统计 admin 行中已经匹配平台的订单，使用 admin 订单原始金额（毛额）。
        if MATCH_STATUS_COL in result_df.columns:
            matched_mask = result_df[MATCH_STATUS_COL].astype(str).str.strip().eq("是")
        else:
            matched_mask = result_df[SETTLEMENT_AMOUNT_COL].astype(str).str.strip().ne("")
        matched = result_df[
            matched_mask &
            result_df[ADMIN_PAYMENT_COL].astype(str).str.strip().ne("苹果支付Lua")
        ].copy()
        if not matched.empty and ADMIN_AMOUNT_COL in matched.columns:
            matched["月份"] = matched[TRANSACTION_DATE_COL].astype(str).str[:7]
            matched["_amt"] = pd.to_numeric(
                matched[ADMIN_AMOUNT_COL].astype(str).str.replace(",", "", regex=False),
                errors="coerce",
            ).fillna(0.0)
            admin_monthly = (
                matched.groupby(grp_cols, as_index=False, sort=True)
                .agg(
                    Admin结算币种=(PLATFORM_CURRENCY_COL, "first"),
                    Admin结算金额=("_amt", "sum"),
                )
            )
            admin_monthly["Admin结算金额"] = admin_monthly["Admin结算金额"].round(2)
        else:
            admin_monthly = pd.DataFrame(columns=grp_cols + ["Admin结算币种", "Admin结算金额"])

        monthly = plat_monthly.merge(admin_monthly, on=grp_cols, how="left")
        monthly["Admin结算币种"] = monthly["Admin结算币种"].fillna("")
        monthly["Admin结算金额"] = monthly["Admin结算金额"].fillna(0.0).round(2)
        monthly["平台净到账"] = monthly["平台净到账"].round(2)
        segments.append(monthly[comp_cols])

    # ── 苹果部分 ─────────────────────────────────────────────
    apple_plat = summary_df[
        summary_df[ADMIN_PAYMENT_COL].astype(str).str.strip() == "苹果支付Lua"
    ].copy()
    if not apple_plat.empty:
        apple_plat["月份"] = apple_plat[TRANSACTION_DATE_COL].astype(str).str[:7]
        plat_grp = (
            apple_plat.groupby(["月份", SETTLEMENT_CURRENCY_COL], as_index=False, sort=True)
            .agg(平台净到账=("净交易金额", "sum"))
        )
        plat_grp[ADMIN_PAYMENT_COL] = "苹果支付Lua"

        # Admin侧：按月份聚合 apple_admin_df 正常订单金额
        apple_admin_monthly: dict = {}
        if not apple_admin_df.empty and ADMIN_DATE_COL in apple_admin_df.columns and ADMIN_AMOUNT_COL in apple_admin_df.columns:
            adf = apple_admin_df.copy()
            adf["_月份"] = (
                pd.to_datetime(adf[ADMIN_DATE_COL].astype(str).str.strip(), errors="coerce")
                .dt.strftime("%Y-%m").fillna("")
            )
            adf["_金额"] = pd.to_numeric(
                adf[ADMIN_AMOUNT_COL].astype(str).str.replace(",", "", regex=False),
                errors="coerce",
            ).fillna(0.0)
            flag = (
                adf[ADMIN_REFUND_COL].astype(str).str.strip()
                if ADMIN_REFUND_COL in adf.columns
                else pd.Series("正常", index=adf.index)
            )
            for month, grp in adf[adf["_月份"] != ""].groupby("_月份", sort=True):
                normal_mask = flag.reindex(grp.index) == "正常"
                apple_admin_monthly[month] = round(grp.loc[normal_mask, "_金额"].sum(), 2)

        # 将 admin 月汇总金额填入苹果平台行（同月多币种：仅填第一个币种行）
        filled: set = set()
        admin_amts = []
        for _, r in plat_grp.iterrows():
            month = r["月份"]
            if month not in filled and month in apple_admin_monthly:
                admin_amts.append(apple_admin_monthly[month])
                filled.add(month)
            else:
                admin_amts.append(0.0)
        plat_grp["Admin结算币种"] = ""
        plat_grp["Admin结算金额"] = admin_amts
        plat_grp["平台净到账"] = plat_grp["平台净到账"].round(2)
        segments.append(plat_grp[comp_cols])

    if not segments:
        return pd.DataFrame(columns=comp_cols)

    result = pd.concat(segments, ignore_index=True)
    result = result.sort_values(
        ["月份", ADMIN_PAYMENT_COL, SETTLEMENT_CURRENCY_COL], kind="stable"
    ).reset_index(drop=True)
    return result[comp_cols]


def _append_comparison_table(
    ws,
    comparison_df: pd.DataFrame,
    start_row: int,
) -> None:
    """将月度对比表追加到 openpyxl worksheet 的 start_row 行位置。"""
    from openpyxl.styles import Font

    headers = list(comparison_df.columns)
    numeric_headers = {"Admin结算金额", "平台净到账"}

    row = start_row
    ws.cell(row, 1).value = "月度对比（对账差异）"
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci)
        c.value = h
        c.font = Font(bold=True)
    row += 1

    for _, dr in comparison_df.iterrows():
        for ci, val in enumerate(dr, 1):
            cell = ws.cell(row, ci)
            if headers[ci - 1] in numeric_headers:
                try:
                    cell.value = float(val)
                except (ValueError, TypeError):
                    cell.value = val
            else:
                cell.value = str(val) if pd.notna(val) else ""
        row += 1


def _write_apple_sheet(
    writer: pd.ExcelWriter,
    apple_admin_df: pd.DataFrame,
    apple_raw_df: Optional[pd.DataFrame],
) -> None:
    """在 ExcelWriter 中写入苹果支付 Sheet。

    布局（从上到下）：
      合并日期汇总：admin 与平台同结算日期数据并排，一行一个日期
      Admin 苹果订单详情 + 统计
      苹果平台报表详情 + 合计
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    ws = writer.book.create_sheet(OUTPUT_APPLE_SHEET_3)
    writer.sheets[OUTPUT_APPLE_SHEET_3] = ws

    row = 1

    def _title(r, text, size=12):
        ws.cell(r, 1).value = text
        ws.cell(r, 1).font = Font(bold=True, size=size)

    def _header(r, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(r, ci)
            c.value = h
            c.font = Font(bold=True)

    def _coerce(val):
        """将字符串型数字转为 float，非数字保持字符串，供公式正常计算。"""
        s = str(val).strip()
        if s == "":
            return ""
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return s

    def _sum_col(r, ci, r_start, r_end):
        if r_end >= r_start:
            cl = get_column_letter(ci)
            ws.cell(r, ci).value = f"=SUM({cl}{r_start}:{cl}{r_end})"
        else:
            ws.cell(r, ci).value = 0

    # ══════════════════════════════════════════════════════════════
    # 合并日期汇总（admin 与平台同行并排）
    # ══════════════════════════════════════════════════════════════
    _title(row, "苹果订单日期汇总")
    row += 1

    merged_headers = [
        "日期",
        "笔数", "正常笔数", "退款笔数", "正常金额", "退款金额", "合计金额",
        "Quantity合计", "Extended Partner Share合计", "客户支付合计", "手续费合计",
    ]
    _header(row, merged_headers)
    row += 1

    # -- Admin 按支付日期聚合 --
    admin_summary: dict = {}
    if not apple_admin_df.empty:
        adf = apple_admin_df.copy()
        adf["_date"] = (
            pd.to_datetime(adf[ADMIN_DATE_COL].astype(str).str.strip(), errors="coerce")
            .dt.strftime("%Y-%m-%d").fillna("")
            if ADMIN_DATE_COL in adf.columns else ""
        )
        adf["_amount"] = (
            pd.to_numeric(
                adf[ADMIN_AMOUNT_COL].astype(str).str.replace(",", "", regex=False),
                errors="coerce",
            ).fillna(0.0)
            if ADMIN_AMOUNT_COL in adf.columns else 0.0
        )
        flag = (
            adf[ADMIN_REFUND_COL].astype(str).str.strip()
            if ADMIN_REFUND_COL in adf.columns
            else pd.Series("正常", index=adf.index)
        )
        adf["_normal"] = flag == "正常"
        adf["_refund"] = flag == "已退款"
        for dt, grp in adf.groupby("_date", sort=True):
            n_amt = round(grp.loc[grp["_normal"], "_amount"].sum(), 2)
            r_amt = round(grp.loc[grp["_refund"], "_amount"].sum(), 2)
            admin_summary[dt] = [
                len(grp), int(grp["_normal"].sum()), int(grp["_refund"].sum()),
                n_amt, r_amt, round(n_amt - r_amt, 2),
            ]

    # -- 平台 按结算日期聚合 --
    platform_summary: dict = {}
    if apple_raw_df is not None and not apple_raw_df.empty:
        plat_date_col = next(
            (c for c in ["Settlement Date", "Transaction Date", "Begin Date", "End Date", "Report Date", "Date", "日期"]
             if c in apple_raw_df.columns),
            None,
        )
        if plat_date_col:
            pdf = apple_raw_df.copy()
            pdf["_date"] = (
                pd.to_datetime(pdf[plat_date_col].astype(str).str.strip(), errors="coerce")
                .dt.strftime("%Y-%m-%d").fillna("")
            )
            pdf = pdf[pdf["_date"] != ""]

            def _num(col):
                if col not in pdf.columns:
                    return pd.Series(0.0, index=pdf.index)
                return pd.to_numeric(
                    pdf[col].astype(str).str.replace(",", "", regex=False), errors="coerce"
                ).fillna(0.0)

            pdf["_qty"] = _num("Quantity")
            pdf["_eps"] = _num("Extended Partner Share")
            pdf["_tcp"] = pdf["_qty"] * _num("Customer Price")
            pdf["_fee"] = pdf["_tcp"] - pdf["_eps"]

            for dt, grp in pdf.groupby("_date", sort=True):
                platform_summary[dt] = [
                    int(grp["_qty"].sum()),
                    round(grp["_eps"].sum(), 2),
                    round(grp["_tcp"].sum(), 2),
                    round(grp["_fee"].sum(), 2),
                ]

    # -- 合并写出：admin 与平台 outer join 按日期排序 --
    all_dates = sorted(set(admin_summary) | set(platform_summary))
    summary_start = row
    for dt in all_dates:
        a = admin_summary.get(dt, ["", "", "", "", "", ""])
        p = platform_summary.get(dt, ["", "", "", ""])
        vals = [dt] + a + p
        for ci, v in enumerate(vals, 1):
            ws.cell(row, ci).value = v
        row += 1
    summary_end = row - 1

    ws.cell(row, 1).value = "合计"
    ws.cell(row, 1).font = Font(bold=True)
    for ci in range(2, len(merged_headers) + 1):
        _sum_col(row, ci, summary_start, summary_end)
    row += 2

    # ══════════════════════════════════════════════════════════════
    # Admin 苹果订单详情
    # ══════════════════════════════════════════════════════════════
    _title(row, "Admin 苹果订单详情")
    row += 1

    admin_cols = list(apple_admin_df.columns)
    _header(row, admin_cols)
    row += 1

    admin_data_start = row
    for _, dr in apple_admin_df.iterrows():
        for ci, val in enumerate(dr, 1):
            ws.cell(row, ci).value = _coerce(val)
        row += 1
    admin_data_end = row - 1
    row += 1

    ws.cell(row, 1).value = "统计"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    _header(row, ["", "笔数", "合计金额"])
    row += 1

    ridx = (admin_cols.index(ADMIN_REFUND_COL) + 1) if ADMIN_REFUND_COL in admin_cols else None
    aidx = (admin_cols.index(ADMIN_AMOUNT_COL) + 1) if ADMIN_AMOUNT_COL in admin_cols else None

    if admin_data_end >= admin_data_start and ridx and aidx:
        rc = get_column_letter(ridx)
        ac = get_column_letter(aidx)
        rng_r = f"{rc}{admin_data_start}:{rc}{admin_data_end}"
        rng_a = f"{ac}{admin_data_start}:{ac}{admin_data_end}"

        normal_stat_row = row
        ws.cell(row, 1).value = "正常订单"
        ws.cell(row, 2).value = f'=COUNTIF({rng_r},"正常")'
        ws.cell(row, 3).value = f'=SUMIF({rng_r},"正常",{rng_a})'
        row += 1
        refund_stat_row = row
        ws.cell(row, 1).value = "退款订单"
        ws.cell(row, 2).value = f'=COUNTIF({rng_r},"已退款")'
        ws.cell(row, 3).value = f'=SUMIF({rng_r},"已退款",{rng_a})'
        row += 1
        ws.cell(row, 1).value = "合计"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = f"=SUM(B{normal_stat_row}:B{refund_stat_row})"
        ws.cell(row, 3).value = f"=SUM(C{normal_stat_row}:C{refund_stat_row})"
        row += 2
    else:
        row += 1

    # ══════════════════════════════════════════════════════════════
    # 苹果平台报表详情
    # ══════════════════════════════════════════════════════════════
    _title(row, "苹果平台报表详情")
    row += 1

    if apple_raw_df is None or apple_raw_df.empty:
        ws.cell(row, 1).value = "（未找到苹果平台文件）"
        return

    # 追加手续费计算列（= Customer Price × Quantity - Extended Partner Share）
    def _to_num_s(series):
        return pd.to_numeric(
            series.astype(str).str.replace(",", "", regex=False), errors="coerce"
        ).fillna(0.0)

    plat_df = apple_raw_df.copy()
    if all(c in plat_df.columns for c in ["Quantity", "Customer Price", "Extended Partner Share"]):
        plat_df["手续费"] = (
            _to_num_s(plat_df["Quantity"]) * _to_num_s(plat_df["Customer Price"])
            - _to_num_s(plat_df["Extended Partner Share"])
        ).round(4)
    else:
        plat_df["手续费"] = ""

    pcols = list(plat_df.columns)
    _header(row, pcols)
    row += 1

    plat_data_start = row
    for _, dr in plat_df.iterrows():
        for ci, val in enumerate(dr, 1):
            ws.cell(row, ci).value = _coerce(val)
        row += 1
    plat_data_end = row - 1
    row += 1

    # 平台数据合计行
    ps_idx  = (pcols.index("Partner Share") + 1) if "Partner Share" in pcols else None
    ext_idx = (pcols.index("Extended Partner Share") + 1) if "Extended Partner Share" in pcols else None
    cp_idx  = (pcols.index("Customer Price") + 1) if "Customer Price" in pcols else None
    fee_idx = (pcols.index("手续费") + 1) if "手续费" in pcols else None

    ws.cell(row, 2).value = "Partner Share（结算单价）"
    ws.cell(row, 3).value = "Extended Partner Share（合计结算）"
    ws.cell(row, 4).value = "Customer Price（客户单价）"
    ws.cell(row, 5).value = "手续费合计（客户支付-结算）"
    for ci in (2, 3, 4, 5):
        ws.cell(row, ci).font = Font(bold=True)
    row += 1

    ws.cell(row, 1).value = "合计"
    ws.cell(row, 1).font = Font(bold=True)
    for out_ci, src_idx in ((2, ps_idx), (3, ext_idx), (4, cp_idx)):
        if src_idx and plat_data_end >= plat_data_start:
            data_col = get_column_letter(src_idx)
            ws.cell(row, out_ci).value = f"=SUM({data_col}{plat_data_start}:{data_col}{plat_data_end})"
    if fee_idx and plat_data_end >= plat_data_start:
        fee_col = get_column_letter(fee_idx)
        ws.cell(row, 5).value = f"=SUM({fee_col}{plat_data_start}:{fee_col}{plat_data_end})"


def write_output(
    result_df: pd.DataFrame,
    output_dir: Path,
    apple_raw_df: Optional[pd.DataFrame] = None,
    huawei_settle_df: Optional[pd.DataFrame] = None,
    adyen_settle_df: Optional[pd.DataFrame] = None,
    google_raw_df: Optional[pd.DataFrame] = None,
) -> Path:
    """将结果写入 data/output/订单匹配结果_{YYYYMMDD}.xlsx。"""
    today = date.today().strftime("%Y%m%d")
    filename = OUTPUT_FILE_TEMPLATE.format(date=today)
    output_path = output_dir / filename
    output_dir.mkdir(parents=True, exist_ok=True)

    main_df = result_df.reset_index(drop=True)

    if ADMIN_PAYMENT_COL in main_df.columns:
        is_apple = main_df[ADMIN_PAYMENT_COL].astype(str).str.strip().str.contains("苹果支付", na=False)
        apple_admin_df = main_df.loc[is_apple].copy()
        main_df = main_df.loc[~is_apple].reset_index(drop=True)
    else:
        apple_admin_df = pd.DataFrame()

    def _money_series(col_name: str) -> "pd.Series":
        if col_name not in main_df.columns:
            return pd.Series([float("nan")] * len(main_df), index=main_df.index)
        return pd.to_numeric(
            main_df[col_name].astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        )

    if MATCH_STATUS_COL in main_df.columns:
        match_status = main_df[MATCH_STATUS_COL].astype(str).str.strip()
    else:
        match_status = pd.Series("", index=main_df.index)

    admin_amount = _money_series(ADMIN_AMOUNT_COL)
    platform_amount = _money_series(PLATFORM_AMOUNT_COL)
    amount_diff = (admin_amount - platform_amount).abs()

    diff_df = main_df.loc[
        (match_status.eq("是") & amount_diff.gt(1.0)) | match_status.eq("退款待确认")
    ].copy()
    failed_df = main_df.loc[match_status.eq("否")].copy()

    summary_df = build_summary_sheet(main_df, huawei_settle_df, adyen_settle_df, google_raw_df)
    apple_summary_df = build_apple_platform_summary(apple_raw_df)
    if not apple_summary_df.empty:
        summary_df = pd.concat([summary_df, apple_summary_df], ignore_index=True)

    for df in (main_df, diff_df, failed_df):
        if GOOGLE_CHARGE_NET_INTERNAL_COL in df.columns:
            df.drop(columns=[GOOGLE_CHARGE_NET_INTERNAL_COL], inplace=True)

    def _format_detail_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
        ws = writer.sheets[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        main_df.to_excel(writer, sheet_name=OUTPUT_SHEET_3, index=False)
        diff_df.to_excel(writer, sheet_name=OUTPUT_DIFF_SHEET_3, index=False)
        failed_df.to_excel(writer, sheet_name=OUTPUT_FAILED_SHEET_3, index=False)
        for sheet_name in (OUTPUT_SHEET_3, OUTPUT_DIFF_SHEET_3, OUTPUT_FAILED_SHEET_3):
            _format_detail_sheet(writer, sheet_name)
        summary_df.to_excel(writer, sheet_name=OUTPUT_SUMMARY_SHEET_3, index=False)
        _write_apple_sheet(writer, apple_admin_df, apple_raw_df)

    return output_path


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    files = scan_source_files_3(INPUT_DIR_3)
    if not files["admin"]:
        logging.error(
            "未找到 admin 订单文件。请将文件放入 %s，文件名须以 'admin'（不区分大小写）开头，"
            "例如：admin收入订单明细-xxx.xlsx",
            INPUT_DIR_3,
        )
        return 1

    admin_frames = []
    for fp in files["admin"]:
        logging.info("读取 admin 文件: %s", fp.name)
        try:
            admin_frames.append(read_admin(fp))
        except Exception:
            logging.error("读取 admin 文件失败: %s", fp.name, exc_info=True)
            return 1
    admin_df = pd.concat(admin_frames, ignore_index=True) if len(admin_frames) > 1 else admin_frames[0]
    logging.info("  admin 共 %d 条记录", len(admin_df))

    adyen_lk = huawei_lk = google_lk = None
    google_raw_df = None

    if files["adyen"]:
        adyen_frames = []
        for fp in files["adyen"]:
            logging.info("读取 Adyen 文件: %s", fp.name)
            try:
                adyen_frames.append(read_adyen(fp))
            except Exception:
                logging.error("读取 Adyen 文件失败: %s", fp.name, exc_info=True)
        if adyen_frames:
            raw = pd.concat(adyen_frames, ignore_index=True) if len(adyen_frames) > 1 else adyen_frames[0]
            adyen_lk = build_adyen_lookup(raw)
            logging.info("  Adyen 去重后共 %d 个唯一 PSP Reference", len(adyen_lk))
    else:
        logging.warning("未找到 Adyen 文件，跳过 Adyen 匹配")

    if files["huawei"]:
        huawei_frames = []
        for fp in files["huawei"]:
            logging.info("读取华为文件: %s", fp.name)
            try:
                huawei_frames.append(read_huawei(fp))
            except Exception:
                logging.error("读取华为文件失败: %s", fp.name, exc_info=True)
        if huawei_frames:
            raw = pd.concat(huawei_frames, ignore_index=True) if len(huawei_frames) > 1 else huawei_frames[0]
            huawei_lk = build_huawei_lookup(raw)
            logging.info("  华为共 %d 条记录", len(huawei_lk))
    else:
        logging.warning("未找到华为文件，跳过华为匹配")

    huawei_settle_df = None
    if files["huawei_settlement"]:
        settle_frames = []
        for fp in files["huawei_settlement"]:
            logging.info("读取华为结算文件: %s", fp.name)
            try:
                settle_frames.append(read_huawei_settlement(fp))
            except Exception:
                logging.error("读取华为结算文件失败: %s", fp.name, exc_info=True)
        if settle_frames:
            huawei_settle_df = pd.concat(settle_frames, ignore_index=True) if len(settle_frames) > 1 else settle_frames[0]
            logging.info("  华为结算文件共 %d 行", len(huawei_settle_df))
    else:
        logging.info("未找到华为结算文件（华为平台结算-*.xlsx），手续费HKD行将不生成")

    if files["google"]:
        google_frames = []
        for fp in files["google"]:
            logging.info("读取 Google Play 文件: %s", fp.name)
            try:
                google_frames.append(read_google(fp))
            except Exception:
                logging.error("读取 Google Play 文件失败: %s", fp.name, exc_info=True)
        if google_frames:
            google_raw_df = pd.concat(google_frames, ignore_index=True) if len(google_frames) > 1 else google_frames[0]
            google_lk = build_google_lookup(google_raw_df)
            logging.info("  Google Play 共 %d 个唯一订单", len(google_lk))
    else:
        logging.warning("未找到 Google Play 文件，跳过 Google 匹配")

    apple_raw_df = None
    if files["apple"]:
        apple_frames = []
        for fp in files["apple"]:
            logging.info("读取苹果文件: %s", fp.name)
            try:
                apple_frames.append(read_apple(fp))
            except Exception:
                logging.error("读取苹果文件失败: %s", fp.name, exc_info=True)
        if apple_frames:
            apple_raw_df = pd.concat(apple_frames, ignore_index=True) if len(apple_frames) > 1 else apple_frames[0]
            logging.info("  苹果平台报表共 %d 行", len(apple_raw_df))
    else:
        logging.info("未找到苹果文件（苹果 Sheet 的平台数据部分将为空）")

    adyen_settle_df = None
    if files["adyen_settlement"]:
        settle_frames = []
        for fp in files["adyen_settlement"]:
            logging.info("读取 ADYEN 结算文件: %s", fp.name)
            try:
                settle_frames.append(read_adyen_settlement(fp))
            except Exception:
                logging.error("读取 ADYEN 结算文件失败: %s", fp.name, exc_info=True)
        if settle_frames:
            adyen_settle_df = pd.concat(settle_frames, ignore_index=True) if len(settle_frames) > 1 else settle_frames[0]
            logging.info("  ADYEN 结算文件共 %d 行", len(adyen_settle_df))
    else:
        logging.info("未找到 ADYEN 结算文件（ADYEN-settlement_*.xlsx），手续费将不从结算文件计算")

    result_df = enrich_admin(admin_df, adyen_lk, huawei_lk, google_lk)
    log_match_stats(result_df)

    try:
        output_path = write_output(result_df, OUTPUT_DIR, apple_raw_df, huawei_settle_df, adyen_settle_df, google_raw_df)
        logging.info("结果文件已写入: %s", output_path)
    except PermissionError:
        logging.error("无法写入输出文件，请确认文件未在 Excel 中打开后重试。")
        return 1
    except Exception:
        logging.error("写入输出文件失败", exc_info=True)
        return 1

    return 0
