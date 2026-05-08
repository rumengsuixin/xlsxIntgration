"""Balance extraction and balance-sheet row helpers."""

import calendar
import logging
import re
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel as excel_to_date

from .config import BALANCE_BANK_ORDER, BANK_BALANCE_COL, BANK_DATE_COL
from .config2 import BALANCE_BLOCK_SIZE_2, COMPANY_COL_START_2

FILL_MISSING_BALANCE_BANKS = {"招商银行", "中信银行", "中国银行"}


def get_last_balance(df, bank_name: str):
    """
    按日期找最新有效余额（同日取靠后行）。
    返回 ("YYYY-MM-DD", float) 或 (None, None)。
    """
    balance_col = BANK_BALANCE_COL.get(bank_name, "")
    date_col = BANK_DATE_COL.get(bank_name, "")

    if balance_col not in df.columns:
        logging.warning(f"  余额列 '{balance_col}' 不存在，跳过余额更新")
        return None, None
    if date_col not in df.columns:
        logging.warning(f"  日期列 '{date_col}' 不存在，跳过余额更新")
        return None, None

    latest_date = None
    latest_balance = None

    for _, row in df.iterrows():
        bal_str = str(row.get(balance_col, "")).strip().replace(",", "").replace("+", "")
        if not bal_str:
            continue
        try:
            balance = float(bal_str)
        except ValueError:
            continue
        if balance == 0.0:
            continue

        date_raw = str(row.get(date_col, "")).strip()
        row_date = _parse_date_str(date_raw)
        if row_date is not None:
            if latest_date is None or row_date >= latest_date:
                latest_date = row_date
                latest_balance = balance

    if latest_date is not None:
        return latest_date.isoformat(), latest_balance

    return None, None


def get_monthly_balances(
    df,
    bank_name: str,
    balance_col_map: Optional[Dict] = None,
    date_col_map: Optional[Dict] = None,
) -> List[Tuple[str, float]]:
    """
    按月提取期末余额，每月取最新日期行（同日取靠后行）。
    返回按日期升序的 [("YYYY-MM-DD", float)] 列表。

    balance_col_map / date_col_map 不传时使用代号1配置；
    代号2传入 BANK_BALANCE_COL_2 / BANK_DATE_COL_2。
    """
    _bal_map = balance_col_map or BANK_BALANCE_COL
    _date_map = date_col_map or BANK_DATE_COL

    balance_col = _bal_map.get(bank_name, "")
    date_col = _date_map.get(bank_name, "")

    if not balance_col:
        logging.info(f"  [{bank_name}] 无余额列配置，跳过余额提取")
        return []
    if not date_col:
        logging.info(f"  [{bank_name}] 无日期列配置，跳过余额提取")
        return []

    # 对汇丰银行等余额列名含前缀的银行做模糊匹配
    actual_balance_col = _resolve_col(df, balance_col)
    if actual_balance_col is None:
        logging.warning(f"  余额列 '{balance_col}' 不存在，跳过余额更新")
        return []
    if date_col not in df.columns:
        logging.warning(f"  日期列 '{date_col}' 不存在，跳过余额更新")
        return []

    month_best: dict = {}
    transactions = []

    for seq, (_, row) in enumerate(df.iterrows()):
        balance = _parse_amount(row.get(actual_balance_col, ""))
        if balance is None:
            continue
        if balance == 0.0:
            continue

        date_raw = str(row.get(date_col, "")).strip()
        row_date = _parse_date_str(date_raw)
        if row_date is None:
            continue

        month_key = (row_date.year, row_date.month)
        prev = month_best.get(month_key)
        if prev is None or row_date >= prev[0]:
            month_best[month_key] = (row_date, balance)

        amount = _extract_net_amount(row, bank_name)
        if amount is not None:
            transactions.append((row_date, seq, balance, amount))

    if balance_col_map is None and date_col_map is None:
        _fill_missing_monthly_balances(df, bank_name, month_best, transactions)

    return [
        (d.isoformat(), bal)
        for _, (d, bal) in sorted(month_best.items())
    ]


def _fill_missing_monthly_balances(df, bank_name: str, month_best: dict, transactions: list) -> None:
    if bank_name not in FILL_MISSING_BALANCE_BANKS:
        return

    date_range = df.attrs.get("statement_date_range")
    if not date_range:
        return
    start_date, end_date = date_range
    if start_date is None or end_date is None or start_date > end_date:
        return
    if not transactions:
        logging.info(f"  [{bank_name}] 无可推算交易行，跳过缺失月份余额补全")
        return

    transactions = sorted(transactions, key=lambda item: (item[0], item[1]))
    first_date, _, first_balance, first_amount = transactions[0]
    balance_before_first = first_balance - first_amount

    for year, month, month_end in _iter_month_ends(start_date, end_date):
        month_key = (year, month)
        if month_key in month_best:
            continue

        balance = None
        for row_date, _, row_balance, _ in transactions:
            if row_date <= month_end:
                balance = row_balance
            else:
                break

        if balance is None and month_end < first_date:
            balance = balance_before_first

        if balance is None:
            continue

        month_best[month_key] = (month_end, balance)
        logging.info(f"  [{bank_name}] 补全 {year}年{month}月 余额: {balance}")


def _iter_month_ends(start_date: date, end_date: date):
    year = start_date.year
    month = start_date.month
    while (year, month) <= (end_date.year, end_date.month):
        last_day = calendar.monthrange(year, month)[1]
        yield year, month, date(year, month, last_day)
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1


def _extract_net_amount(row, bank_name: str) -> Optional[float]:
    if bank_name == "招商银行":
        debit = _parse_amount(row.get("借方金额", "")) or 0.0
        credit = _parse_amount(row.get("贷方金额", "")) or 0.0
        return credit - debit
    if bank_name == "中信银行":
        debit = _parse_amount(row.get("借方发生额", "")) or 0.0
        credit = _parse_amount(row.get("贷方发生额", "")) or 0.0
        return credit - debit
    if bank_name == "中国银行":
        return _parse_amount(row.get("交易金额", ""))
    return None


def _parse_amount(value) -> Optional[float]:
    text = str(value or "").strip().replace(",", "").replace("\t", "").replace(" ", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _resolve_col(df, col_name: str) -> Optional[str]:
    """直接匹配列名，失败时尝试前缀匹配（用于汇丰银行等含动态后缀的列名）。"""
    if col_name in df.columns:
        return col_name
    for c in df.columns:
        if c.startswith(col_name):
            return c
    return None


def _parse_date_str(date_raw: str) -> Optional[date]:
    """
    支持多种日期格式：
      YYYY-MM-DD / YYYYMMDD / YYYY/MM/DD（以及含时间的变体）
      YYYY年M月D日（中文，含单位数月/日）
      DD/MM/YYYY（大华银行UOB格式）
    """
    if not date_raw:
        return None

    # 标准：YYYY-MM-DD 或 YYYYMMDD 或 ISO 8601
    m = re.search(r"(\d{4})[-/](\d{2})[-/](\d{2})", date_raw)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # YYYYMMDD（无分隔符）
    m = re.match(r"(\d{4})(\d{2})(\d{2})$", date_raw.strip()[:8])
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # 中文：YYYY年M月D日
    m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", date_raw)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # DD/MM/YYYY（大华银行）
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", date_raw.strip())
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    return None


def parse_cell_date(val) -> Optional[date]:
    """解析 openpyxl 单元格值为 Python date。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)) and val > 0:
        try:
            return excel_to_date(int(val)).date()
        except Exception:
            pass
    if isinstance(val, str):
        return _parse_date_str(val.strip())
    return None


def detect_balance_sheet_year(ws) -> Optional[int]:
    """扫描 A 列，返回找到的最大年份。"""
    max_year = None
    for row_idx in range(1, ws.max_row + 1):
        d = parse_cell_date(ws.cell(row=row_idx, column=1).value)
        if d is not None and (max_year is None or d.year > max_year):
            max_year = d.year
    return max_year


def collect_balance_month_blocks(ws) -> List[Tuple[int, date]]:
    """返回所有月份块的 (起始行, 月末日期) 列表。"""
    blocks = []
    for row_idx in range(1, ws.max_row + 1):
        d = parse_cell_date(ws.cell(row=row_idx, column=1).value)
        if d is not None:
            blocks.append((row_idx, d))
    return blocks


def build_target_dates(current_year: int) -> List[date]:
    """生成目标月末日期：上一年12月末 + 当年1-12月末。"""
    targets = [date(current_year - 1, 12, 31)]
    for month in range(1, 13):
        last_day = calendar.monthrange(current_year, month)[1]
        targets.append(date(current_year, month, last_day))
    return targets


def refresh_balance_sheet_dates(ws, current_year: int) -> None:
    """刷新代号1余额表日期，清空公司余额列（E-Z）。"""
    target_dates = build_target_dates(current_year)
    existing_blocks = collect_balance_month_blocks(ws)
    block_size = len(BALANCE_BANK_ORDER)
    last_company_col = max(_last_company_header_col(ws), 26)

    for i, target_date in enumerate(target_dates):
        if i < len(existing_blocks):
            block_start, _ = existing_blocks[i]
            ws.cell(row=block_start, column=1).value = target_date
            for row in range(block_start, block_start + block_size + 1):
                for col in range(5, last_company_col + 1):
                    ws.cell(row=row, column=col).value = None
        else:
            append_balance_month_block(ws, target_date, start_row=ws.max_row + 2)

    logging.info(f"银行余额表日期已更新至 {current_year} 年度（上一年12月末+当前年1-12月末）")


def refresh_balance_sheet_dates_2(ws, current_year: int) -> None:
    """刷新代号2余额表日期，清空公司余额列（G-U）。"""
    target_dates = build_target_dates(current_year)
    existing_blocks = collect_balance_month_blocks(ws)
    last_company_col = max(_last_company_header_col_2(ws), COMPANY_COL_START_2 + 14)

    for i, target_date in enumerate(target_dates):
        if i < len(existing_blocks):
            block_start, _ = existing_blocks[i]
            ws.cell(row=block_start, column=1).value = target_date
            for row in range(block_start, block_start + BALANCE_BLOCK_SIZE_2):
                for col in range(COMPANY_COL_START_2, last_company_col + 1):
                    ws.cell(row=row, column=col).value = None
        else:
            logging.warning(
                f"代号2余额表中未找到第 {i+1} 个月份块，跳过 {target_date} 日期更新"
            )

    logging.info(f"海外银行余额表日期已更新至 {current_year} 年度")


def append_balance_month_block(ws, month_end: date, start_row: Optional[int] = None) -> int:
    """追加代号1月末余额块，返回起始行。"""
    if start_row is None:
        start_row = ws.max_row + 1
    first_company_col = 5
    last_company_col = max(_last_company_header_col(ws), 26)

    for i, bk in enumerate(BALANCE_BANK_ORDER):
        row = start_row + i
        if i == 0:
            ws.cell(row=row, column=1).value = month_end
            ws.cell(row=row, column=2).value = "银行存款"
        ws.cell(row=row, column=3).value = bk
        ws.cell(row=row, column=4).value = (
            f"=SUM({get_column_letter(first_company_col)}{row}:"
            f"{get_column_letter(last_company_col)}{row})"
        )

    summary_row = start_row + len(BALANCE_BANK_ORDER)
    ws.cell(row=summary_row, column=2).value = "货币资金合计"
    bank_start = start_row
    bank_end = start_row + len(BALANCE_BANK_ORDER) - 1
    for col_idx in range(4, last_company_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=summary_row, column=col_idx).value = (
            f"=SUM({col_letter}{bank_start}:{col_letter}{bank_end})"
        )

    return start_row


def find_or_create_date_block(ws, year: int, month: int) -> int:
    """找代号1月份块起始行，不存在时追加新块。"""
    for row_idx in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row_idx, column=1).value
        if a_val is not None:
            cell_date = parse_cell_date(a_val)
            if cell_date is not None and cell_date.year == year and cell_date.month == month:
                return row_idx

    start_row = ws.max_row + 2
    last_day = calendar.monthrange(year, month)[1]
    month_end = date(year, month, last_day)
    append_balance_month_block(ws, month_end, start_row=start_row)
    summary_row = start_row + len(BALANCE_BANK_ORDER)

    logging.info(f"  银行余额表新增 {year}年{month}月 数据块（行 {start_row}-{summary_row}）")
    return start_row


def find_date_block_2(ws, year: int, month: int) -> int:
    """找代号2月份块起始行（扫描 A 列），不存在时返回 -1 并警告。"""
    for row_idx in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row_idx, column=1).value
        if a_val is not None:
            cell_date = parse_cell_date(a_val)
            if cell_date is not None and cell_date.year == year and cell_date.month == month:
                return row_idx
    logging.warning(f"  代号2余额表未找到 {year}年{month}月 数据块，跳过余额写入")
    return -1


def update_balance_sheet(ws, bank_name: str, company_code: str, date_str: str, balance: float) -> None:
    """更新代号1余额表单元格。"""
    m = re.match(r"(\d{4})-(\d{2})", date_str)
    if not m:
        logging.warning(f"  日期格式异常 '{date_str}'，跳过余额更新")
        return

    year, month = int(m.group(1)), int(m.group(2))
    block_start = find_or_create_date_block(ws, year, month)
    col_idx = find_or_append_company_column(ws, company_code)
    if col_idx is None:
        return

    target_row = None
    for row_idx in range(block_start, block_start + len(BALANCE_BANK_ORDER)):
        c_val = str(ws.cell(row=row_idx, column=3).value or "").strip()
        if c_val == bank_name:
            target_row = row_idx
            break

    if target_row is None:
        logging.warning(f"  银行余额块中未找到银行行 [{bank_name}]，跳过")
        return

    ws.cell(row=target_row, column=col_idx).value = balance
    logging.info(f"  银行余额已更新: {year}年{month}月 / {bank_name} / 公司{company_code} = {balance}")


def find_or_append_company_column(ws, company_code: str) -> Optional[int]:
    """Find a company column by header text, appending it when missing."""
    company = str(company_code or "").strip()
    if not company:
        logging.warning("  公司前缀为空，跳过余额更新")
        return None

    last_header_col = _last_company_header_col(ws)
    for col_idx in range(5, last_header_col + 1):
        header = ws.cell(row=1, column=col_idx).value
        if str(header or "").strip() == company:
            return col_idx

    new_col = max(last_header_col, 4) + 1
    ws.cell(row=1, column=new_col).value = company
    _refresh_balance_total_formulas(ws, new_col)
    logging.info(f"  银行余额表新增公司列: {company} ({get_column_letter(new_col)}列)")
    return new_col


def _last_company_header_col(ws) -> int:
    """Return the last non-empty company header column on row 1."""
    last_col = 4
    for col_idx in range(5, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if str(value or "").strip():
            last_col = col_idx
    return last_col


def _refresh_balance_total_formulas(ws, last_company_col: int) -> None:
    """Extend D-column bank totals to include all company columns."""
    if last_company_col < 5:
        return

    last_letter = get_column_letter(last_company_col)
    for row_idx in range(1, ws.max_row + 1):
        bank_name = str(ws.cell(row=row_idx, column=3).value or "").strip()
        if bank_name in BALANCE_BANK_ORDER:
            ws.cell(row=row_idx, column=4).value = f"=SUM(E{row_idx}:{last_letter}{row_idx})"


def update_balance_sheet_2(
    ws,
    bank_name: str,
    currency: str,
    company_code: str,
    date_str: str,
    balance: float,
) -> None:
    """
    更新代号2余额表单元格。

    余额表列结构（1-indexed）：
      A(1)=日期  B(2)=银行类型  C(3)=银行名  D(4)=折合人民币
      E(5)=合计（模板公式，不写入）  F(6)=币种  G(7)=公司A … U(21)=公司O

    D列写入 Excel 公式，由汇率工作表动态折算：
      =IFERROR(E行 * INDEX(汇率!$B:$F, 月份匹配, 币种列匹配), E行)
    IFERROR 回落到 E 列，适用于 CNY 等汇率表中无对应列的币种。
    """
    m = re.match(r"(\d{4})-(\d{2})", date_str)
    if not m:
        logging.warning(f"  日期格式异常 '{date_str}'，跳过余额更新")
        return

    year, month = int(m.group(1)), int(m.group(2))
    block_start = find_date_block_2(ws, year, month)
    if block_start == -1:
        return

    current_bank = ""
    target_row = None
    for row_idx in range(block_start, block_start + BALANCE_BLOCK_SIZE_2):
        c_val = str(ws.cell(row=row_idx, column=3).value or "").strip()
        if c_val:
            current_bank = c_val
        f_val = str(ws.cell(row=row_idx, column=6).value or "").strip()
        if current_bank == bank_name and f_val == currency:
            target_row = row_idx
            break

    if target_row is None:
        logging.warning(f"  余额表中未找到行 [{bank_name} {currency}]，跳过")
        return

    col_idx = find_or_append_company_column_2(ws, company_code)
    if col_idx is None:
        return

    ws.cell(row=target_row, column=col_idx).value = balance
    # D列写入公式：汇率表按月份+币种列头动态查找，IFERROR兜底（CNY等无对应列时等于E列）
    ws.cell(row=target_row, column=4).value = (
        f"=IFERROR(E{target_row}*INDEX(汇率!$B:$F,"
        f"MATCH($A${block_start},汇率!$A:$A,0),"
        f'MATCH(F{target_row}&"/CNY",汇率!$B$1:$F$1,0)),E{target_row})'
    )
    logging.info(
        f"  余额已更新: {year}年{month}月 / {bank_name} {currency} / 公司{company_code} = {balance}"
    )


def find_or_append_company_column_2(ws, company_code: str) -> Optional[int]:
    """Find a mode2 company column by header text, appending it when missing."""
    company = str(company_code or "").strip()
    if not company:
        logging.warning("  公司前缀为空，跳过余额更新")
        return None

    last_header_col = _last_company_header_col_2(ws)
    for col_idx in range(COMPANY_COL_START_2, last_header_col + 1):
        header = ws.cell(row=1, column=col_idx).value
        if str(header or "").strip() == company:
            return col_idx

    new_col = max(last_header_col, COMPANY_COL_START_2 - 1) + 1
    ws.cell(row=1, column=new_col).value = company
    _refresh_balance_total_formulas_2(ws, new_col)
    logging.info(f"  海外银行余额表新增公司列: {company} ({get_column_letter(new_col)}列)")
    return new_col


def _last_company_header_col_2(ws) -> int:
    """Return the last non-empty mode2 company header column on row 1."""
    last_col = COMPANY_COL_START_2 - 1
    for col_idx in range(COMPANY_COL_START_2, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if str(value or "").strip():
            last_col = col_idx
    return last_col


def _refresh_balance_total_formulas_2(ws, last_company_col: int) -> None:
    """Extend E-column mode2 totals to include all company columns."""
    if last_company_col < COMPANY_COL_START_2:
        return

    start_letter = get_column_letter(COMPANY_COL_START_2)
    last_letter = get_column_letter(last_company_col)
    for row_idx in range(2, ws.max_row + 1):
        currency = str(ws.cell(row=row_idx, column=6).value or "").strip()
        if currency:
            ws.cell(row=row_idx, column=5).value = (
                f"=SUM({start_letter}{row_idx}:{last_letter}{row_idx})"
            )
