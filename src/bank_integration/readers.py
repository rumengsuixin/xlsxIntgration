"""Readers for bank source files."""

import re
from datetime import date, datetime
from typing import Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from .config import BANK_DATE_COL, BANK_READ_CONFIG
from .pdf_daily_balance import read_huamei_daily_balance_pdf


def read_bank_file(
    filepath: str,
    bank_name: str,
    bank_read_config: Optional[Dict] = None,
    bank_date_col: Optional[Dict] = None,
) -> pd.DataFrame:
    """
    读取银行流水文件，返回清洗后的 DataFrame（所有列保留为字符串）。

    bank_read_config / bank_date_col 默认使用代号1配置；
    传入代号2配置字典时，自动应用额外的读取选项。

    额外选项（代号2用，写在 bank_read_config[bank_name] 中）：
      encoding         强制编码（覆盖默认的三次编码尝试）
      col_map          读取后重命名列 {旧名: 新名}
      strip_col_suffix_char  列名含该字符时截断至该字符之前（去空白）
      row_filter_col   只保留该列以 row_filter_prefix 开头的行
      row_filter_prefix 行过滤前缀
      row_filter_val   只保留该列去空白后等于此值的行
    """
    cfg = (bank_read_config or BANK_READ_CONFIG)[bank_name]
    date_col_map = bank_date_col or BANK_DATE_COL
    header = cfg["header"]
    date_range = None

    if cfg.get("is_pdf"):
        df = read_huamei_daily_balance_pdf(filepath)
    elif cfg["is_csv"]:
        if bank_read_config is None and bank_name == "中国银行":
            date_range = _extract_boc_date_range(filepath, header, cfg)
        df = _read_csv(filepath, header, cfg)
    else:
        if bank_read_config is None:
            date_range = _extract_xlsx_date_range(filepath, bank_name)
        df = pd.read_excel(
            filepath,
            header=header,
            dtype=str,
            engine=cfg["engine"],
        )

    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    df = df.fillna("")

    # --- 代号2额外处理 ---
    suffix_char = cfg.get("strip_col_suffix_char")
    if suffix_char:
        df.columns = [
            c.split(suffix_char)[0].strip() if suffix_char in c else c
            for c in df.columns
        ]

    col_map = cfg.get("col_map")
    if col_map:
        rename_map = {
            old: new
            for old, new in col_map.items()
            if old in df.columns and (new not in df.columns or old == new)
        }
        if rename_map:
            df = df.rename(columns=rename_map)

    row_filter_col = cfg.get("row_filter_col")
    if row_filter_col and row_filter_col in df.columns:
        prefix = cfg.get("row_filter_prefix")
        val = cfg.get("row_filter_val")
        if prefix is not None:
            df = df[df[row_filter_col].str.startswith(prefix, na=False)]
        elif val is not None:
            df = df[df[row_filter_col].str.strip() == val]

    # --- 代号1特殊处理：农业银行行过滤 ---
    if bank_read_config is None and bank_name == "农业银行":
        date_col = date_col_map.get(bank_name, "")
        if date_col in df.columns:
            df = df[df[date_col].str.match(r"\d{4}-\d{2}-\d{2}")]

    if date_range is not None:
        df.attrs["statement_date_range"] = date_range

    return df


def _read_csv(filepath: str, header, cfg: Dict) -> pd.DataFrame:
    forced_enc = cfg.get("encoding")
    encodings = [forced_enc] if forced_enc else ("utf-8-sig", "gbk", "utf-8")
    df = None
    for enc in encodings:
        try:
            df = pd.read_csv(filepath, header=header, dtype=str, encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    if df is None:
        raise ValueError(f"无法解码 CSV 文件（尝试编码: {encodings}）: {filepath}")
    # 去除列名中 [英文] 后缀（中国银行格式，其他银行无影响）
    df.columns = [c.split("[")[0].strip() for c in df.columns]
    return df


def _extract_xlsx_date_range(filepath: str, bank_name: str) -> Optional[Tuple[date, date]]:
    labels = {
        "招商银行": ("查询开始日期", "查询结束日期"),
        "中信银行": ("起始日期", "截止日期"),
    }.get(bank_name)
    if labels is None:
        return None

    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.active
        values = {}
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            cells = list(row)
            for idx, cell in enumerate(cells):
                label = str(cell or "").strip()
                if label in labels:
                    next_value = cells[idx + 1] if idx + 1 < len(cells) else None
                    parsed = _parse_date_value(next_value)
                    if parsed is not None:
                        values[label] = parsed
        start = values.get(labels[0])
        end = values.get(labels[1])
        if start is not None and end is not None:
            return start, end
        return None
    finally:
        wb.close()


def _extract_boc_date_range(filepath: str, header, cfg: Dict) -> Optional[Tuple[date, date]]:
    forced_enc = cfg.get("encoding")
    encodings = [forced_enc] if forced_enc else ("utf-8-sig", "gbk", "utf-8")
    for enc in encodings:
        try:
            rows = pd.read_csv(filepath, header=None, dtype=str, encoding=enc, nrows=header)
            break
        except UnicodeDecodeError:
            continue
    else:
        return None

    for _, row in rows.fillna("").iterrows():
        label = str(row.iloc[0] if len(row) else "").strip()
        if label.startswith("查询时间范围"):
            raw = str(row.iloc[1] if len(row) > 1 else "").strip()
            m = re.search(r"(\d{8})\s*-\s*(\d{8})", raw)
            if not m:
                return None
            start = _parse_date_value(m.group(1))
            end = _parse_date_value(m.group(2))
            if start is not None and end is not None:
                return start, end
    return None


def _parse_date_value(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if not m:
        m = re.search(r"(\d{4})(\d{2})(\d{2})", text)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None
