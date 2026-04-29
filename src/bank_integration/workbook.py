"""Workbook copy preparation and summary writing."""

import logging
import shutil
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Union

import openpyxl

from .balances import (
    detect_balance_sheet_year,
    refresh_balance_sheet_dates,
    refresh_balance_sheet_dates_2,
    update_balance_sheet,
    update_balance_sheet_2,
)
from .config import BALANCE_SHEET, OUTPUT_PATH, PROTECTED_SHEETS, SUMMARY_FILE, TEMPLATE_PATH
from .config2 import (
    BALANCE_SHEET_PREFIX_2,
    OUTPUT_PATH_2,
    PROTECTED_SHEETS_2,
    SUMMARY_FILE_2,
    TEMPLATE_PATH_2,
)


def _find_balance_sheet_2(wb) -> Optional[str]:
    """找代号2余额工作表（名称以 BALANCE_SHEET_PREFIX_2 开头）。"""
    for name in wb.sheetnames:
        if name.startswith(BALANCE_SHEET_PREFIX_2):
            return name
    return None


def prepare_work_copy(
    output_dir: Optional[Union[str, Path]] = None,
    template_path: Optional[Path] = None,
    output_path: Optional[Path] = None,
    summary_file: Optional[str] = None,
    balance_sheet: Optional[str] = None,
) -> Optional[str]:
    """
    准备工作副本（从模板复制），必要时刷新余额表年份。

    不传额外参数时行为与原来完全一致（代号1）。
    代号2传入 template_path_2 / output_path_2 / summary_file_2 等参数。
    """
    _template = template_path or TEMPLATE_PATH
    _summary = summary_file or SUMMARY_FILE
    _output = (
        Path(output_dir) / _summary
        if output_dir is not None
        else (output_path or OUTPUT_PATH)
    )
    _output.parent.mkdir(parents=True, exist_ok=True)
    current_year = date.today().year

    # --- 判断是否为代号2模式 ---
    is_mode2 = template_path is not None

    def copy_from_template() -> bool:
        if not _template.exists():
            logging.error(
                f"缺少模板文件: {_template}。请把 {_summary} 放入对应 template 文件夹后再运行。"
            )
            return False
        try:
            shutil.copy2(_template, _output)
            logging.info(f"已从模板复制工作副本: {_output}")
        except PermissionError:
            logging.error(f"无法覆盖工作副本，请先关闭 Excel 中打开的 {_summary}，然后重新运行。")
            return False
        return True

    def check_year(path: Path) -> Optional[int]:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except PermissionError:
            logging.error(f"无法读取 {_summary}，请先关闭 Excel 中打开的该文件，然后重新运行。")
            return -1
        try:
            sheet_name = (
                _find_balance_sheet_2(wb) if is_mode2 else balance_sheet or BALANCE_SHEET
            )
            if sheet_name and sheet_name in wb.sheetnames:
                return detect_balance_sheet_year(wb[sheet_name])
            return None
        finally:
            wb.close()

    needs_copy = False
    if not _output.exists():
        needs_copy = True
    else:
        year = check_year(_output)
        if year == -1:
            return None
        if year != current_year:
            logging.info(f"工作副本年份为 {year}，当前年份为 {current_year}，重新从模板复制")
            needs_copy = True

    if needs_copy:
        if not copy_from_template():
            return None
        try:
            wb = openpyxl.load_workbook(_output)
        except PermissionError:
            logging.error(f"无法打开刚复制的工作副本 {_summary}，请确认文件没有被 Excel 占用。")
            return None
        sheet_name = (
            _find_balance_sheet_2(wb) if is_mode2 else balance_sheet or BALANCE_SHEET
        )
        if sheet_name and sheet_name in wb.sheetnames:
            year = detect_balance_sheet_year(wb[sheet_name])
            if year != current_year:
                if is_mode2:
                    refresh_balance_sheet_dates_2(wb[sheet_name], current_year)
                else:
                    refresh_balance_sheet_dates(wb[sheet_name], current_year)
                wb.save(_output)
                logging.info(f"工作副本日期已更新至 {current_year} 年度")
        wb.close()

    return str(_output)


def write_all_to_summary(results: List[Dict], summary_path: str) -> None:
    """将代号1明细表和余额更新写入汇总文件。"""
    try:
        wb = openpyxl.load_workbook(summary_path)
    except PermissionError:
        logging.error(f"无法打开汇总文件，请先关闭 Excel 中打开的 {SUMMARY_FILE}，然后重新运行。")
        raise

    for item in results:
        sheet_name = item["sheet_name"]
        df = item["df"]

        if sheet_name in PROTECTED_SHEETS:
            logging.warning(f"  [{sheet_name}] 是受保护表，跳过写入")
            continue

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        ws.append(list(df.columns))
        for row_tuple in df.itertuples(index=False, name=None):
            ws.append([v if v != "" else None for v in row_tuple])

        logging.info(f"  [{sheet_name}] 写入 {len(df)} 行数据")

    if BALANCE_SHEET in wb.sheetnames:
        ws_bal = wb[BALANCE_SHEET]
        for item in results:
            for date_str, balance in item.get("monthly_balances", []):
                update_balance_sheet(
                    ws_bal,
                    item["bank_name"],
                    item["company_code"],
                    date_str,
                    balance,
                )
    else:
        logging.warning(f"汇总文件中未找到 '{BALANCE_SHEET}' 工作表，跳过余额更新")

    try:
        wb.save(summary_path)
        logging.info(f"汇总文件保存完成: {summary_path}")
    except PermissionError:
        logging.error(f"无法保存汇总文件，请先关闭 Excel 中打开的 {SUMMARY_FILE}，然后重新运行。")
        raise


def write_all_to_summary_2(results: List[Dict], summary_path: str) -> None:
    """将代号2明细表和余额更新写入汇总文件。"""
    try:
        wb = openpyxl.load_workbook(summary_path)
    except PermissionError:
        logging.error(f"无法打开汇总文件，请先关闭 Excel 中打开的 {SUMMARY_FILE_2}，然后重新运行。")
        raise

    for item in results:
        sheet_name = item["sheet_name"]
        df = item["df"]

        if sheet_name in PROTECTED_SHEETS_2:
            logging.warning(f"  [{sheet_name}] 是受保护表，跳过写入")
            continue

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        ws.append(list(df.columns))
        for row_tuple in df.itertuples(index=False, name=None):
            ws.append([v if v != "" else None for v in row_tuple])

        logging.info(f"  [{sheet_name}] 写入 {len(df)} 行数据")

    balance_sheet_name = _find_balance_sheet_2(wb)
    if balance_sheet_name:
        ws_bal = wb[balance_sheet_name]
        for item in results:
            currency = item.get("currency", "")
            for date_str, balance in item.get("monthly_balances", []):
                update_balance_sheet_2(
                    ws_bal,
                    item["bank_name"],
                    currency,
                    item["company_code"],
                    date_str,
                    balance,
                )
    else:
        logging.warning(f"汇总文件中未找到以 '{BALANCE_SHEET_PREFIX_2}' 开头的余额工作表，跳过余额更新")

    try:
        wb.save(summary_path)
        logging.info(f"汇总文件保存完成: {summary_path}")
    except PermissionError:
        logging.error(f"无法保存汇总文件，请先关闭 Excel 中打开的 {SUMMARY_FILE_2}，然后重新运行。")
        raise


def align_workbook_year(summary_path: str, source_year: int) -> None:
    """若代号1工作副本年份与源文件年份不一致，刷新至 source_year。"""
    try:
        wb = openpyxl.load_workbook(summary_path)
    except PermissionError:
        logging.error(f"无法打开汇总文件，请先关闭 Excel 中打开的 {SUMMARY_FILE}，然后重新运行。")
        raise

    if BALANCE_SHEET not in wb.sheetnames:
        wb.close()
        return

    ws = wb[BALANCE_SHEET]
    current_year = detect_balance_sheet_year(ws)

    if current_year != source_year:
        logging.info(
            f"工作副本年份为 {current_year}，源文件年份为 {source_year}，"
            f"刷新余额表至 {source_year} 年度"
        )
        refresh_balance_sheet_dates(ws, source_year)
        try:
            wb.save(summary_path)
            logging.info(f"工作副本已更新至 {source_year} 年度")
        except PermissionError:
            logging.error(f"无法保存汇总文件，请先关闭 Excel 中打开的 {SUMMARY_FILE}，然后重新运行。")
            wb.close()
            raise

    wb.close()


def align_workbook_year_2(summary_path: str, source_year: int) -> None:
    """若代号2工作副本年份与源文件年份不一致，刷新至 source_year。"""
    try:
        wb = openpyxl.load_workbook(summary_path)
    except PermissionError:
        logging.error(f"无法打开汇总文件，请先关闭 Excel 中打开的 {SUMMARY_FILE_2}，然后重新运行。")
        raise

    balance_sheet_name = _find_balance_sheet_2(wb)
    if not balance_sheet_name:
        wb.close()
        return

    ws = wb[balance_sheet_name]
    current_year = detect_balance_sheet_year(ws)

    if current_year != source_year:
        logging.info(
            f"工作副本年份为 {current_year}，源文件年份为 {source_year}，"
            f"刷新海外余额表至 {source_year} 年度"
        )
        refresh_balance_sheet_dates_2(ws, source_year)
        try:
            wb.save(summary_path)
            logging.info(f"工作副本已更新至 {source_year} 年度")
        except PermissionError:
            logging.error(f"无法保存汇总文件，请先关闭 Excel 中打开的 {SUMMARY_FILE_2}，然后重新运行。")
            wb.close()
            raise

    wb.close()
