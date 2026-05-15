# -*- coding: utf-8 -*-
import json
import unittest
import tempfile
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.bank_integration.balances import (
    get_last_balance,
    get_monthly_balances,
    update_balance_sheet,
    update_balance_sheet_2,
)
from src.bank_integration.config2 import (
    BANK_BALANCE_COL_2,
    BANK_DATE_COL_2,
    BANK_READ_CONFIG_2,
)
from src.bank_integration.app3 import (
    _format_date,
    build_adyen_lookup,
    build_apple_platform_summary,
    build_google_lookup,
    build_monthly_comparison,
    build_summary_sheet,
    enrich_admin,
    read_admin,
    read_adyen,
    read_adyen_settlement,
    read_apple,
    read_google,
    read_huawei,
    read_huawei_settlement,
    scan_source_files_3,
    write_output,
)
from src.bank_integration.app4 import (
    build_chrome_args,
    build_export_url,
    build_export_urls,
    chunk_list,
    configure_chrome_downloads,
    expected_export_stem,
    export_batches_with_retries,
    get_previous_month_range,
    has_chrome_cookie_store,
    missing_export_dates,
    parse_date_args,
    parse_date,
)
from src.bank_integration.config3 import (
    ADMIN_AMOUNT_COL,
    ADMIN_DATE_COL,
    ADMIN_JOIN_COL,
    ADMIN_PAYMENT_COL,
    ADMIN_REFUND_COL,
    ADMIN_SHEET,
    ADYEN_AMOUNT_COL,
    ADYEN_CURRENCY_COL,
    ADYEN_DATE_COL,
    ADYEN_INTERCHANGE_COL,
    ADYEN_JOIN_COL,
    ADYEN_MARKUP_COL,
    ADYEN_PAYABLE_COL,
    ADYEN_RECORD_TYPE_COL,
    ADYEN_SCHEME_FEES_COL,
    ADYEN_SETTLE_AMOUNT_COL,
    ADYEN_SETTLE_CURRENCY_COL,
    ADYEN_SETTLE_DATE_COL,
    ADYEN_SETTLE_HEADER,
    ADYEN_SETTLE_JOURNAL_COL,
    ADYEN_SETTLEMENT_CURRENCY_COL,
    COUNTRY_TAX_COL,
    GOOGLE_CHARGE_TYPE,
    GOOGLE_DATE_COL,
    GOOGLE_BUYER_AMOUNT_COL,
    GOOGLE_BUYER_CURRENCY_COL,
    GOOGLE_CONVERSION_RATE_COL,
    GOOGLE_FEE_REFUND_TYPE,
    GOOGLE_FEE_TYPE,
    GOOGLE_MERCHANT_AMOUNT_COL,
    GOOGLE_JOIN_COL,
    GOOGLE_MERCHANT_CURRENCY_COL,
    GOOGLE_REFUND_TYPE,
    GOOGLE_TRANSACTION_TYPE_COL,
    HUAWEI_AMOUNT_COL,
    HUAWEI_CURRENCY_COL,
    HUAWEI_DATE_COL,
    HUAWEI_JOIN_COL,
    HUAWEI_SETTLE_AMOUNT_COL,
    HUAWEI_SETTLE_CURRENCY_COL,
    HUAWEI_SETTLE_DATE_COL,
    MATCH_STATUS_COL,
    ORIGINAL_CHARGE_AMOUNT_COL,
    OUTPUT_DIFF_SHEET_3,
    OUTPUT_FAILED_SHEET_3,
    OUTPUT_SHEET_3,
    OUTPUT_APPLE_SHEET_3,
    OUTPUT_SUMMARY_SHEET_3,
    PLATFORM_AMOUNT_COL,
    PLATFORM_CURRENCY_COL,
    SETTLEMENT_CURRENCY_COL,
    STATUS_COL,
    TRANSACTION_DATE_COL,
)
from src.bank_integration.config4 import EXPORT_BATCH_WAIT_SECONDS_4
from src.bank_integration.readers import read_bank_file
from src.bank_integration.scanner import scan_source_files, scan_source_files_2


ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "data" / "input" / "1"
INPUT_DIR_2 = ROOT / "data" / "input" / "2"
HUAMEI_PDF = ROOT / "华美银行电子对账单-2025.02.pdf"


class BankIntegrationSampleTests(unittest.TestCase):
    def test_format_date_normalizes_platform_dates(self):
        cases = {
            "Jan 1, 2026": "2026-01-01",
            "2026-01-30 12:34:56": "2026-01-30",
            "2026-01-30": "2026-01-30",
            "": "",
            "not a date": "",
        }
        for raw, expected in cases.items():
            with self.subTest(raw=raw):
                self.assertEqual(_format_date(raw), expected)

    def _adyen_df(self, rows):
        defaults = {
            ADYEN_AMOUNT_COL: "10.00",
            ADYEN_CURRENCY_COL: "USD",
            ADYEN_SETTLEMENT_CURRENCY_COL: "HKD",
            ADYEN_PAYABLE_COL: "9.50",
            ADYEN_MARKUP_COL: "0.10",
            ADYEN_SCHEME_FEES_COL: "0.20",
            ADYEN_INTERCHANGE_COL: "0.20",
            ADYEN_DATE_COL: "2026-03-01",
        }
        return pd.DataFrame([{**defaults, **row} for row in rows])

    def _admin_df(self, psp):
        return pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: psp,
                    ADMIN_AMOUNT_COL: "10.00",
                    ADMIN_PAYMENT_COL: "Adyen",
                    ADMIN_REFUND_COL: "正常",
                    ADMIN_DATE_COL: "2026-03-01 12:00:00",
                }
            ]
        )

    def _write_admin_workbook(self, path, sheets):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _write_workbook(self, path, sheets, header=0):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, startrow=header, index=False)

    def _write_csv(self, path, df, header=0, encoding="utf-8-sig"):
        prefix = ""
        if header:
            filler = ",".join([f"说明{i}" for i in range(len(df.columns))])
            prefix = "\n".join([filler] * header) + "\n"
        path.write_text(prefix + df.to_csv(index=False), encoding=encoding)

    def _platform_sheet_cases(self):
        return [
            (
                "adyen",
                read_adyen,
                "Data",
                0,
                pd.DataFrame({ADYEN_JOIN_COL: ["PSP-FALLBACK"], ADYEN_RECORD_TYPE_COL: ["SentForSettle"]}),
                ADYEN_JOIN_COL,
                "PSP-FALLBACK",
            ),
            (
                "adyen settlement",
                read_adyen_settlement,
                "Data",
                ADYEN_SETTLE_HEADER,
                pd.DataFrame(
                    {
                        ADYEN_SETTLE_JOURNAL_COL: ["MerchantPayout"],
                        ADYEN_SETTLE_DATE_COL: ["2026-03-31"],
                        ADYEN_SETTLE_CURRENCY_COL: ["HKD"],
                        ADYEN_SETTLE_AMOUNT_COL: ["10.00"],
                    }
                ),
                ADYEN_SETTLE_JOURNAL_COL,
                "MerchantPayout",
            ),
            (
                "huawei",
                read_huawei,
                "Sheet0",
                0,
                pd.DataFrame(
                    {
                        HUAWEI_JOIN_COL: ["HUAWEI-FALLBACK"],
                        HUAWEI_AMOUNT_COL: ["10.00"],
                        HUAWEI_CURRENCY_COL: ["HKD"],
                    }
                ),
                HUAWEI_JOIN_COL,
                "HUAWEI-FALLBACK",
            ),
            (
                "huawei settlement",
                read_huawei_settlement,
                0,
                1,
                pd.DataFrame(
                    {
                        HUAWEI_SETTLE_DATE_COL: ["202603"],
                        HUAWEI_SETTLE_AMOUNT_COL: ["10.00"],
                        HUAWEI_SETTLE_CURRENCY_COL: ["HKD"],
                    }
                ),
                HUAWEI_SETTLE_DATE_COL,
                "202603",
            ),
            (
                "google",
                read_google,
                0,
                0,
                pd.DataFrame(
                    {
                        GOOGLE_JOIN_COL: ["GPA.123"],
                        GOOGLE_TRANSACTION_TYPE_COL: ["Charge"],
                        GOOGLE_BUYER_AMOUNT_COL: ["10.00"],
                    }
                ),
                GOOGLE_JOIN_COL,
                "GPA.123",
            ),
            (
                "apple",
                read_apple,
                0,
                3,
                pd.DataFrame(
                    {
                        "Settlement Date": ["2026-03-01"],
                        "Extended Partner Share": ["8.00"],
                    }
                ),
                "Extended Partner Share",
                "8.00",
            ),
        ]

    def test_read_admin_uses_configured_summary_sheet_when_present(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "admin.xlsx"
            self._write_admin_workbook(
                path,
                {
                    "Other": pd.DataFrame({"not_join": ["skip"]}),
                    ADMIN_SHEET: self._admin_df("PSP-PREFERRED"),
                },
            )

            result = read_admin(path)

        self.assertEqual(result.loc[0, ADMIN_JOIN_COL], "PSP-PREFERRED")

    def test_read_admin_falls_back_to_sheet_with_join_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "admin.xlsx"
            self._write_admin_workbook(
                path,
                {
                    "Readme": pd.DataFrame({"not_join": ["skip"]}),
                    "Orders": self._admin_df("PSP-FALLBACK"),
                },
            )

            result = read_admin(path)

        self.assertEqual(result.loc[0, ADMIN_JOIN_COL], "PSP-FALLBACK")

    def test_read_admin_error_lists_sheets_when_no_candidate_sheet_exists(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "admin.xlsx"
            self._write_admin_workbook(
                path,
                {
                    "Readme": pd.DataFrame({"not_join": ["skip"]}),
                    "Orders": pd.DataFrame({"also_not_join": ["skip"]}),
                },
            )

            with self.assertRaisesRegex(ValueError, "Readme.*Orders"):
                read_admin(path)

    def test_read_admin_falls_back_when_configured_sheet_has_wrong_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "admin.xlsx"
            self._write_admin_workbook(
                path,
                {
                    ADMIN_SHEET: pd.DataFrame({"not_join": ["skip"]}),
                    "Orders": self._admin_df("PSP-FALLBACK"),
                },
            )

            result = read_admin(path)

        self.assertEqual(result.loc[0, ADMIN_JOIN_COL], "PSP-FALLBACK")

    def test_platform_readers_fall_back_when_default_sheet_missing(self):
        for label, reader, default_sheet, header, data, expected_col, expected_val in self._platform_sheet_cases():
            if not isinstance(default_sheet, str):
                continue
            with self.subTest(platform=label), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / f"{label}.xlsx"
                self._write_workbook(
                    path,
                    {
                        "Readme": pd.DataFrame({"not_expected": ["skip"]}),
                        "Actual": data,
                    },
                    header=header,
                )

                result = reader(path)

            self.assertEqual(result.loc[0, expected_col], expected_val)

    def test_platform_readers_fall_back_when_default_sheet_has_wrong_columns(self):
        for label, reader, default_sheet, header, data, expected_col, expected_val in self._platform_sheet_cases():
            with self.subTest(platform=label), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / f"{label}.xlsx"
                first_sheet = "Data" if default_sheet == "Data" else "Sheet0" if default_sheet == "Sheet0" else "Readme"
                self._write_workbook(
                    path,
                    {
                        first_sheet: pd.DataFrame({"not_expected": ["skip"]}),
                        "Actual": data,
                    },
                    header=header,
                )

                result = reader(path)

            self.assertEqual(result.loc[0, expected_col], expected_val)

    def test_platform_reader_error_lists_sheets_and_required_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "adyen.xlsx"
            self._write_workbook(
                path,
                {
                    "Data": pd.DataFrame({"not_expected": ["skip"]}),
                    "Readme": pd.DataFrame({"also_not_expected": ["skip"]}),
                },
            )

            with self.assertRaises(ValueError) as ctx:
                read_adyen(path)
            msg = str(ctx.exception)
            for expected in ("adyen.xlsx", "Data", "Readme", "Psp Reference", "Record Type"):
                self.assertIn(expected, msg)

    def test_scan_source_files_3_detects_google_csv_case_insensitive(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            (tmp_path / "Google-PlayApps_20260301.csv").write_text("", encoding="utf-8")
            (tmp_path / "admin-orders.xlsx").write_text("", encoding="utf-8")

            files = scan_source_files_3(tmp_path)

        self.assertEqual([path.name for path in files["google"]], ["Google-PlayApps_20260301.csv"])
        self.assertEqual([path.name for path in files["admin"]], ["admin-orders.xlsx"])

    def test_scan_source_files_3_detects_all_csv_sources(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            for name in (
                "admin-orders.csv",
                "adyen-orders.csv",
                "adyen-settlement-202603.csv",
                "华为订单.csv",
                "华为平台结算-202603.csv",
                "Google-PlayApps_20260301.csv",
                "苹果-202603.csv",
                "ignore.txt",
            ):
                (tmp_path / name).write_text("", encoding="utf-8")

            files = scan_source_files_3(tmp_path)

        self.assertEqual([path.name for path in files["admin"]], ["admin-orders.csv"])
        self.assertEqual([path.name for path in files["adyen"]], ["adyen-orders.csv"])
        self.assertEqual([path.name for path in files["adyen_settlement"]], ["adyen-settlement-202603.csv"])
        self.assertEqual([path.name for path in files["huawei"]], ["华为订单.csv"])
        self.assertEqual([path.name for path in files["huawei_settlement"]], ["华为平台结算-202603.csv"])
        self.assertEqual([path.name for path in files["google"]], ["Google-PlayApps_20260301.csv"])
        self.assertEqual([path.name for path in files["apple"]], ["苹果-202603.csv"])

    def test_read_google_supports_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "Google-PlayApps_20260301.csv"
            path.write_text(
                "Description,Transaction Type,Amount (Buyer Currency)\n"
                "GPA.123,Charge,10.00\n",
                encoding="utf-8-sig",
            )

            result = read_google(path)

        self.assertEqual(result.loc[0, GOOGLE_JOIN_COL], "GPA.123")
        self.assertEqual(result.loc[0, GOOGLE_TRANSACTION_TYPE_COL], "Charge")

    def test_platform_readers_support_csv(self):
        for label, reader, _default_sheet, header, data, expected_col, expected_val in self._platform_sheet_cases():
            with self.subTest(platform=label), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / f"{label.replace(' ', '-')}.csv"
                self._write_csv(path, data, header=header)

                result = reader(path)

            self.assertEqual(result.loc[0, expected_col], expected_val)

    def test_read_admin_supports_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "admin-orders.csv"
            self._write_csv(path, self._admin_df("PSP-CSV"))

            result = read_admin(path)

        self.assertEqual(result.loc[0, ADMIN_JOIN_COL], "PSP-CSV")

    def test_read_adyen_settlement_csv_detects_metadata_header(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "adyen-settlement.csv"
            path.write_text(
                "Merchant/Company,Start date,End date,Aggregation,Country codes\n"
                "MIG_InteractiveECOM,01/01/2026,10/05/2026,Store Terminal,\n"
                "Company Account,Merchant Account,Country Code,Store,Terminal ID,Batch Number,"
                "Batch Closed Date,Payment Method,Creation Date,TimeZone,Journal Type,Gross Currency,"
                "Gross Debit (GC),Gross Credit (GC),Exchange Rate,Net Currency,Net Debit (NC),"
                "Net Credit (NC),Bank/Card Commission (NC),DCC Markup (NC),Num Txs\n"
                "MIG_Interactive,MIG_InteractiveECOM,HK,,,762,2026-05-07 20:15:28.663,,"
                "2026-05-08,HKT,MerchantPayout,,,,1,USD,6984.78,,,,1\n",
                encoding="utf-8-sig",
            )

            result = read_adyen_settlement(path)

        self.assertEqual(result.loc[0, ADYEN_SETTLE_JOURNAL_COL], "MerchantPayout")
        self.assertEqual(result.loc[0, ADYEN_SETTLE_CURRENCY_COL], "USD")
        self.assertEqual(result.loc[0, ADYEN_SETTLE_AMOUNT_COL], "6984.78")

    def _adyen_status_for(self, psp, rows):
        lookup = build_adyen_lookup(self._adyen_df(rows))
        result = enrich_admin(self._admin_df(psp), lookup, None, None)
        return result.loc[0, STATUS_COL], result.loc[0, PLATFORM_AMOUNT_COL], lookup

    def test_adyen_refused_overrides_sent_for_settle(self):
        status, amount, lookup = self._adyen_status_for(
            "PSP-REFUSED-SETTLED",
            [
                {ADYEN_JOIN_COL: "PSP-REFUSED-SETTLED", ADYEN_RECORD_TYPE_COL: "SentForSettle"},
                {ADYEN_JOIN_COL: "PSP-REFUSED-SETTLED", ADYEN_RECORD_TYPE_COL: "Refused"},
            ],
        )

        self.assertNotIn("PSP-REFUSED-SETTLED", lookup.index)
        self.assertEqual(status, "失败")
        self.assertEqual(amount, "")

    def test_adyen_refused_overrides_authorised(self):
        status, amount, lookup = self._adyen_status_for(
            "PSP-REFUSED-AUTH",
            [
                {ADYEN_JOIN_COL: "PSP-REFUSED-AUTH", ADYEN_RECORD_TYPE_COL: "Authorised"},
                {ADYEN_JOIN_COL: "PSP-REFUSED-AUTH", ADYEN_RECORD_TYPE_COL: "Refused"},
            ],
        )

        self.assertNotIn("PSP-REFUSED-AUTH", lookup.index)
        self.assertEqual(status, "失败")
        self.assertEqual(amount, "")

    def test_adyen_received_and_refused_is_failed(self):
        status, amount, lookup = self._adyen_status_for(
            "PSP-RECEIVED-REFUSED",
            [
                {ADYEN_JOIN_COL: "PSP-RECEIVED-REFUSED", ADYEN_RECORD_TYPE_COL: "Received"},
                {ADYEN_JOIN_COL: "PSP-RECEIVED-REFUSED", ADYEN_RECORD_TYPE_COL: "Refused"},
            ],
        )

        self.assertNotIn("PSP-RECEIVED-REFUSED", lookup.index)
        self.assertEqual(status, "失败")
        self.assertEqual(amount, "")

    def test_adyen_success_types_still_match_without_refused(self):
        status, amount, lookup = self._adyen_status_for(
            "PSP-SUCCESS",
            [
                {ADYEN_JOIN_COL: "PSP-SUCCESS", ADYEN_RECORD_TYPE_COL: "Authorised", ADYEN_AMOUNT_COL: "8.00"},
                {ADYEN_JOIN_COL: "PSP-SUCCESS", ADYEN_RECORD_TYPE_COL: "SentForSettle", ADYEN_AMOUNT_COL: "10.00"},
            ],
        )

        self.assertIn("PSP-SUCCESS", lookup.index)
        self.assertEqual(lookup.at["PSP-SUCCESS", ADYEN_AMOUNT_COL], "10.00")
        self.assertEqual(status, "成功")
        self.assertEqual(amount, "10.00")

    def test_enrich_admin_adds_settlement_currency_after_platform_currency(self):
        lookup = build_adyen_lookup(
            self._adyen_df(
                [
                    {
                        ADYEN_JOIN_COL: "PSP-SETTLEMENT-CCY",
                        ADYEN_RECORD_TYPE_COL: "SentForSettle",
                        ADYEN_CURRENCY_COL: "TRY",
                        ADYEN_SETTLEMENT_CURRENCY_COL: "USD",
                    }
                ]
            )
        )

        result = enrich_admin(self._admin_df("PSP-SETTLEMENT-CCY"), lookup, None, None)
        columns = list(result.columns)

        self.assertEqual(result.loc[0, PLATFORM_CURRENCY_COL], "TRY")
        self.assertEqual(result.loc[0, SETTLEMENT_CURRENCY_COL], "USD")
        self.assertEqual(columns.index(ORIGINAL_CHARGE_AMOUNT_COL), columns.index(PLATFORM_AMOUNT_COL) - 1)
        self.assertEqual(columns.index(SETTLEMENT_CURRENCY_COL), columns.index(PLATFORM_CURRENCY_COL) + 1)
        self.assertEqual(columns.index(MATCH_STATUS_COL), columns.index(SETTLEMENT_CURRENCY_COL) + 1)
        self.assertEqual(columns.index(STATUS_COL), columns.index(MATCH_STATUS_COL) + 1)

    def test_enrich_admin_marks_match_status_for_admin_and_platform_only_rows(self):
        lookup = build_adyen_lookup(
            self._adyen_df(
                [
                    {ADYEN_JOIN_COL: "PSP-MATCH", ADYEN_RECORD_TYPE_COL: "SentForSettle"},
                    {ADYEN_JOIN_COL: "PSP-PLATFORM-ONLY", ADYEN_RECORD_TYPE_COL: "SentForSettle"},
                ]
            )
        )
        admin = pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: "PSP-MATCH",
                    ADMIN_AMOUNT_COL: "10.00",
                    ADMIN_PAYMENT_COL: "Adyen",
                    ADMIN_REFUND_COL: "正常",
                    ADMIN_DATE_COL: "2026-03-01 12:00:00",
                },
                {
                    ADMIN_JOIN_COL: "PSP-NO-MATCH",
                    ADMIN_AMOUNT_COL: "12.00",
                    ADMIN_PAYMENT_COL: "Adyen",
                    ADMIN_REFUND_COL: "正常",
                    ADMIN_DATE_COL: "2026-03-01 12:00:00",
                },
            ]
        )

        result = enrich_admin(admin, lookup, None, None)
        by_key = result.set_index(ADMIN_JOIN_COL)

        self.assertEqual(by_key.at["PSP-MATCH", MATCH_STATUS_COL], "是")
        self.assertEqual(by_key.at["PSP-NO-MATCH", MATCH_STATUS_COL], "否")
        self.assertEqual(by_key.at["PSP-PLATFORM-ONLY", MATCH_STATUS_COL], "平台多余")

    def test_enrich_admin_defaults_settlement_currency_for_huawei_and_google(self):
        admin = pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: "HUAWEI-1",
                    ADMIN_AMOUNT_COL: "10.00",
                    ADMIN_PAYMENT_COL: "华为支付",
                    ADMIN_REFUND_COL: "正常",
                    ADMIN_DATE_COL: "2026-03-01 12:00:00",
                },
                {
                    ADMIN_JOIN_COL: "GOOGLE-1",
                    ADMIN_AMOUNT_COL: "12.00",
                    ADMIN_PAYMENT_COL: "Google支付",
                    ADMIN_REFUND_COL: "正常",
                    ADMIN_DATE_COL: "2026-03-01 12:00:00",
                },
            ]
        )
        huawei_lk = pd.DataFrame(
            [
                {
                    HUAWEI_JOIN_COL: "HUAWEI-1",
                    HUAWEI_AMOUNT_COL: "10.00",
                    HUAWEI_CURRENCY_COL: "HKD",
                    HUAWEI_DATE_COL: "2026-03-01",
                }
            ]
        ).set_index(HUAWEI_JOIN_COL)
        google_lk = pd.DataFrame(
            [
                {
                    GOOGLE_JOIN_COL: "GOOGLE-1",
                    "charge_amt": "10.00",
                    "fee_amt": "-1.00",
                    "refund_amt": "",
                    "fee_refund_amt": "",
                    "merchant_charge_amt": "2.00",
                    "merchant_fee_amt": "-0.30",
                    "merchant_refund_amt": "",
                    "merchant_fee_refund_amt": "",
                    "conversion_rate": "0.2",
                    "ccy": "TRY",
                    "merchant_ccy": "HKD",
                    "transaction_date": "2026-03-01",
                }
            ]
        ).set_index(GOOGLE_JOIN_COL)

        result = enrich_admin(admin, None, huawei_lk, google_lk)

        self.assertEqual(result.loc[0, ORIGINAL_CHARGE_AMOUNT_COL], "")
        self.assertEqual(result.loc[0, PLATFORM_CURRENCY_COL], "HKD")
        self.assertEqual(result.loc[0, SETTLEMENT_CURRENCY_COL], "HKD")
        self.assertEqual(result.loc[1, PLATFORM_CURRENCY_COL], "TRY")
        self.assertEqual(result.loc[1, SETTLEMENT_CURRENCY_COL], "HKD")
        self.assertEqual(result.loc[1, ORIGINAL_CHARGE_AMOUNT_COL], "10.0")
        self.assertEqual(result.loc[1, PLATFORM_AMOUNT_COL], "12.0")
        self.assertEqual(result.loc[1, "结算金额"], "1.7")
        self.assertEqual(result.loc[1, "手续费"], "0.3")
        self.assertEqual(result.loc[1, COUNTRY_TAX_COL], "0.4")

    def test_google_lookup_uses_refund_merchant_currency_when_charge_missing(self):
        lookup = build_google_lookup(
            pd.DataFrame(
                [
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-REFUND-ONLY",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_REFUND_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "-10.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_CONVERSION_RATE_COL: "0.2",
                        GOOGLE_MERCHANT_AMOUNT_COL: "-2.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    }
                ]
            )
        )

        self.assertEqual(lookup.at["GOOGLE-REFUND-ONLY", "ccy"], "TRY")
        self.assertEqual(lookup.at["GOOGLE-REFUND-ONLY", "merchant_ccy"], "HKD")
        self.assertEqual(lookup.at["GOOGLE-REFUND-ONLY", "conversion_rate"], "0.2")
        self.assertEqual(lookup.at["GOOGLE-REFUND-ONLY", "merchant_refund_amt"], "-2.00")

    def test_google_refund_uses_merchant_amount_for_settlement(self):
        admin = pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: "GOOGLE-REFUND-ONLY",
                    ADMIN_AMOUNT_COL: "10.00",
                    ADMIN_PAYMENT_COL: "Google支付",
                    ADMIN_REFUND_COL: "已退款",
                    ADMIN_DATE_COL: "2026-03-01 12:00:00",
                }
            ]
        )
        lookup = build_google_lookup(
            pd.DataFrame(
                [
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-REFUND-ONLY",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_REFUND_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "-10.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_CONVERSION_RATE_COL: "0.2",
                        GOOGLE_MERCHANT_AMOUNT_COL: "-2.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-REFUND-ONLY",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_REFUND_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "1.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_MERCHANT_AMOUNT_COL: "0.30",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                ]
            )
        )

        result = enrich_admin(admin, None, None, lookup)

        self.assertEqual(result.loc[0, ORIGINAL_CHARGE_AMOUNT_COL], "10.0")
        self.assertEqual(result.loc[0, PLATFORM_AMOUNT_COL], "12.0")
        self.assertEqual(result.loc[0, PLATFORM_CURRENCY_COL], "TRY")
        self.assertEqual(result.loc[0, SETTLEMENT_CURRENCY_COL], "HKD")
        self.assertEqual(result.loc[0, "结算金额"], "1.7")
        self.assertEqual(result.loc[0, "手续费"], "0.3")
        self.assertEqual(result.loc[0, COUNTRY_TAX_COL], "0.4")

    def test_google_refund_summary_supplements_original_charge_net(self):
        admin = pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: "GOOGLE-REFUND",
                    ADMIN_AMOUNT_COL: "100.00",
                    ADMIN_PAYMENT_COL: "Google支付",
                    ADMIN_REFUND_COL: "已退款",
                    ADMIN_DATE_COL: "2026-02-01 12:00:00",
                },
                {
                    ADMIN_JOIN_COL: "GOOGLE-PENDING",
                    ADMIN_AMOUNT_COL: "50.00",
                    ADMIN_PAYMENT_COL: "Google支付",
                    ADMIN_REFUND_COL: "已退款",
                    ADMIN_DATE_COL: "2026-02-01 13:00:00",
                },
                {
                    ADMIN_JOIN_COL: "ADMIN-ONLY",
                    ADMIN_AMOUNT_COL: "10.00",
                    ADMIN_PAYMENT_COL: "Google支付",
                    ADMIN_REFUND_COL: "正常",
                    ADMIN_DATE_COL: "2026-01-01 10:00:00",
                },
            ]
        )
        lookup = build_google_lookup(
            pd.DataFrame(
                [
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-REFUND",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_CHARGE_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "100.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_CONVERSION_RATE_COL: "0.2",
                        GOOGLE_DATE_COL: "Jan 31, 2026",
                        GOOGLE_MERCHANT_AMOUNT_COL: "20.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-REFUND",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "-15.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_MERCHANT_AMOUNT_COL: "-3.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-REFUND",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_REFUND_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "-60.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_CONVERSION_RATE_COL: "0.2",
                        GOOGLE_DATE_COL: "Jan 31, 2026",
                        GOOGLE_MERCHANT_AMOUNT_COL: "-12.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-REFUND",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_REFUND_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "10.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_MERCHANT_AMOUNT_COL: "2.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-PENDING",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_CHARGE_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "30.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_CONVERSION_RATE_COL: "0.2",
                        GOOGLE_DATE_COL: "Jan 30, 2026",
                        GOOGLE_MERCHANT_AMOUNT_COL: "6.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-PENDING",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "-2.50",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_MERCHANT_AMOUNT_COL: "-0.50",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-PLATFORM-REFUND",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_CHARGE_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "25.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_CONVERSION_RATE_COL: "0.2",
                        GOOGLE_DATE_COL: "Jan 29, 2026",
                        GOOGLE_MERCHANT_AMOUNT_COL: "5.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-PLATFORM-REFUND",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "-1.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_MERCHANT_AMOUNT_COL: "-0.20",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-PLATFORM-REFUND",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_REFUND_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "-20.00",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_CONVERSION_RATE_COL: "0.2",
                        GOOGLE_DATE_COL: "Jan 29, 2026",
                        GOOGLE_MERCHANT_AMOUNT_COL: "-4.00",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                    {
                        GOOGLE_JOIN_COL: "GOOGLE-PLATFORM-REFUND",
                        GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_REFUND_TYPE,
                        GOOGLE_BUYER_AMOUNT_COL: "0.50",
                        GOOGLE_BUYER_CURRENCY_COL: "TRY",
                        GOOGLE_MERCHANT_AMOUNT_COL: "0.10",
                        GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    },
                ]
            )
        )

        result = enrich_admin(admin, None, None, lookup)
        summary = build_summary_sheet(result)

        self.assertEqual(
            result.loc[result[ADMIN_JOIN_COL] == "GOOGLE-REFUND", TRANSACTION_DATE_COL].iloc[0],
            "2026-01-31",
        )
        google_hkd = summary[
            (summary[TRANSACTION_DATE_COL] == "2026-01")
            & (summary[ADMIN_PAYMENT_COL] == "Google支付")
            & (summary[SETTLEMENT_CURRENCY_COL] == "HKD")
        ].iloc[0]
        self.assertEqual(google_hkd["成功笔数"], 0)
        self.assertEqual(google_hkd["退款笔数"], 2)
        self.assertEqual(google_hkd["退款待确认笔数"], 1)
        self.assertAlmostEqual(google_hkd["成功金额"], 27.30, places=2)
        self.assertAlmostEqual(google_hkd["退款金额"], 13.90, places=2)
        self.assertAlmostEqual(google_hkd["退款待确认金额"], 5.50, places=2)
        self.assertAlmostEqual(google_hkd["净交易金额"], 13.40, places=2)

    def test_google_cashflow_summary_keeps_cross_month_refund_in_refund_month(self):
        admin = pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: "GOOGLE-CROSS-MONTH",
                    ADMIN_AMOUNT_COL: "100.00",
                    ADMIN_PAYMENT_COL: "Google支付",
                    ADMIN_REFUND_COL: "已退款",
                    ADMIN_DATE_COL: "2026-02-05 12:00:00",
                }
            ]
        )
        google_raw = pd.DataFrame(
            [
                {
                    GOOGLE_JOIN_COL: "GOOGLE-CROSS-MONTH",
                    GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_CHARGE_TYPE,
                    GOOGLE_BUYER_AMOUNT_COL: "100.00",
                    GOOGLE_BUYER_CURRENCY_COL: "TRY",
                    GOOGLE_CONVERSION_RATE_COL: "0.2",
                    GOOGLE_DATE_COL: "Jan 31, 2026",
                    GOOGLE_MERCHANT_AMOUNT_COL: "20.00",
                    GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                },
                {
                    GOOGLE_JOIN_COL: "GOOGLE-CROSS-MONTH",
                    GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_TYPE,
                    GOOGLE_BUYER_AMOUNT_COL: "-15.00",
                    GOOGLE_BUYER_CURRENCY_COL: "TRY",
                    GOOGLE_MERCHANT_AMOUNT_COL: "-3.00",
                    GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    GOOGLE_DATE_COL: "Jan 31, 2026",
                },
                {
                    GOOGLE_JOIN_COL: "GOOGLE-CROSS-MONTH",
                    GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_REFUND_TYPE,
                    GOOGLE_BUYER_AMOUNT_COL: "-50.00",
                    GOOGLE_BUYER_CURRENCY_COL: "TRY",
                    GOOGLE_CONVERSION_RATE_COL: "0.2",
                    GOOGLE_DATE_COL: "Feb 3, 2026",
                    GOOGLE_MERCHANT_AMOUNT_COL: "-10.00",
                    GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                },
                {
                    GOOGLE_JOIN_COL: "GOOGLE-CROSS-MONTH",
                    GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_REFUND_TYPE,
                    GOOGLE_BUYER_AMOUNT_COL: "7.50",
                    GOOGLE_BUYER_CURRENCY_COL: "TRY",
                    GOOGLE_MERCHANT_AMOUNT_COL: "1.50",
                    GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    GOOGLE_DATE_COL: "Feb 3, 2026",
                },
            ]
        )

        result = enrich_admin(admin, None, None, build_google_lookup(google_raw))
        summary = build_summary_sheet(result, google_raw_df=google_raw)

        jan = summary[
            (summary[TRANSACTION_DATE_COL] == "2026-01")
            & (summary[ADMIN_PAYMENT_COL] == "Google支付")
            & (summary[SETTLEMENT_CURRENCY_COL] == "HKD")
        ].iloc[0]
        feb = summary[
            (summary[TRANSACTION_DATE_COL] == "2026-02")
            & (summary[ADMIN_PAYMENT_COL] == "Google支付")
            & (summary[SETTLEMENT_CURRENCY_COL] == "HKD")
        ].iloc[0]

        self.assertEqual(jan["成功笔数"], 1)
        self.assertEqual(jan["退款笔数"], 0)
        self.assertAlmostEqual(jan["成功金额"], 17.00, places=2)
        self.assertAlmostEqual(jan["退款金额"], 0.00, places=2)
        self.assertAlmostEqual(jan["净交易金额"], 17.00, places=2)
        self.assertEqual(feb["成功笔数"], 0)
        self.assertEqual(feb["退款笔数"], 1)
        self.assertAlmostEqual(feb["成功金额"], 0.00, places=2)
        self.assertAlmostEqual(feb["退款金额"], 8.50, places=2)
        self.assertAlmostEqual(feb["净交易金额"], -8.50, places=2)

    def test_google_cashflow_summary_keeps_single_month_refund_net_amount(self):
        admin = pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: "GOOGLE-SAME-MONTH",
                    ADMIN_AMOUNT_COL: "100.00",
                    ADMIN_PAYMENT_COL: "Google支付",
                    ADMIN_REFUND_COL: "已退款",
                    ADMIN_DATE_COL: "2026-01-20 12:00:00",
                }
            ]
        )
        google_raw = pd.DataFrame(
            [
                {
                    GOOGLE_JOIN_COL: "GOOGLE-SAME-MONTH",
                    GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_CHARGE_TYPE,
                    GOOGLE_BUYER_AMOUNT_COL: "100.00",
                    GOOGLE_BUYER_CURRENCY_COL: "TRY",
                    GOOGLE_CONVERSION_RATE_COL: "0.2",
                    GOOGLE_DATE_COL: "Jan 10, 2026",
                    GOOGLE_MERCHANT_AMOUNT_COL: "20.00",
                    GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                },
                {
                    GOOGLE_JOIN_COL: "GOOGLE-SAME-MONTH",
                    GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_TYPE,
                    GOOGLE_BUYER_AMOUNT_COL: "-15.00",
                    GOOGLE_BUYER_CURRENCY_COL: "TRY",
                    GOOGLE_MERCHANT_AMOUNT_COL: "-3.00",
                    GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    GOOGLE_DATE_COL: "Jan 10, 2026",
                },
                {
                    GOOGLE_JOIN_COL: "GOOGLE-SAME-MONTH",
                    GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_REFUND_TYPE,
                    GOOGLE_BUYER_AMOUNT_COL: "-60.00",
                    GOOGLE_BUYER_CURRENCY_COL: "TRY",
                    GOOGLE_CONVERSION_RATE_COL: "0.2",
                    GOOGLE_DATE_COL: "Jan 25, 2026",
                    GOOGLE_MERCHANT_AMOUNT_COL: "-12.00",
                    GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                },
                {
                    GOOGLE_JOIN_COL: "GOOGLE-SAME-MONTH",
                    GOOGLE_TRANSACTION_TYPE_COL: GOOGLE_FEE_REFUND_TYPE,
                    GOOGLE_BUYER_AMOUNT_COL: "10.00",
                    GOOGLE_BUYER_CURRENCY_COL: "TRY",
                    GOOGLE_MERCHANT_AMOUNT_COL: "2.00",
                    GOOGLE_MERCHANT_CURRENCY_COL: "HKD",
                    GOOGLE_DATE_COL: "Jan 25, 2026",
                },
            ]
        )

        result = enrich_admin(admin, None, None, build_google_lookup(google_raw))
        summary = build_summary_sheet(result, google_raw_df=google_raw)

        google_hkd = summary[
            (summary[TRANSACTION_DATE_COL] == "2026-01")
            & (summary[ADMIN_PAYMENT_COL] == "Google支付")
            & (summary[SETTLEMENT_CURRENCY_COL] == "HKD")
        ].iloc[0]
        self.assertEqual(google_hkd["成功笔数"], 1)
        self.assertEqual(google_hkd["退款笔数"], 1)
        self.assertEqual(google_hkd["退款待确认笔数"], 0)
        self.assertAlmostEqual(google_hkd["成功金额"], 17.00, places=2)
        self.assertAlmostEqual(google_hkd["退款金额"], 10.00, places=2)
        self.assertAlmostEqual(google_hkd["净交易金额"], 7.00, places=2)

    def test_google_platform_only_uses_buyer_amount_total_for_platform_amount(self):
        admin = pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: "GOOGLE-ADMIN-ONLY",
                    ADMIN_AMOUNT_COL: "10.00",
                    ADMIN_PAYMENT_COL: "Google支付",
                    ADMIN_REFUND_COL: "正常",
                    ADMIN_DATE_COL: "2026-03-01 12:00:00",
                }
            ]
        )
        google_lk = pd.DataFrame(
            [
                {
                    GOOGLE_JOIN_COL: "GOOGLE-PLATFORM-ONLY",
                    "charge_amt": "10.00",
                    "fee_amt": "-1.00",
                    "refund_amt": "",
                    "fee_refund_amt": "",
                    "merchant_charge_amt": "2.00",
                    "merchant_fee_amt": "-0.30",
                    "merchant_refund_amt": "",
                    "merchant_fee_refund_amt": "",
                    "conversion_rate": "0.2",
                    "ccy": "TRY",
                    "merchant_ccy": "HKD",
                    "transaction_date": "2026-03-01",
                }
            ]
        ).set_index(GOOGLE_JOIN_COL)

        result = enrich_admin(admin, None, None, google_lk)
        extra = result[result[ADMIN_JOIN_COL] == "GOOGLE-PLATFORM-ONLY"].iloc[0]

        self.assertEqual(extra[MATCH_STATUS_COL], "平台多余")
        self.assertEqual(extra[ORIGINAL_CHARGE_AMOUNT_COL], "10.0")
        self.assertEqual(extra[PLATFORM_AMOUNT_COL], "12.0")
        self.assertEqual(extra[COUNTRY_TAX_COL], "0.4")

    def test_build_summary_sheet_groups_platform_amounts(self):
        detail = pd.DataFrame(
            [
                {
                    ADMIN_AMOUNT_COL: "100.00",
                    ADMIN_DATE_COL: "2026-03-01 10:00:00",
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "100.50",
                    "结算金额": "10.50",
                    "手续费": "-1.00",
                },
                {
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "200.25",
                    "结算金额": "20.25",
                    "手续费": "-2.00",
                },
                {
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    STATUS_COL: "退款",
                    PLATFORM_AMOUNT_COL: "50.00",
                    "结算金额": "5.00",
                    "手续费": "0.50",
                },
                {
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    STATUS_COL: "失败",
                    PLATFORM_AMOUNT_COL: "999.00",
                    "结算金额": "999.00",
                    "手续费": "-99.00",
                },
                {
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "HKD",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "70.00",
                    "结算金额": "7.00",
                    "手续费": "-0.70",
                },
                {
                    TRANSACTION_DATE_COL: "2026-03-02",
                    ADMIN_PAYMENT_COL: "Google支付",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "HKD",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "80.00",
                    "结算金额": "8.00",
                    "手续费": "-0.80",
                },
                {
                    TRANSACTION_DATE_COL: "2026-03-02",
                    ADMIN_PAYMENT_COL: "Google支付",
                    PLATFORM_CURRENCY_COL: "USD",
                    SETTLEMENT_CURRENCY_COL: "HKD",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "999.00",
                    "结算金额": "2.00",
                    "手续费": "-0.20",
                },
                {
                    TRANSACTION_DATE_COL: "2026-03-02",
                    ADMIN_PAYMENT_COL: "Google支付",
                    PLATFORM_CURRENCY_COL: "USD",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "20.00",
                    "结算金额": "2.00",
                    "手续费": "-0.20",
                },
            ]
        )

        summary = build_summary_sheet(detail)

        self.assertEqual(len(summary), 4)
        self.assertIn(SETTLEMENT_CURRENCY_COL, summary.columns)
        self.assertNotIn(PLATFORM_CURRENCY_COL, summary.columns)
        adyen = summary[
            (summary[TRANSACTION_DATE_COL] == "2026-03")
            & (summary[ADMIN_PAYMENT_COL] == "Adyen")
            & (summary[SETTLEMENT_CURRENCY_COL] == "USD")
        ].iloc[0]
        self.assertEqual(adyen["成功笔数"], 2)
        self.assertAlmostEqual(adyen["成功金额"], 30.75, places=2)
        self.assertEqual(adyen["退款笔数"], 1)
        self.assertAlmostEqual(adyen["退款金额"], 5.00, places=2)
        self.assertAlmostEqual(adyen["净交易金额"], 25.75, places=2)
        self.assertAlmostEqual(adyen["手续费"], 3.50, places=2)
        google_hkd = summary[
            (summary[TRANSACTION_DATE_COL] == "2026-03")
            & (summary[ADMIN_PAYMENT_COL] == "Google支付")
            & (summary[SETTLEMENT_CURRENCY_COL] == "HKD")
        ].iloc[0]
        self.assertEqual(google_hkd["成功笔数"], 2)
        self.assertAlmostEqual(google_hkd["成功金额"], 10.00, places=2)
        self.assertAlmostEqual(google_hkd["手续费"], 1.00, places=2)

    def test_apple_platform_summary_includes_absolute_fee(self):
        apple_raw = pd.DataFrame(
            [
                {
                    "Settlement Date": "2026-03-31",
                    "Currency of Proceeds": "TRY",
                    "Quantity": "2",
                    "Customer Price": "36.00",
                    "Extended Partner Share": "59.00",
                },
                {
                    "Settlement Date": "2026-03-31",
                    "Currency of Proceeds": "TRY",
                    "Quantity": "-1",
                    "Customer Price": "36.00",
                    "Extended Partner Share": "-29.00",
                },
            ]
        )

        summary = build_apple_platform_summary(apple_raw)

        self.assertEqual(len(summary), 1)
        row = summary.iloc[0]
        self.assertEqual(row[TRANSACTION_DATE_COL], "2026-03")
        self.assertEqual(row[ADMIN_PAYMENT_COL], "苹果支付Lua")
        self.assertEqual(row["成功笔数"], 2)
        self.assertAlmostEqual(row["成功金额"], 59.00, places=2)
        self.assertEqual(row["退款笔数"], 1)
        self.assertAlmostEqual(row["退款金额"], 29.00, places=2)
        self.assertAlmostEqual(row["手续费"], 6.00, places=2)

    def test_write_output_includes_reconciliation_sheets(self):
        detail = pd.DataFrame(
            [
                {
                    ADMIN_AMOUNT_COL: "100.00",
                    MATCH_STATUS_COL: "是",
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "100.00",
                    "结算金额": "10.00",
                }
            ]
        )

        with tempfile.TemporaryDirectory() as tmp:
            output_path = write_output(detail, Path(tmp))
            wb = load_workbook(output_path, read_only=True)
            sheet_names = wb.sheetnames
            summary_values = [
                cell
                for row in wb[OUTPUT_SUMMARY_SHEET_3].iter_rows(values_only=True)
                for cell in row
            ]
            wb.close()

        self.assertEqual(
            sheet_names,
            [
                OUTPUT_SHEET_3,
                OUTPUT_DIFF_SHEET_3,
                OUTPUT_FAILED_SHEET_3,
                OUTPUT_SUMMARY_SHEET_3,
                OUTPUT_APPLE_SHEET_3,
            ],
        )
        self.assertNotIn("月度对比（对账差异）", summary_values)

    def test_write_output_filters_reconciliation_sheets_and_formats_headers(self):
        detail = pd.DataFrame(
            [
                {
                    ADMIN_JOIN_COL: "OK-WITHIN-POINT-8",
                    ADMIN_AMOUNT_COL: "100.00",
                    MATCH_STATUS_COL: "是",
                    PLATFORM_AMOUNT_COL: "100.80",
                    STATUS_COL: "成功",
                },
                {
                    ADMIN_JOIN_COL: "OK-WITHIN-ONE",
                    ADMIN_AMOUNT_COL: "100.00",
                    MATCH_STATUS_COL: "是",
                    PLATFORM_AMOUNT_COL: "101.00",
                    STATUS_COL: "成功",
                },
                {
                    ADMIN_JOIN_COL: "DIFF-OVER-ONE",
                    ADMIN_AMOUNT_COL: "100.00",
                    MATCH_STATUS_COL: "是",
                    PLATFORM_AMOUNT_COL: "101.01",
                    STATUS_COL: "成功",
                },
                {
                    ADMIN_JOIN_COL: "FAILED",
                    ADMIN_AMOUNT_COL: "100.00",
                    MATCH_STATUS_COL: "否",
                    PLATFORM_AMOUNT_COL: "",
                    STATUS_COL: "失败",
                },
                {
                    ADMIN_JOIN_COL: "PLATFORM-ONLY",
                    ADMIN_AMOUNT_COL: "",
                    MATCH_STATUS_COL: "平台多余",
                    PLATFORM_AMOUNT_COL: "100.00",
                    STATUS_COL: "成功",
                },
            ]
        )

        with tempfile.TemporaryDirectory() as tmp:
            detail[TRANSACTION_DATE_COL] = "2026-03-01"
            detail[ADMIN_PAYMENT_COL] = "Adyen"
            detail[PLATFORM_CURRENCY_COL] = "TRY"
            detail[SETTLEMENT_CURRENCY_COL] = "USD"
            detail["结算金额"] = detail[PLATFORM_AMOUNT_COL]

            output_path = write_output(detail, Path(tmp))
            main = pd.read_excel(output_path, sheet_name=OUTPUT_SHEET_3, dtype=str).fillna("")
            diff = pd.read_excel(output_path, sheet_name=OUTPUT_DIFF_SHEET_3, dtype=str).fillna("")
            failed = pd.read_excel(output_path, sheet_name=OUTPUT_FAILED_SHEET_3, dtype=str).fillna("")

            wb = load_workbook(output_path)
            try:
                for sheet_name in (OUTPUT_SHEET_3, OUTPUT_DIFF_SHEET_3, OUTPUT_FAILED_SHEET_3):
                    ws = wb[sheet_name]
                    self.assertEqual(ws.freeze_panes, "A2")
                    self.assertEqual(ws.auto_filter.ref, ws.dimensions)
            finally:
                wb.close()

        self.assertEqual(list(main.columns), list(diff.columns))
        self.assertEqual(list(main.columns), list(failed.columns))
        self.assertFalse(any(str(col).startswith("_google_") for col in main.columns))
        self.assertEqual(set(diff[ADMIN_JOIN_COL]), {"DIFF-OVER-ONE"})
        self.assertEqual(set(failed[ADMIN_JOIN_COL]), {"FAILED"})

    def test_write_output_keeps_apple_rows_in_result_sheet(self):
        detail = pd.DataFrame(
            [
                {
                    ADMIN_AMOUNT_COL: "100.00",
                    ADMIN_DATE_COL: "2026-03-01 10:00:00",
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "100.00",
                    "结算金额": "10.00",
                    "手续费": "-1.00",
                },
                {
                    ADMIN_AMOUNT_COL: "50.00",
                    ADMIN_DATE_COL: "2026-03-01 11:00:00",
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "苹果支付Lua",
                    PLATFORM_CURRENCY_COL: "",
                    SETTLEMENT_CURRENCY_COL: "",
                    STATUS_COL: "成功",
                    PLATFORM_AMOUNT_COL: "50.00",
                    "结算金额": "50.00",
                    "手续费": "",
                },
            ]
        )

        with tempfile.TemporaryDirectory() as tmp:
            output_path = write_output(detail, Path(tmp))
            result = pd.read_excel(output_path, sheet_name=OUTPUT_SHEET_3, dtype=str).fillna("")

        self.assertEqual(len(result), 2)
        self.assertIn("苹果支付Lua", set(result[ADMIN_PAYMENT_COL]))

    def test_monthly_comparison_uses_match_status_and_settlement_currency(self):
        summary = pd.DataFrame(
            [
                {
                    TRANSACTION_DATE_COL: "2026-03",
                    ADMIN_PAYMENT_COL: "Adyen",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    "净交易金额": 14.0,
                },
                {
                    TRANSACTION_DATE_COL: "2026-03",
                    ADMIN_PAYMENT_COL: "Adyen",
                    SETTLEMENT_CURRENCY_COL: "HKD",
                    "净交易金额": 7.0,
                },
            ]
        )
        detail = pd.DataFrame(
            [
                {
                    ADMIN_AMOUNT_COL: "100.00",
                    TRANSACTION_DATE_COL: "2026-03-01",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    MATCH_STATUS_COL: "是",
                },
                {
                    ADMIN_AMOUNT_COL: "70.00",
                    TRANSACTION_DATE_COL: "2026-03-02",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "HKD",
                    SETTLEMENT_CURRENCY_COL: "HKD",
                    MATCH_STATUS_COL: "是",
                },
                {
                    ADMIN_AMOUNT_COL: "999.00",
                    TRANSACTION_DATE_COL: "2026-03-03",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    MATCH_STATUS_COL: "否",
                },
                {
                    ADMIN_AMOUNT_COL: "",
                    TRANSACTION_DATE_COL: "2026-03-04",
                    ADMIN_PAYMENT_COL: "Adyen",
                    PLATFORM_CURRENCY_COL: "TRY",
                    SETTLEMENT_CURRENCY_COL: "USD",
                    MATCH_STATUS_COL: "平台多余",
                },
            ]
        )

        comparison = build_monthly_comparison(summary, pd.DataFrame(), detail)
        usd = comparison[
            (comparison[ADMIN_PAYMENT_COL] == "Adyen")
            & (comparison[SETTLEMENT_CURRENCY_COL] == "USD")
        ].iloc[0]
        hkd = comparison[
            (comparison[ADMIN_PAYMENT_COL] == "Adyen")
            & (comparison[SETTLEMENT_CURRENCY_COL] == "HKD")
        ].iloc[0]

        self.assertAlmostEqual(usd["Admin结算金额"], 100.00, places=2)
        self.assertAlmostEqual(usd["平台净到账"], 14.00, places=2)
        self.assertAlmostEqual(hkd["Admin结算金额"], 70.00, places=2)
        self.assertAlmostEqual(hkd["平台净到账"], 7.00, places=2)

    def test_scan_source_files_only_recognizes_prefixed_samples(self):
        sources = scan_source_files(INPUT_DIR)
        names = sorted(path.name for path in (Path(item["filepath"]) for item in sources))
        bank_names = sorted(item["bank_name"] for item in sources)

        self.assertEqual(len(names), 7)
        self.assertIn("智星-中信银行.xlsx", names)
        self.assertEqual(
            bank_names,
            ["中信银行", "中国银行", "农业银行", "工商银行", "建设银行", "招商银行", "浦发银行"],
        )

    def test_read_samples_and_extract_latest_balance_by_date(self):
        cases = {
            "中信银行": ("交易日期", "账户余额"),
            "招商银行": ("交易日", "余额"),
            "建设银行": ("交易时间", "余额"),
            "浦发银行": ("交易日期", "余额"),
            "工商银行": ("交易时间", "余额"),
            "中国银行": ("交易日期", "交易后余额"),
            "农业银行": ("交易时间", "账户余额"),
        }
        sources = {item["bank_name"]: item["filepath"] for item in scan_source_files(INPUT_DIR)}

        for bank_name, (date_col, balance_col) in cases.items():
            with self.subTest(bank_name=bank_name):
                df = read_bank_file(sources[bank_name], bank_name)
                balance_date, balance = get_last_balance(df, bank_name)

                self.assertEqual(len(df), 12)
                self.assertIn(date_col, df.columns)
                self.assertIn(balance_col, df.columns)
                self.assertEqual(balance_date, "2026-12-28")
                self.assertAlmostEqual(balance, 11481.40, places=2)

    def test_summary_template_is_required(self):
        import src.bank_integration.workbook as workbook

        self.assertFalse(hasattr(workbook, "create_summary_file"))

    def test_scan_source_files_accepts_chinese_company_prefix(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            (tmp_path / "A-中信银行.xlsx").touch()
            (tmp_path / "瑞泽商务-中信银行.xlsx").touch()
            (tmp_path / "瑞泽商务-未知银行.xlsx").touch()
            (tmp_path / "中信银行.xlsx").touch()

            sources = scan_source_files(tmp_path)
            by_name = {Path(item["filepath"]).name: item for item in sources}

            self.assertEqual(set(by_name), {"A-中信银行.xlsx", "瑞泽商务-中信银行.xlsx"})
            self.assertEqual(by_name["A-中信银行.xlsx"]["company"], "A")
            self.assertEqual(by_name["瑞泽商务-中信银行.xlsx"]["company"], "瑞泽商务")
            self.assertEqual(by_name["瑞泽商务-中信银行.xlsx"]["bank_name"], "中信银行")

    def test_update_balance_sheet_uses_existing_company_header(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "银行余额"
        ws.append(["日期", "类别", None, "合计", "A", "B"])
        ws.append(["2026-12-31", "银行存款", "中信银行", "=SUM(E2:F2)", None, None])

        update_balance_sheet(ws, "中信银行", "A", "2026-12-28", 123.45)

        self.assertEqual(ws.cell(row=2, column=5).value, 123.45)
        self.assertEqual(ws.cell(row=1, column=7).value, None)
        self.assertEqual(ws.cell(row=2, column=4).value, "=SUM(E2:F2)")

    def test_update_balance_sheet_appends_missing_company_header_and_formula(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "银行余额"
        ws.append(["日期", "类别", None, "合计", "A", "B"])
        ws.append(["2026-12-31", "银行存款", "中信银行", "=SUM(E2:F2)", None, None])
        ws.append([None, None, "招商银行", "=SUM(E3:F3)", None, None])

        update_balance_sheet(ws, "中信银行", "瑞泽商务", "2026-12-28", 456.78)

        self.assertEqual(ws.cell(row=1, column=7).value, "瑞泽商务")
        self.assertEqual(ws.cell(row=2, column=7).value, 456.78)
        self.assertEqual(ws.cell(row=2, column=4).value, "=SUM(E2:G2)")
        self.assertEqual(ws.cell(row=3, column=4).value, "=SUM(E3:G3)")

    def test_boc_date_range_fills_months_before_first_transaction(self):
        df = pd.DataFrame(
            [
                {"交易日期": "20260320", "交易金额": "+0.81", "交易后余额": "6,475.08"},
                {"交易日期": "20260326", "交易金额": "-400.00", "交易后余额": "6,075.08"},
            ]
        )
        df.attrs["statement_date_range"] = (date(2026, 1, 1), date(2026, 3, 31))

        monthly = get_monthly_balances(df, "中国银行")

        self.assertEqual(monthly[0][0], "2026-01-31")
        self.assertAlmostEqual(monthly[0][1], 6474.27, places=2)
        self.assertEqual(monthly[1][0], "2026-02-28")
        self.assertAlmostEqual(monthly[1][1], 6474.27, places=2)
        self.assertEqual(monthly[2][0], "2026-03-26")
        self.assertAlmostEqual(monthly[2][1], 6075.08, places=2)

    def test_cmb_debit_first_transaction_backfills_prior_months(self):
        df = pd.DataFrame(
            [
                {
                    "交易日": "2026-03-03",
                    "借方金额": "100.00",
                    "贷方金额": "",
                    "余额": "900.00",
                }
            ]
        )
        df.attrs["statement_date_range"] = (date(2026, 1, 1), date(2026, 3, 31))

        monthly = get_monthly_balances(df, "招商银行")

        self.assertEqual(monthly, [("2026-01-31", 1000.0), ("2026-02-28", 1000.0), ("2026-03-03", 900.0)])

    def test_citic_credit_first_transaction_backfills_prior_months(self):
        df = pd.DataFrame(
            [
                {
                    "交易日期": "2026-04-10",
                    "借方发生额": "",
                    "贷方发生额": "50.00",
                    "账户余额": "1,050.00",
                }
            ]
        )
        df.attrs["statement_date_range"] = (date(2026, 2, 1), date(2026, 4, 30))

        monthly = get_monthly_balances(df, "中信银行")

        self.assertEqual(monthly, [("2026-02-28", 1000.0), ("2026-03-31", 1000.0), ("2026-04-10", 1050.0)])

    def test_missing_middle_month_uses_previous_transaction_balance(self):
        df = pd.DataFrame(
            [
                {"交易日": "2026-01-15", "借方金额": "", "贷方金额": "100.00", "余额": "1,000.00"},
                {"交易日": "2026-03-10", "借方金额": "", "贷方金额": "200.00", "余额": "1,200.00"},
            ]
        )
        df.attrs["statement_date_range"] = (date(2026, 1, 1), date(2026, 3, 31))

        monthly = get_monthly_balances(df, "招商银行")

        self.assertEqual(monthly, [("2026-01-15", 1000.0), ("2026-02-28", 1000.0), ("2026-03-10", 1200.0)])

    def test_missing_month_fill_requires_mode1_default_maps(self):
        df = pd.DataFrame(
            [{"交易日": "2026-03-03", "借方金额": "100.00", "贷方金额": "", "余额": "900.00"}]
        )
        df.attrs["statement_date_range"] = (date(2026, 1, 1), date(2026, 3, 31))

        monthly = get_monthly_balances(
            df,
            "招商银行",
            balance_col_map=BANK_BALANCE_COL_2,
            date_col_map=BANK_DATE_COL_2,
        )

        self.assertEqual(monthly, [("2026-03-03", 900.0)])

    def test_mode2_east_asia_has_no_balance_column(self):
        cases = [
            "B-东亚银行-HKD.csv",
            "B-东亚银行-USD.csv",
        ]

        for filename in cases:
            with self.subTest(filename=filename):
                df = read_bank_file(
                    str(INPUT_DIR_2 / filename),
                    "东亚银行",
                    bank_read_config=BANK_READ_CONFIG_2,
                    bank_date_col=BANK_DATE_COL_2,
                )
                monthly = get_monthly_balances(
                    df,
                    "东亚银行",
                    balance_col_map=BANK_BALANCE_COL_2,
                    date_col_map=BANK_DATE_COL_2,
                )

                self.assertFalse(df.empty)
                self.assertEqual(monthly, [])

    def test_mode2_ocbc_extracts_monthly_balance(self):
        cases = {
            "D-华侨银行-HKD.csv": ("2026-04-13", 1716.58),
            "D-华侨银行-USD.csv": ("2026-04-24", 96049.90),
        }

        for filename, (expected_date, expected_balance) in cases.items():
            with self.subTest(filename=filename):
                df = read_bank_file(
                    str(INPUT_DIR_2 / filename),
                    "华侨银行",
                    bank_read_config=BANK_READ_CONFIG_2,
                    bank_date_col=BANK_DATE_COL_2,
                )
                monthly = get_monthly_balances(
                    df,
                    "华侨银行",
                    balance_col_map=BANK_BALANCE_COL_2,
                    date_col_map=BANK_DATE_COL_2,
                )

                self.assertIn("余额", df.columns)
                self.assertEqual(list(df.columns).count("交易日期"), 1)
                self.assertEqual(monthly[0][0], expected_date)
                self.assertAlmostEqual(monthly[0][1], expected_balance, places=2)

    def test_mode2_scanner_accepts_only_huamei_pdf(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            (tmp_path / "A-华美银行-USD.pdf").touch()
            (tmp_path / "瑞泽商务-华侨银行-HKD.csv").touch()
            (tmp_path / "B-汇丰银行-USD.pdf").touch()
            (tmp_path / "C-华侨银行-HKD.csv").touch()
            (tmp_path / "瑞泽商务-未知银行-HKD.csv").touch()

            sources = scan_source_files_2(tmp_path)
            by_name = {Path(item["filepath"]).name: item for item in sources}

            self.assertEqual(
                set(by_name),
                {"A-华美银行-USD.pdf", "C-华侨银行-HKD.csv", "瑞泽商务-华侨银行-HKD.csv"},
            )
            self.assertEqual(by_name["A-华美银行-USD.pdf"]["company"], "A")
            self.assertEqual(by_name["瑞泽商务-华侨银行-HKD.csv"]["company"], "瑞泽商务")
            self.assertEqual(by_name["瑞泽商务-华侨银行-HKD.csv"]["bank_name"], "华侨银行")
            self.assertEqual(by_name["瑞泽商务-华侨银行-HKD.csv"]["currency"], "HKD")

    def test_update_balance_sheet_2_uses_existing_company_header(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "MIG银行余额（20261231）"
        ws.append(["日期", None, None, "折合人民币", "合计", "币种", "A", "B"])
        ws.append(["2026-04-30", None, "华侨银行", None, "=SUM(G2:H2)", "HKD", None, None])
        ws.append([None, None, None, None, "=SUM(G3:H3)", "USD", None, None])

        update_balance_sheet_2(ws, "华侨银行", "HKD", "A", "2026-04-13", 123.45)

        self.assertEqual(ws.cell(row=2, column=7).value, 123.45)
        self.assertEqual(ws.cell(row=1, column=9).value, None)
        self.assertEqual(ws.cell(row=2, column=5).value, "=SUM(G2:H2)")
        self.assertIn("MATCH(F2", ws.cell(row=2, column=4).value)

    def test_update_balance_sheet_2_appends_missing_company_header_and_formula(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "MIG银行余额（20261231）"
        ws.append(["日期", None, None, "折合人民币", "合计", "币种", "A", "B"])
        ws.append(["2026-04-30", None, "华侨银行", None, "=SUM(G2:H2)", "HKD", None, None])
        ws.append([None, None, None, None, "=SUM(G3:H3)", "USD", None, None])

        update_balance_sheet_2(ws, "华侨银行", "USD", "瑞泽商务", "2026-04-24", 456.78)

        self.assertEqual(ws.cell(row=1, column=5).value, "合计")
        self.assertEqual(ws.cell(row=1, column=9).value, "瑞泽商务")
        self.assertEqual(ws.cell(row=3, column=9).value, 456.78)
        self.assertEqual(ws.cell(row=2, column=5).value, "=SUM(G2:I2)")
        self.assertEqual(ws.cell(row=3, column=5).value, "=SUM(G3:I3)")
        self.assertIn("MATCH(F3", ws.cell(row=3, column=4).value)

    @unittest.skipIf(not HUAMEI_PDF.exists(), "华美银行 PDF 测试文件不存在")
    def test_mode2_huamei_pdf_extracts_statement_month_last_daily_balance(self):
        df = read_bank_file(
            str(HUAMEI_PDF),
            "华美银行",
            bank_read_config=BANK_READ_CONFIG_2,
            bank_date_col=BANK_DATE_COL_2,
        )
        monthly = get_monthly_balances(
            df,
            "华美银行",
            balance_col_map=BANK_BALANCE_COL_2,
            date_col_map=BANK_DATE_COL_2,
        )

        self.assertEqual(list(df.columns), ["Date", "Amount"])
        self.assertEqual(len(df), 1)
        self.assertEqual(df.iloc[0]["Date"], "2025-02-04")
        self.assertAlmostEqual(df.iloc[0]["Amount"], 3297.03, places=2)
        self.assertEqual(monthly[0][0], "2025-02-04")
        self.assertAlmostEqual(monthly[0][1], 3297.03, places=2)

    def test_mode4_parse_date_accepts_strict_iso_date(self):
        self.assertEqual(parse_date("2025-05-12"), date(2025, 5, 12))

    def test_mode4_parse_date_rejects_invalid_formats_and_values(self):
        for value in ("2025/05/12", "20250512", "2025-02-30", "2025-5-12"):
            with self.subTest(value=value):
                with self.assertRaises(ValueError):
                    parse_date(value)

    def test_mode4_previous_month_range_uses_previous_natural_month(self):
        self.assertEqual(get_previous_month_range(date(2026, 5, 13)), (date(2026, 4, 1), date(2026, 4, 30)))

    def test_mode4_previous_month_range_handles_year_boundary(self):
        self.assertEqual(get_previous_month_range(date(2026, 1, 10)), (date(2025, 12, 1), date(2025, 12, 31)))

    def test_mode4_parse_date_args_defaults_to_previous_month(self):
        self.assertEqual(
            parse_date_args([], today=date(2026, 5, 13)),
            (date(2026, 4, 1), date(2026, 4, 30), EXPORT_BATCH_WAIT_SECONDS_4),
        )

    def test_mode4_parse_date_args_accepts_manual_date_range(self):
        args = ["--date-range", "2026-04-01", "2026-04-30"]

        self.assertEqual(parse_date_args(args), (date(2026, 4, 1), date(2026, 4, 30), EXPORT_BATCH_WAIT_SECONDS_4))

    def test_mode4_parse_date_args_accepts_custom_wait_seconds(self):
        self.assertEqual(
            parse_date_args(["--wait-seconds", "60"], today=date(2026, 5, 13)),
            (date(2026, 4, 1), date(2026, 4, 30), 60),
        )

    def test_mode4_parse_date_args_accepts_date_range_with_custom_wait_seconds(self):
        args = ["--date-range", "2026-04-01", "2026-04-30", "--wait-seconds", "45"]

        self.assertEqual(parse_date_args(args), (date(2026, 4, 1), date(2026, 4, 30), 45))

    def test_mode4_parse_date_args_rejects_bad_arguments(self):
        bad_args = [
            ["--date-range"],
            ["--date-range", "2026-04-01"],
            ["--date-range", "2026/04/01", "2026-04-30"],
            ["--date-range", "2026-05-01", "2026-04-30"],
            ["--start", "2026-04-01", "2026-04-30"],
            ["--wait-seconds"],
            ["--wait-seconds", "0"],
            ["--wait-seconds", "-1"],
            ["--wait-seconds", "abc"],
        ]
        for args in bad_args:
            with self.subTest(args=args):
                with self.assertRaises(ValueError):
                    parse_date_args(args)

    def test_mode4_build_urls_for_each_day_and_preserves_page_placeholder(self):
        urls = build_export_urls(date(2025, 5, 12), date(2025, 5, 14))

        self.assertEqual(len(urls), 3)
        self.assertIn("pay_sdate=2025-05-12", urls[0])
        self.assertIn("pay_edate=2025-05-12", urls[0])
        self.assertIn("pay_sdate=2025-05-13", urls[1])
        self.assertIn("pay_edate=2025-05-13", urls[1])
        self.assertIn("pay_sdate=2025-05-14", urls[2])
        self.assertIn("pay_edate=2025-05-14", urls[2])
        self.assertTrue(all("p=[PAGE]" in url for url in urls))

    def test_mode4_chunks_month_into_five_day_batches(self):
        days = list(range(31))

        batches = chunk_list(days, 5)

        self.assertEqual(len(batches), 7)
        self.assertEqual([len(batch) for batch in batches], [5, 5, 5, 5, 5, 5, 1])

    def test_mode4_build_urls_rejects_reversed_date_range(self):
        with self.assertRaises(ValueError):
            build_export_urls(date(2025, 5, 14), date(2025, 5, 12))

    def test_mode4_build_single_export_url_keeps_page_placeholder(self):
        url = build_export_url(date(2025, 5, 12))

        self.assertIn("pay_sdate=2025-05-12", url)
        self.assertIn("pay_edate=2025-05-12", url)
        self.assertIn("p=[PAGE]", url)

    def test_mode4_configure_chrome_downloads_writes_preferences(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            profile_dir = root / "profile"
            download_dir = root / "downloads"

            prefs_path = configure_chrome_downloads(profile_dir, download_dir)

            self.assertTrue(download_dir.exists())
            self.assertTrue(prefs_path.exists())
            prefs = json.loads(prefs_path.read_text(encoding="utf-8"))
            self.assertEqual(prefs["download"]["default_directory"], str(download_dir.resolve()))
            self.assertFalse(prefs["download"]["prompt_for_download"])
            self.assertTrue(prefs["download"]["directory_upgrade"])

    def test_mode4_expected_export_file_matches_strict_date_filename(self):
        with tempfile.TemporaryDirectory() as tmp:
            download_dir = Path(tmp)
            target = date(2026, 5, 1)
            (download_dir / "2026-05-01,2026-05-01.xlsx").write_text("xlsx", encoding="utf-8")
            (download_dir / "2026-05-02,2026-05-02.xls").write_text("xls", encoding="utf-8")
            (download_dir / "2026-05-03,2026-05-03.csv").write_text("csv", encoding="utf-8")
            (download_dir / "2026-05-04,2026-05-04.xlsx.crdownload").write_text("tmp", encoding="utf-8")
            (download_dir / "temp.tmp").write_text("tmp", encoding="utf-8")
            (download_dir / "2026-05-05,2026-05-05.txt").write_text("txt", encoding="utf-8")
            (download_dir / "prefix-2026-05-06,2026-05-06.xlsx").write_text("bad", encoding="utf-8")
            (download_dir / "2026-05-07,2026-05-07 (1).xlsx").write_text("bad", encoding="utf-8")

            self.assertEqual(expected_export_stem(target), "2026-05-01,2026-05-01")
            missing = missing_export_dates(
                download_dir,
                [date(2026, 5, day) for day in range(1, 8)],
                (".xls", ".xlsx", ".csv"),
            )

            self.assertEqual(missing, [date(2026, 5, 4), date(2026, 5, 5), date(2026, 5, 6), date(2026, 5, 7)])

    def test_mode4_export_batch_success_does_not_retry(self):
        dated_urls = [(date(2025, 5, day), f"https://example.test/{day}") for day in range(1, 4)]
        launched = []

        def fake_launcher(_chrome_path, _profile_dir, urls):
            launched.append(list(urls))

        def fake_waiter(_download_dir, _days, _wait_seconds, _suffixes):
            return []

        failures = export_batches_with_retries(
            "chrome",
            Path("profile"),
            Path("downloads"),
            dated_urls,
            batch_size=5,
            wait_seconds=20,
            retry_limit=2,
            launcher=fake_launcher,
            waiter=fake_waiter,
        )

        self.assertEqual(failures, [])
        self.assertEqual(len(launched), 1)
        self.assertEqual(len(launched[0]), 3)

    def test_mode4_export_batch_retries_only_missing_dates(self):
        dated_urls = [(date(2025, 5, day), f"https://example.test/{day}") for day in range(1, 4)]
        launched = []
        wait_calls = []

        def fake_launcher(_chrome_path, _profile_dir, urls):
            launched.append(list(urls))

        def fake_waiter(_download_dir, days, _wait_seconds, _suffixes):
            wait_calls.append(list(days))
            if len(wait_calls) == 1:
                return [date(2025, 5, 2)]
            return []

        failures = export_batches_with_retries(
            "chrome",
            Path("profile"),
            Path("downloads"),
            dated_urls,
            batch_size=5,
            wait_seconds=20,
            retry_limit=2,
            launcher=fake_launcher,
            waiter=fake_waiter,
        )

        self.assertEqual(failures, [])
        self.assertEqual(launched, [["https://example.test/1", "https://example.test/2", "https://example.test/3"], ["https://example.test/2"]])
        self.assertEqual(wait_calls, [[date(2025, 5, 1), date(2025, 5, 2), date(2025, 5, 3)], [date(2025, 5, 2)]])

    def test_mode4_export_batch_records_missing_dates_after_retry_limit(self):
        dated_urls = [(date(2025, 5, day), f"https://example.test/{day}") for day in range(1, 4)]
        launched = []

        def fake_launcher(_chrome_path, _profile_dir, urls):
            launched.append(list(urls))

        def fake_waiter(_download_dir, days, _wait_seconds, _suffixes):
            return list(days)

        failures = export_batches_with_retries(
            "chrome",
            Path("profile"),
            Path("downloads"),
            dated_urls,
            batch_size=5,
            wait_seconds=20,
            retry_limit=2,
            launcher=fake_launcher,
            waiter=fake_waiter,
        )

        self.assertEqual(len(launched), 3)
        self.assertEqual(failures, [date(2025, 5, 1), date(2025, 5, 2), date(2025, 5, 3)])

    def test_mode4_build_chrome_args_uses_default_profile_and_preserves_url(self):
        with tempfile.TemporaryDirectory() as tmp:
            profile_dir = Path(tmp) / "profile"
            url = build_export_url(date(2025, 5, 12))

            args = build_chrome_args("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome", profile_dir, [url])

            self.assertEqual(args[0], "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")
            self.assertIn(f"--user-data-dir={profile_dir.resolve()}", args)
            self.assertIn("--profile-directory=Default", args)
            self.assertIn("--remote-debugging-port=0", args)
            self.assertIn("--class=bank-integration-export", args)
            self.assertIn(url, args)
            self.assertIn("p=[PAGE]", url)

    def test_mode4_cookie_store_requires_default_cookies_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            profile_dir = Path(tmp) / "profile"

            self.assertFalse(has_chrome_cookie_store(profile_dir))

            network_cookie = profile_dir / "Default" / "Network" / "Cookies"
            network_cookie.parent.mkdir(parents=True, exist_ok=True)
            network_cookie.write_text("network cookie data", encoding="utf-8")

            self.assertFalse(has_chrome_cookie_store(profile_dir))

            primary_cookie = profile_dir / "Default" / "Cookies"
            primary_cookie.write_text("primary cookie data", encoding="utf-8")

            self.assertTrue(has_chrome_cookie_store(profile_dir))


if __name__ == "__main__":
    unittest.main()
