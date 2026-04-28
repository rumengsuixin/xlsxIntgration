"""
银行流水数据整合脚本

使用方法：
  1. 将各银行源文件重命名为 {公司代号}-{银行全称}.{扩展名}
     例如：B-建设银行.xls、A-招商银行.xlsx、C-工商银行.xlsx
  2. 将重命名后的文件放入本脚本同目录
  3. 运行：python 整合.py

功能：
  - 将源文件的交易数据原样写入 国内银行汇总.xlsx 对应的公司-银行子表
  - 提取最后一笔交易的期末余额，更新"银行余额"工作表对应月份的单元格
"""

import os
import re
import logging
from datetime import date, datetime
from typing import Optional, List, Dict, Tuple

import pandas as pd
import openpyxl
from openpyxl.utils.datetime import from_excel as excel_to_date

# ─────────────────────────── 常量 ───────────────────────────────

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
SUMMARY_FILE = "国内银行汇总.xlsx"

BANK_ABBR = {
    "招商银行": "招行",
    "建设银行": "建行",
    "工商银行": "工行",
    "中信银行": "中信",
    "浦发银行": "浦发",
    "农业银行": "农行",
    "中国银行": "中行",
}

# 各银行读取配置：header 行（0-indexed），文件引擎
BANK_READ_CONFIG = {
    "工商银行": {"header": 1,  "engine": "openpyxl", "is_csv": False},
    "中信银行": {"header": 15, "engine": "openpyxl", "is_csv": False},
    "招商银行": {"header": 12, "engine": "openpyxl", "is_csv": False},
    "农业银行": {"header": 2,  "engine": "xlrd",     "is_csv": False},
    "建设银行": {"header": 9,  "engine": "xlrd",     "is_csv": False},
    "浦发银行": {"header": 4,  "engine": "xlrd",     "is_csv": False},
    "中国银行": {"header": 7,  "engine": None,        "is_csv": True},
}

# 各银行余额列名（源文件原始列名，清洗后）
BANK_BALANCE_COL = {
    "工商银行": "余额",
    "中信银行": "账户余额",
    "招商银行": "余额",
    "农业银行": "账户余额",
    "建设银行": "余额",
    "浦发银行": "余额",
    "中国银行": "交易后余额",
}

# 各银行日期列名（源文件原始列名，清洗后）
BANK_DATE_COL = {
    "工商银行": "交易时间",
    "中信银行": "交易日期",
    "招商银行": "交易日",
    "农业银行": "交易时间",
    "建设银行": "交易时间",
    "浦发银行": "交易日期",
    "中国银行": "交易日期",
}

BALANCE_SHEET = "银行余额"
PROTECTED_SHEETS = {"Sheet9", BALANCE_SHEET}


# ─────────────────────────── 读取源文件 ─────────────────────────

def read_bank_file(filepath: str, bank_name: str) -> pd.DataFrame:
    """读取银行源文件，从表头行开始，返回清洗后的 DataFrame（dtype 全为 str）。"""
    cfg = BANK_READ_CONFIG[bank_name]
    header = cfg["header"]

    if cfg["is_csv"]:
        df = None
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                df = pd.read_csv(filepath, header=header, dtype=str, encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        if df is None:
            raise ValueError(f"无法解码 CSV 文件（尝试 utf-8-sig/gbk/utf-8）: {filepath}")
        # 中国银行列名含双语，如 "交易类型[ Transaction Type ]"，只取中文部分
        df.columns = [c.split('[')[0].strip() for c in df.columns]
    else:
        df = pd.read_excel(
            filepath,
            header=header,
            dtype=str,
            engine=cfg["engine"],
        )

    # 去除全空行
    df = df.dropna(how='all')
    # 列名去除首尾空格（建设银行等存在此问题）
    df.columns = [str(c).strip() for c in df.columns]
    # NaN 统一填充为空字符串
    df = df.fillna('')

    # 农业银行：末尾有统计汇总行，交易时间不是日期格式，过滤掉
    if bank_name == "农业银行":
        date_col = BANK_DATE_COL[bank_name]
        if date_col in df.columns:
            df = df[df[date_col].str.match(r'\d{4}-\d{2}-\d{2}')]

    return df


# ─────────────────────────── 提取期末余额 ───────────────────────

def get_last_balance(df: pd.DataFrame, bank_name: str):
    """
    从 DataFrame 末尾往前找最后一个余额非空非零的行。
    返回 (date_str "YYYY-MM-DD", balance float)，找不到则返回 (None, None)。
    """
    balance_col = BANK_BALANCE_COL.get(bank_name, "")
    date_col = BANK_DATE_COL.get(bank_name, "")

    if balance_col not in df.columns:
        logging.warning(f"  余额列 '{balance_col}' 不存在，跳过余额更新")
        return None, None

    for _, row in df.iloc[::-1].iterrows():
        bal_str = str(row.get(balance_col, '')).strip().replace(',', '').replace('+', '')
        if not bal_str:
            continue
        try:
            balance = float(bal_str)
        except ValueError:
            continue
        if balance == 0.0:
            continue

        # 提取日期字符串（兼容 YYYY-MM-DD、YYYY/MM/DD、YYYYMMDD、含时间的格式）
        date_raw = str(row.get(date_col, '')).strip()
        m = re.search(r'(\d{4})[-/]?(\d{2})[-/]?(\d{2})', date_raw)
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}", balance

    return None, None


# ─────────────────────────── 单元格日期解析 ─────────────────────

def parse_cell_date(val) -> Optional[date]:
    """将 openpyxl 单元格值解析为 Python date 对象（兼容 datetime、date、Excel 序列数、字符串）。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)) and val > 0:
        try:
            return excel_to_date(int(val)).date()
        except Exception:
            pass
    if isinstance(val, str):
        m = re.match(r'(\d{4})-(\d{2})-(\d{2})', val.strip())
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


# ─────────────────────────── 更新银行余额表 ─────────────────────

def _find_or_create_date_block(ws, year: int, month: int) -> int:
    """
    在银行余额表中查找目标年月的数据块起始行。
    找不到时，在表末追加新块（9行：8家银行 + 货币资金合计），返回其起始行号。
    """
    import calendar
    from openpyxl.utils import get_column_letter

    # 扫描 A 列查找目标年月
    block_start = None
    for row_idx in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row_idx, column=1).value
        if a_val is not None:
            cell_date = parse_cell_date(a_val)
            if cell_date is not None and cell_date.year == year and cell_date.month == month:
                block_start = row_idx
                break

    if block_start is not None:
        return block_start

    # 未找到：在末尾追加新月份块
    # 空一行后开始（若最后有内容则再空一行）
    start_row = ws.max_row + 2

    # 月末日期（与现有格式一致）
    last_day = calendar.monthrange(year, month)[1]
    month_end = date(year, month, last_day)

    company_codes = [chr(c) for c in range(ord('A'), ord('V') + 1)]  # A-V

    for i, bk in enumerate(BALANCE_BANK_ORDER):
        row = start_row + i
        if i == 0:
            # 第一行写月末日期和"银行存款"类别
            ws.cell(row=row, column=1).value = month_end
            ws.cell(row=row, column=2).value = "银行存款"
        ws.cell(row=row, column=3).value = bk
        # D 列写合计公式（E..Z 列公司余额之和）
        e_col = get_column_letter(5)
        z_col = get_column_letter(4 + len(company_codes))
        ws.cell(row=row, column=4).value = f"=SUM({e_col}{row}:{z_col}{row})"

    # 第 9 行：货币资金合计
    summary_row = start_row + len(BALANCE_BANK_ORDER)
    ws.cell(row=summary_row, column=2).value = "货币资金合计"
    d_start = get_column_letter(4)
    ws.cell(row=summary_row, column=4).value = (
        f"=SUM({d_start}{start_row}:{d_start}{start_row + len(BALANCE_BANK_ORDER) - 1})"
    )

    logging.info(f"  银行余额表新增 {year}年{month}月 数据块（行 {start_row}-{summary_row}）")
    return start_row


def update_balance_sheet(
    ws,
    bank_name: str,
    company_code: str,
    date_str: str,
    balance: float,
) -> None:
    """在"银行余额"工作表中找到对应月份、银行行，更新指定公司列的余额。
    若目标月份块不存在，自动在表末追加新块后再更新。"""
    m = re.match(r'(\d{4})-(\d{2})', date_str)
    if not m:
        logging.warning(f"  日期格式异常 '{date_str}'，跳过余额更新")
        return

    year, month = int(m.group(1)), int(m.group(2))

    # 公司列号（1-based）：A→5(E列)，B→6(F列)，…，V→26(Z列)
    col_idx = ord(company_code) - ord('A') + 5

    # 确保目标月份块存在
    block_start = _find_or_create_date_block(ws, year, month)

    # 在块内 C 列找目标银行行
    target_row = None
    for row_idx in range(block_start, block_start + len(BALANCE_BANK_ORDER)):
        c_val = str(ws.cell(row=row_idx, column=3).value or '').strip()
        if c_val == bank_name:
            target_row = row_idx
            break

    if target_row is None:
        logging.warning(f"  银行余额块中未找到银行行 [{bank_name}]，跳过")
        return

    ws.cell(row=target_row, column=col_idx).value = balance
    logging.info(
        f"  银行余额已更新: {year}年{month}月 / {bank_name} / 公司{company_code} = {balance}"
    )


# ─────────────────────────── 创建汇总文件 ──────────────────────

# 银行在余额表中的固定顺序
BALANCE_BANK_ORDER = [
    "招商银行", "浦发银行", "中国银行", "工商银行",
    "建设银行", "农业银行", "兴业银行", "中信银行",
]

def create_summary_file(summary_path: str) -> None:
    """
    当汇总文件不存在时，创建含基础结构的空白汇总文件：
    - Sheet9：公司代号映射表（A-V）
    - 银行余额：含列头行（日期 | 类别 | 银行名 | 合计 | A…V）
    """
    wb = openpyxl.Workbook()

    # ── Sheet9：公司代号映射 ──────────────────────────────────────
    ws9 = wb.active
    ws9.title = "Sheet9"
    ws9.append(["编号", "公司名称"])
    for code in [chr(c) for c in range(ord('A'), ord('V') + 1)]:
        ws9.append([code, ""])

    # ── 银行余额：列头行 ─────────────────────────────────────────
    ws_bal = wb.create_sheet(BALANCE_SHEET)
    company_codes = [chr(c) for c in range(ord('A'), ord('V') + 1)]  # A-V
    ws_bal.append(["日期", "类别", "", "合计"] + company_codes)

    wb.save(summary_path)
    logging.info(f"汇总文件不存在，已创建空白模板: {summary_path}")


# ─────────────────────────── 扫描源文件目录 ─────────────────────

def scan_source_files(work_dir: str) -> List[Dict]:
    """
    扫描目录，匹配命名格式为 {公司代号}-{银行全称}.{扩展名} 的源文件。
    返回列表，每项包含 company、bank_name、filepath。
    """
    pattern = re.compile(r'^([A-Z])-(.+)\.(xls|xlsx|csv)$')
    results = []

    for fname in sorted(os.listdir(work_dir)):
        if fname.startswith('~$') or fname == SUMMARY_FILE:
            continue
        m = pattern.match(fname)
        if not m:
            continue
        company = m.group(1)
        bank_name = m.group(2)
        if bank_name not in BANK_ABBR:
            logging.warning(f"未知银行名称 '{bank_name}'，跳过文件: {fname}")
            continue
        results.append({
            "company": company,
            "bank_name": bank_name,
            "filepath": os.path.join(work_dir, fname),
        })

    return results


# ─────────────────────────── 批量写入汇总文件 ───────────────────

def write_all_to_summary(results: List[Dict], summary_path: str) -> None:
    """
    一次性打开汇总文件，写入所有明细子表并更新银行余额表，最后保存。
    """
    try:
        wb = openpyxl.load_workbook(summary_path)
    except PermissionError:
        logging.error(f"无法打开汇总文件，请先关闭 Excel 中的 {SUMMARY_FILE}")
        raise

    # 写入各明细子表
    for item in results:
        sheet_name = item["sheet_name"]
        df = item["df"]

        if sheet_name in PROTECTED_SHEETS:
            logging.warning(f"  [{sheet_name}] 是受保护表，跳过写入")
            continue

        # 删除旧表后重建，确保数据干净
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        # 第一行写原始列名
        ws.append(list(df.columns))
        # 逐行写原始数据（空字符串转为 None，单元格更干净）
        for row_tuple in df.itertuples(index=False, name=None):
            ws.append([v if v != '' else None for v in row_tuple])

        logging.info(f"  [{sheet_name}] 写入 {len(df)} 行数据")

    # 更新银行余额表
    if BALANCE_SHEET in wb.sheetnames:
        ws_bal = wb[BALANCE_SHEET]
        for item in results:
            if item.get("balance") is not None and item.get("balance_date"):
                update_balance_sheet(
                    ws_bal,
                    item["bank_name"],
                    item["company_code"],
                    item["balance_date"],
                    item["balance"],
                )
    else:
        logging.warning(f"汇总文件中未找到 '{BALANCE_SHEET}' 工作表，跳过余额更新")

    try:
        wb.save(summary_path)
        logging.info(f"汇总文件保存完成: {summary_path}")
    except PermissionError:
        logging.error(f"无法保存汇总文件，请先关闭 Excel 中的 {SUMMARY_FILE}")
        raise


# ─────────────────────────── 主流程 ────────────────────────────

def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
    )

    work_dir = WORK_DIR
    summary_path = os.path.join(work_dir, SUMMARY_FILE)

    if not os.path.exists(summary_path):
        create_summary_file(summary_path)

    sources = scan_source_files(work_dir)
    if not sources:
        logging.warning(
            "未找到符合命名规范的源文件。"
            "请将文件重命名为 {公司代号}-{银行全称}.{扩展名}，例如：B-建设银行.xls"
        )
        return

    logging.info(f"共找到 {len(sources)} 个源文件待处理")
    results = []

    for item in sources:
        company = item["company"]
        bank_name = item["bank_name"]
        filepath = item["filepath"]
        sheet_name = f"{company}-{BANK_ABBR[bank_name]}"

        logging.info(f"处理: {os.path.basename(filepath)} → [{sheet_name}]")
        try:
            df = read_bank_file(filepath, bank_name)

            if df.empty:
                logging.warning(f"  [{sheet_name}] 无有效数据行，仅写入列头")
            else:
                logging.info(f"  [{sheet_name}] 共 {len(df)} 条记录")

            balance_date, balance = get_last_balance(df, bank_name)
            if balance is not None:
                logging.info(f"  期末余额: {balance}（日期: {balance_date}）")
            else:
                logging.info(f"  未提取到期末余额，跳过余额更新")

            results.append({
                "sheet_name": sheet_name,
                "df": df,
                "bank_name": bank_name,
                "company_code": company,
                "balance": balance,
                "balance_date": balance_date,
            })
        except Exception as e:
            logging.error(f"  处理失败: {e}", exc_info=True)
            continue

    if not results:
        logging.warning("没有可写入的数据，退出")
        return

    write_all_to_summary(results, summary_path)
    logging.info("全部处理完成")


if __name__ == "__main__":
    main()
